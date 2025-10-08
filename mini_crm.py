#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════════════════════
  ENTERPRISE CRM/ERP SYSTEM — AI-READY CORE
  Version: 2.0.0 (Enhanced & Hardened)
  
  Монолитная архитектура для максимальной стабильности
  Подготовлено для интеграции с AI-агентами
═══════════════════════════════════════════════════════════════════════════════
"""

# ===== BLOCK: IMPORTS =====

import os
import sys
import re
import json
import uuid
import secrets
import hashlib
import hmac
import time
import threading
import sqlite3
import base64
import mimetypes
import urllib.parse
import urllib.request
import socket
import ipaddress
import imaplib
import smtplib
import email
import logging
from datetime import datetime, timedelta, timezone
from functools import wraps
from typing import Any, Literal
from collections import defaultdict
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from queue import Queue, Empty

# Third-party imports
try:
    from flask import (
        Flask, g, request, session, redirect, url_for, 
        render_template_string, jsonify, send_file, Response, stream_with_context
    )
    from werkzeug.security import generate_password_hash, check_password_hash
    from werkzeug.utils import secure_filename
    import requests as _rq
    import bcrypt  # ✅ НОВОЕ: для 2FA backup codes
    import bleach  # ✅ НОВОЕ: для HTML sanitization
    from PIL import Image  # ✅ НОВОЕ: для avatar resizing
except ImportError as e:
    print(f"❌ КРИТИЧНАЯ ОШИБКА: Отсутствует зависимость: {e}")
    print("Установите: pip install flask werkzeug requests bcrypt bleach pillow")
    sys.exit(1)

# Optional dependencies (graceful degradation)
try:
    import redis
    REDIS_AVAILABLE = True
except ImportError:
    REDIS_AVAILABLE = False
    print("⚠️  Redis не установлен — rate limiting будет in-memory (не масштабируется)")

try:
    import pyotp
    PYOTP_AVAILABLE = True
except ImportError:
    PYOTP_AVAILABLE = False
    print("⚠️  pyotp не установлен — 2FA отключена")


# ===== BLOCK: CONFIGURATION =====

# Environment detection
ENV = os.getenv("ENV", "development")
DEBUG = ENV == "development"

# Server config
HOST = os.getenv("HOST", "0.0.0.0")
PORT = int(os.getenv("PORT", "5000"))

# Security
SECRET_KEY = os.getenv("SECRET_KEY", secrets.token_hex(32))
if SECRET_KEY == secrets.token_hex(32) and not DEBUG:
    print("⚠️  ВНИМАНИЕ: SECRET_KEY не установлен в production!")

# Database
DATABASE_PATH = os.getenv("DATABASE_PATH", "./crm.db")
DB_TIMEOUT = float(os.getenv("DB_TIMEOUT", "30.0"))

# Storage
STORAGE_BACKEND = os.getenv("STORAGE_BACKEND", "local")  # local | s3
LOCAL_UPLOAD_DIR = os.getenv("LOCAL_UPLOAD_DIR", "./uploads")
S3_BUCKET = os.getenv("S3_BUCKET", "")
S3_REGION = os.getenv("S3_REGION", "us-east-1")
S3_ENDPOINT = os.getenv("S3_ENDPOINT", "")

# Redis
REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379/0")

# AI Provider
AI_PROVIDER = os.getenv("AI_PROVIDER", "openai")  # openai | anthropic | custom
AI_API_KEY = os.getenv("AI_API_KEY", "")
AI_MODEL = os.getenv("AI_MODEL", "gpt-3.5-turbo")
AI_TEMPERATURE = float(os.getenv("AI_TEMPERATURE", "0.3"))
AI_MAX_TOKENS = int(os.getenv("AI_MAX_TOKENS", "512"))

# Email (SMTP)
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")
SMTP_FROM = os.getenv("SMTP_FROM", "noreply@crm.local")

# Jitsi (for meetings)
JITSI_BASE = os.getenv("JITSI_BASE", "https://meet.jit.si")

# Limits
MAX_UPLOAD_SIZE = int(os.getenv("MAX_UPLOAD_SIZE", str(50 * 1024 * 1024)))  # 50 MB
AVATAR_MAX_SIZE = int(os.getenv("AVATAR_MAX_SIZE", str(5 * 1024 * 1024)))  # 5 MB
AVATAR_ALLOWED_TYPES = {"image/jpeg", "image/png", "image/webp"}
AVATAR_RESIZE_TO = (256, 256)
AVATAR_CONTENT_SNIFF = True

# Rate limiting
RATE_LIMIT_ENABLED = os.getenv("RATE_LIMIT_ENABLED", "true").lower() == "true"
GLOBAL_RATE_LIMIT_PER_MIN = int(os.getenv("GLOBAL_RATE_LIMIT_PER_MIN", "600"))

# SSE (Server-Sent Events)
SSE_ENABLED = os.getenv("SSE_ENABLED", "true").lower() == "true"
SSE_MAX_CONN_PER_USER = int(os.getenv("SSE_MAX_CONN_PER_USER", "3"))

# Content Security Policy
CSP_ENABLED = os.getenv("CSP_ENABLED", "true").lower() == "true"

# Logging
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO")
LOG_FORMAT = os.getenv("LOG_FORMAT", "json")  # json | text


# ===== BLOCK: LOGGING SETUP =====

class JSONFormatter(logging.Formatter):
    """JSON log formatter для structured logging"""
    def format(self, record):
        log_obj = {
            "timestamp": datetime.utcnow().isoformat() + "Z",
            "level": record.levelname,
            "message": record.getMessage(),
            "module": record.module,
            "function": record.funcName,
            "line": record.lineno,
        }
        if record.exc_info:
            log_obj["exception"] = self.formatException(record.exc_info)
        return json.dumps(log_obj)

logging.basicConfig(
    level=getattr(logging, LOG_LEVEL.upper(), logging.INFO),
    format='%(message)s' if LOG_FORMAT == "json" else '%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)

if LOG_FORMAT == "json":
    for handler in logging.root.handlers:
        handler.setFormatter(JSONFormatter())

logger = logging.getLogger(__name__)


def log(level: str, message: str, **kwargs):
    """Unified logging helper"""
    level_map = {
        "DEBUG": logger.debug,
        "INFO": logger.info,
        "WARN": logger.warning,
        "ERROR": logger.error,
        "CRITICAL": logger.critical,
    }
    log_func = level_map.get(level.upper(), logger.info)
    
    if kwargs:
        if LOG_FORMAT == "json":
            # Merge kwargs into JSON
            extra_data = json.dumps(kwargs)
            log_func(f"{message} | {extra_data}")
        else:
            log_func(f"{message} | {kwargs}")
    else:
        log_func(message)


# ===== BLOCK: FLASK APP INITIALIZATION =====

app = Flask(__name__)
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_SIZE


# ===== BLOCK: UTILITIES =====

def utc_now() -> str:
    """Returns current UTC time in ISO format (without timezone suffix)"""
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def ensure_iso_datetime(s: str | None) -> str:
    """
    Normalize datetime strings to ISO format.
    Handles: YYYY-MM-DD HH:MM:SS, YYYY-MM-DDTHH:MM:SS, ISO with timezone
    """
    if not s:
        return ""
    s = str(s).strip()
    
    # Already in correct format
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", s):
        return s
    
    # ISO with T separator
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}", s):
        return s.replace("T", " ")
    
    # Try parsing with timezone
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass
    
    # Fallback: just date → add midnight
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return f"{s} 00:00:00"
    
    return s


def normalize_phone(phone: str | None) -> str:
    """
    Normalize phone to E.164 format (+7XXXXXXXXXX for RU)
    Handles: 8XXXXXXXXXX, +7XXXXXXXXXX, 7XXXXXXXXXX, XXXXXXXXXX
    """
    if not phone:
        return ""
    
    digits = re.sub(r"\D", "", phone)
    
    # Russian numbers
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    if len(digits) == 10:
        return "+7" + digits
    
    # International
    if len(digits) > 10:
        return "+" + digits
    
    return phone  # Return as-is if can't normalize


def validate_email(email: str | None) -> bool:
    """Simple email validation"""
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return bool(re.match(pattern, email))


def validate_inn(inn: str | None) -> bool:
    """Validate Russian INN (10 or 12 digits)"""
    if not inn:
        return False
    return bool(re.fullmatch(r"\d{10}|\d{12}", inn))


def generate_random_code(length: int = 6, charset: str = "0123456789") -> str:
    """Generate random code for 2FA/backup codes"""
    return "".join(secrets.choice(charset) for _ in range(length))


def detect_mime_from_bytes(data: bytes, filename: str = "") -> str | None:
    """
    Detect MIME type from file content (security: don't trust client headers).
    Uses python-magic if available, otherwise falls back to mimetypes.
    """
    try:
        import magic
        mime = magic.from_buffer(data, mime=True)
        return mime
    except ImportError:
        # Fallback to extension-based detection
        if filename:
            mime, _ = mimetypes.guess_type(filename)
            return mime
        return None


# ===== BLOCK: DATABASE HELPERS =====

def get_db():
    """
    Get database connection from Flask g context.
    ✅ ИСПРАВЛЕНО: isolation_level="DEFERRED" вместо None (autocommit)
    """
    if "db" not in g:
        g.db = sqlite3.connect(
            DATABASE_PATH,
            isolation_level="DEFERRED",  # ✅ ИЗМЕНЕНО для транзакций
            timeout=DB_TIMEOUT,
            check_same_thread=False,
        )
        g.db.row_factory = sqlite3.Row
        # Enable foreign keys
        g.db.execute("PRAGMA foreign_keys = ON")
        # Enable WAL mode for better concurrency
        g.db.execute("PRAGMA journal_mode = WAL")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    """Close database connection at end of request"""
    db = g.pop("db", None)
    if db is not None:
        db.close()


def query_db(query: str, args=(), one=False):
    """
    Execute SELECT query and return results.
    Returns dict-like Row objects.
    """
    try:
        cur = get_db().execute(query, args)
        rv = cur.fetchall()
        cur.close()
        return (rv[0] if rv else None) if one else rv
    except sqlite3.Error as e:
        log("ERROR", f"query_db error: {e}", query=query[:100])
        raise


def exec_db(query: str, args=()):
    """
    Execute INSERT/UPDATE/DELETE query.
    Returns lastrowid for INSERT.
    ✅ ИСПРАВЛЕНО: добавлен явный commit
    """
    try:
        con = get_db()
        cur = con.execute(query, args)
        last_id = cur.lastrowid
        con.commit()  # ✅ ДОБАВЛЕНО
        cur.close()
        return last_id
    except sqlite3.Error as e:
        log("ERROR", f"exec_db error: {e}", query=query[:100])
        get_db().rollback()
        raise


# ===== BLOCK: AUTH & CRYPTO =====

def hash_password(password: str) -> str:
    """Hash password using Werkzeug (pbkdf2:sha256)"""
    return generate_password_hash(password, method="pbkdf2:sha256")


def verify_password(password: str, password_hash: str) -> bool:
    """Verify password against hash"""
    return check_password_hash(password_hash, password)


def generate_csrf_token() -> str:
    """Generate CSRF token"""
    return secrets.token_hex(32)


def generate_api_token() -> str:
    """Generate API token (64 chars hex)"""
    return secrets.token_hex(32)


def hash_api_token(token: str) -> str:
    """Hash API token for storage"""
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def backup_code_hash(code: str) -> str:
    """
    Hash 2FA backup code using bcrypt.
    ✅ НОВОЕ: заменили SHA256+pepper на bcrypt с random salt
    """
    raw = code.strip().upper().replace("-", "")
    return bcrypt.hashpw(raw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def backup_code_verify(code: str, hashed: str) -> bool:
    """Verify 2FA backup code against hash"""
    raw = code.strip().upper().replace("-", "")
    try:
        return bcrypt.checkpw(raw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def generate_totp_secret() -> str:
    """Generate TOTP secret for 2FA"""
    if not PYOTP_AVAILABLE:
        return ""
    return pyotp.random_base32()


def verify_totp(secret: str, token: str) -> bool:
    """Verify TOTP token"""
    if not PYOTP_AVAILABLE:
        return False
    try:
        totp = pyotp.TOTP(secret)
        return totp.verify(token, valid_window=1)
    except Exception:
        return False


# ===== BLOCK: REDIS HELPERS =====

_redis_client = None

def get_redis():
    """Get Redis client (cached)"""
    global _redis_client
    if not REDIS_AVAILABLE:
        return None
    if _redis_client is None:
        try:
            _redis_client = redis.from_url(REDIS_URL, decode_responses=True)
            _redis_client.ping()
        except Exception as e:
            log("ERROR", f"Redis connection failed: {e}")
            return None
    return _redis_client


# ===== BLOCK: RATE LIMITING =====

_rate_buckets = defaultdict(list)  # In-memory fallback (per-worker)
_rate_lock = threading.Lock()


def rate_limit(key: str, per_min: int = 60) -> bool:
    """
    Check if rate limit exceeded.
    Returns True if OK, False if exceeded.
    ✅ УЛУЧШЕНО: приоритет Redis, fallback на in-memory
    """
    if not RATE_LIMIT_ENABLED:
        return True
    
    r = get_redis()
    now = time.time()
    
    if r:
        # Redis-based (distributed)
        try:
            pipe = r.pipeline()
            bucket_key = f"crm:ratelimit:{key}"
            pipe.zadd(bucket_key, {str(now): now})
            pipe.zremrangebyscore(bucket_key, 0, now - 60)
            pipe.zcard(bucket_key)
            pipe.expire(bucket_key, 120)
            _, _, count, _ = pipe.execute()
            return count <= per_min
        except Exception as e:
            log("WARN", f"Rate limit Redis error: {e}")
            # Fallback to in-memory
    
    # In-memory fallback
    with _rate_lock:
        bucket = _rate_buckets[key]
        # Remove old entries
        bucket[:] = [t for t in bucket if now - t < 60]
        # Check limit
        if len(bucket) >= per_min:
            return False
        bucket.append(now)
        return True


def _rate_limit(key_prefix: str, per_min: int = 60):
    """
    Decorator for rate limiting routes.
    Uses IP + user_id for key.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_id = g.get("user", {}).get("id", "anon")
            ip = request.headers.get("X-Forwarded-For", request.remote_addr)
            key = f"{key_prefix}:{user_id}:{ip}"
            
            if not rate_limit(key, per_min):
                return jsonify(ok=False, error="Rate limit exceeded"), 429
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# ===== BLOCK: SESSION & AUTH DECORATORS =====

def _get_current_user():
    """Get current user from session"""
    user_id = session.get("user_id")
    if not user_id:
        return None
    
    user = query_db(
        "SELECT * FROM users WHERE id=? AND active=1",
        (user_id,),
        one=True
    )
    return dict(user) if user else None


def _login_required(f):
    """Decorator: require authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = _get_current_user()
        if not user:
            if request.path.startswith("/api/"):
                return jsonify(ok=False, error="Unauthorized"), 401
            return redirect(url_for("login"))
        
        g.user = user
        return f(*args, **kwargs)
    return decorated_function


def _require_role(role: str):
    """Decorator: require specific role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = g.get("user")
            if not user:
                return jsonify(ok=False, error="Unauthorized"), 401
            
            if user.get("role") != role and user.get("role") != "admin":
                return jsonify(ok=False, error="Forbidden"), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def _csrf_protect(f):
    """Decorator: CSRF protection for POST/PUT/DELETE"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method in ("POST", "PUT", "DELETE", "PATCH"):
            token = request.headers.get("X-CSRFToken") or request.form.get("csrf_token")
            expected = session.get("csrf_token")
            
            if not token or not expected or token != expected:
                return jsonify(ok=False, error="CSRF validation failed"), 403
        
        return f(*args, **kwargs)
    return decorated_function


# ===== BLOCK: METRICS & MONITORING =====

_metrics = {
    "requests_total": 0,
    "requests_by_endpoint": defaultdict(int),
    "errors_total": 0,
    "db_queries_total": 0,
    "api_calls_total": defaultdict(int),  # ✅ НОВОЕ: трекинг AI calls
}
_metrics_lock = threading.Lock()


def _increment_metric(key: str, labels: dict = None):
    """Increment metric counter"""
    with _metrics_lock:
        if labels:
            _metrics[key][frozenset(labels.items())] += 1
        else:
            _metrics[key] += 1


@app.before_request
def before_request():
    """Request lifecycle: setup"""
    # Generate CSRF token if not exists
    if "csrf_token" not in session:
        session["csrf_token"] = generate_csrf_token()
    
    # Store request start time
    g.request_start_time = time.time()
    
    # Increment metrics
    _increment_metric("requests_total")
    _increment_metric("requests_by_endpoint", {"endpoint": request.endpoint or "unknown"})


@app.after_request
def after_request(response):
    """Request lifecycle: cleanup & headers"""
    # Calculate request duration
    if hasattr(g, "request_start_time"):
        duration = time.time() - g.request_start_time
        response.headers["X-Request-Duration"] = f"{duration:.3f}"
    
    # CSP header
    if CSP_ENABLED and not request.path.startswith("/api/"):
        nonce = secrets.token_hex(16)
        g.csp_nonce = nonce
        csp = (
            f"default-src 'self'; "
            f"script-src 'self' 'nonce-{nonce}' https://cdn.socket.io; "
            f"style-src 'self' 'unsafe-inline'; "
            f"img-src 'self' data: https:; "
            f"font-src 'self' data:; "
            f"connect-src 'self' wss: https:; "
            f"frame-src 'self' https://meet.jit.si; "
            f"object-src 'none'; "
            f"base-uri 'self'"
        )
        response.headers["Content-Security-Policy"] = csp
    
    return response


# ===== END OF CORE PART 1/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 2/10 — DATABASE SCHEMA & MIGRATIONS
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: DATABASE SCHEMA VERSION =====

SCHEMA_VERSION = 12  # ✅ УВЕЛИЧЕНО для новых таблиц


# ===== BLOCK: SCHEMA DEFINITION =====

SCHEMA_SQL = """
-- Organizations (multi-tenant root)
CREATE TABLE IF NOT EXISTS orgs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    slug TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    inn TEXT,
    settings_json TEXT DEFAULT '{}',
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT
);

-- Users
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    username TEXT NOT NULL,
    email TEXT,
    password_hash TEXT NOT NULL,
    role TEXT DEFAULT 'agent',
    department_id INTEGER,
    first_name TEXT,
    last_name TEXT,
    phone TEXT,
    avatar_url TEXT,
    timezone TEXT DEFAULT 'UTC',
    locale TEXT DEFAULT 'ru',
    active INTEGER DEFAULT 1,
    totp_secret TEXT,
    totp_enabled INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    last_login_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(department_id) REFERENCES departments(id) ON DELETE SET NULL,
    UNIQUE(org_id, username)
);

-- 2FA Backup Codes
CREATE TABLE IF NOT EXISTS user_2fa_backup_codes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    code_hash TEXT NOT NULL,
    used INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);

-- API Tokens
CREATE TABLE IF NOT EXISTS api_tokens (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    user_id INTEGER,
    name TEXT NOT NULL,
    token_hash TEXT NOT NULL,
    scopes TEXT DEFAULT 'read,write',
    active INTEGER DEFAULT 1,
    expires_at TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    last_used_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);

-- Departments
CREATE TABLE IF NOT EXISTS departments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    slug TEXT,
    parent_id INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(parent_id) REFERENCES departments(id) ON DELETE SET NULL
);

-- Channels (communication channels)
CREATE TABLE IF NOT EXISTS channels (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    type TEXT NOT NULL,
    name TEXT,
    config_json TEXT DEFAULT '{}',
    secret TEXT,
    active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- Companies (clients/customers)
CREATE TABLE IF NOT EXISTS companies (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    inn TEXT,
    phone TEXT,
    phone_norm TEXT,
    email TEXT,
    address TEXT,
    notes TEXT,
    score INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- Contacts (people associated with companies)
CREATE TABLE IF NOT EXISTS contacts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    company_id INTEGER,
    name TEXT NOT NULL,
    position TEXT,
    phone TEXT,
    phone_norm TEXT,
    email TEXT,
    notes TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL
);

-- Inbox Threads (unified inbox)
CREATE TABLE IF NOT EXISTS inbox_threads (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    channel_id INTEGER,
    subject TEXT,
    status TEXT DEFAULT 'open',
    priority TEXT DEFAULT 'normal',
    assignee_id INTEGER,
    company_id INTEGER,
    contact_id INTEGER,
    external_thread_id TEXT,
    tags_csv TEXT,
    first_response_at TEXT,
    first_response_due_at TEXT,
    last_message_at TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(channel_id) REFERENCES channels(id) ON DELETE SET NULL,
    FOREIGN KEY(assignee_id) REFERENCES users(id) ON DELETE SET NULL,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL,
    FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
);

-- Inbox Messages
CREATE TABLE IF NOT EXISTS inbox_messages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    thread_id INTEGER NOT NULL,
    sender_type TEXT NOT NULL,
    user_id INTEGER,
    external_user_id TEXT,
    body TEXT,
    internal_note INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(thread_id) REFERENCES inbox_threads(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Message Attachments
CREATE TABLE IF NOT EXISTS message_attachments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    message_id INTEGER NOT NULL,
    file_id INTEGER NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(message_id) REFERENCES inbox_messages(id) ON DELETE CASCADE,
    FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
);

-- Tasks
CREATE TABLE IF NOT EXISTS tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    description TEXT,
    status TEXT DEFAULT 'open',
    priority TEXT DEFAULT 'normal',
    assignee_id INTEGER,
    company_id INTEGER,
    contact_id INTEGER,
    due_at TEXT,
    completed_at TEXT,
    monthly_fee REAL DEFAULT 0,
    address TEXT,
    contact_phone TEXT,
    last_commented_at TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(assignee_id) REFERENCES users(id) ON DELETE SET NULL,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL,
    FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
);

-- Task Statuses (custom workflow states)
CREATE TABLE IF NOT EXISTS task_statuses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    color TEXT DEFAULT '#888',
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    UNIQUE(org_id, name)
);

-- Task Comments
CREATE TABLE IF NOT EXISTS task_comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    user_id INTEGER,
    body TEXT NOT NULL,
    format TEXT DEFAULT 'plain',
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Task Comment Attachments
CREATE TABLE IF NOT EXISTS task_comment_attachments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    comment_id INTEGER NOT NULL,
    file_id INTEGER NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(comment_id) REFERENCES task_comments(id) ON DELETE CASCADE,
    FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
);

-- Task Participants
CREATE TABLE IF NOT EXISTS task_participants (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    user_id INTEGER NOT NULL,
    role TEXT DEFAULT 'watcher',
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
    UNIQUE(task_id, user_id)
);

-- Task Checklists
CREATE TABLE IF NOT EXISTS task_checklists (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    item TEXT NOT NULL,
    checked INTEGER DEFAULT 0,
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE
);

-- Task Reminders
CREATE TABLE IF NOT EXISTS task_reminders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    user_id INTEGER NOT NULL,
    remind_at TEXT NOT NULL,
    message TEXT,
    sent INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);

-- Task Files (pinned files)
CREATE TABLE IF NOT EXISTS task_files (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    file_id INTEGER NOT NULL,
    pinned INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE,
    FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE,
    UNIQUE(task_id, file_id)
);

-- Task Activity Log
CREATE TABLE IF NOT EXISTS task_activity (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    user_id INTEGER,
    kind TEXT NOT NULL,
    details_json TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(task_id) REFERENCES tasks(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Deals
CREATE TABLE IF NOT EXISTS deals (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    amount REAL DEFAULT 0,
    currency TEXT DEFAULT 'RUB',
    status TEXT DEFAULT 'open',
    stage TEXT,
    pipeline_key TEXT DEFAULT 'default',
    assignee_id INTEGER,
    company_id INTEGER,
    contact_id INTEGER,
    due_at TEXT,
    won_at TEXT,
    lost_at TEXT,
    score INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(assignee_id) REFERENCES users(id) ON DELETE SET NULL,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL,
    FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
);

-- Workflow Stages (pipeline stages)
CREATE TABLE IF NOT EXISTS workflow_stages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    entity_type TEXT NOT NULL,
    pipeline_key TEXT DEFAULT 'default',
    key TEXT NOT NULL,
    name TEXT NOT NULL,
    sort_order INTEGER DEFAULT 0,
    sla_hours INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    UNIQUE(org_id, entity_type, pipeline_key, key)
);

-- Stage Transitions (audit trail)
CREATE TABLE IF NOT EXISTS stage_transitions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    entity_type TEXT NOT NULL,
    entity_id INTEGER NOT NULL,
    from_stage TEXT,
    to_stage TEXT,
    user_id INTEGER,
    comment TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Workflow Definitions
CREATE TABLE IF NOT EXISTS workflow_definitions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    trigger_event TEXT NOT NULL,
    trigger_filter_json TEXT,
    graph_json TEXT NOT NULL,
    active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- Workflow Tasks (execution queue)
CREATE TABLE IF NOT EXISTS workflow_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    workflow_id INTEGER NOT NULL,
    node_id TEXT NOT NULL,
    entity_type TEXT,
    entity_id INTEGER,
    context_json TEXT,
    status TEXT DEFAULT 'pending',
    scheduled_at TEXT,
    started_at TEXT,
    completed_at TEXT,
    error TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(workflow_id) REFERENCES workflow_definitions(id) ON DELETE CASCADE
);

-- Calls (telephony CDR)
CREATE TABLE IF NOT EXISTS calls (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    channel_id INTEGER,
    external_call_id TEXT,
    direction TEXT,
    from_e164 TEXT,
    to_e164 TEXT,
    agent_id INTEGER,
    company_id INTEGER,
    contact_id INTEGER,
    status TEXT,
    duration_sec INTEGER DEFAULT 0,
    recording_url TEXT,
    started_at TEXT,
    ended_at TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(channel_id) REFERENCES channels(id) ON DELETE SET NULL,
    FOREIGN KEY(agent_id) REFERENCES users(id) ON DELETE SET NULL,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL,
    FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL
);

-- Meetings
CREATE TABLE IF NOT EXISTS meetings (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    title TEXT,
    room TEXT NOT NULL,
    start_at TEXT,
    end_at TEXT,
    created_by INTEGER,
    participants_json TEXT DEFAULT '[]',
    recording_started_at TEXT,
    recording_stopped_at TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(created_by) REFERENCES users(id) ON DELETE SET NULL
);

-- Chat Channels
CREATE TABLE IF NOT EXISTS chat_channels (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    type TEXT DEFAULT 'public',
    title TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- Chat Members
CREATE TABLE IF NOT EXISTS chat_members (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    channel_id INTEGER NOT NULL,
    user_id INTEGER,
    department_id INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(channel_id) REFERENCES chat_channels(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE,
    FOREIGN KEY(department_id) REFERENCES departments(id) ON DELETE CASCADE
);

-- Chat Messages
CREATE TABLE IF NOT EXISTS chat_messages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    channel_id INTEGER NOT NULL,
    user_id INTEGER,
    body TEXT NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(channel_id) REFERENCES chat_channels(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Files (unified file storage metadata)
CREATE TABLE IF NOT EXISTS files (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    storage_key TEXT NOT NULL,
    name TEXT NOT NULL,
    content_type TEXT,
    size_bytes INTEGER DEFAULT 0,
    uploaded_by INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(uploaded_by) REFERENCES users(id) ON DELETE SET NULL
);

-- Documents
CREATE TABLE IF NOT EXISTS documents (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    template_id INTEGER,
    doc_type TEXT,
    title TEXT NOT NULL,
    content_html TEXT,
    company_id INTEGER,
    user_id INTEGER,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(template_id) REFERENCES document_templates(id) ON DELETE SET NULL,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Document Templates
CREATE TABLE IF NOT EXISTS document_templates (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    type TEXT NOT NULL,
    tkey TEXT,
    name TEXT NOT NULL,
    body_template TEXT NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- Products (CPQ/Warehouse)
CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    sku TEXT,
    name TEXT NOT NULL,
    description TEXT,
    price REAL DEFAULT 0,
    currency TEXT DEFAULT 'RUB',
    qty REAL DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    UNIQUE(org_id, sku)
);

-- Webhooks
CREATE TABLE IF NOT EXISTS webhooks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    event TEXT NOT NULL,
    url TEXT NOT NULL,
    secret TEXT,
    active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- Webhook Delivery Queue
CREATE TABLE IF NOT EXISTS webhook_queue (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    webhook_id INTEGER NOT NULL,
    event TEXT NOT NULL,
    payload_json TEXT NOT NULL,
    status TEXT DEFAULT 'pending',
    attempts INTEGER DEFAULT 0,
    next_try_at TEXT,
    last_error TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(webhook_id) REFERENCES webhooks(id) ON DELETE CASCADE
);

-- Audit Logs
CREATE TABLE IF NOT EXISTS audit_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    user_id INTEGER,
    action TEXT NOT NULL,
    entity_type TEXT,
    entity_id INTEGER,
    details TEXT,
    ip_address TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- AI Jobs
CREATE TABLE IF NOT EXISTS ai_jobs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    user_id INTEGER,
    job_type TEXT NOT NULL,
    entity_type TEXT,
    entity_id INTEGER,
    input_text TEXT,
    output_text TEXT,
    model TEXT,
    tokens_used INTEGER DEFAULT 0,
    status TEXT DEFAULT 'pending',
    error TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    completed_at TEXT,
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);

-- Saved Views (filters)
CREATE TABLE IF NOT EXISTS saved_views (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    user_id INTEGER,
    entity_type TEXT NOT NULL,
    name TEXT NOT NULL,
    filter_json TEXT NOT NULL,
    is_default INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);

-- Schema Migrations Tracking
CREATE TABLE IF NOT EXISTS schema_migrations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    version INTEGER NOT NULL UNIQUE,
    applied_at TEXT DEFAULT (datetime('now'))
);

-- ✅ НОВЫЕ ТАБЛИЦЫ ДЛЯ РАСШИРЕННОГО ФУНКЦИОНАЛА

-- Email Sequences (Drip Campaigns)
CREATE TABLE IF NOT EXISTS email_sequences (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS sequence_steps (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sequence_id INTEGER NOT NULL,
    step_num INTEGER NOT NULL,
    delay_hours INTEGER DEFAULT 24,
    subject TEXT,
    body_template TEXT,
    send_condition TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(sequence_id) REFERENCES email_sequences(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS sequence_enrollments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sequence_id INTEGER NOT NULL,
    contact_id INTEGER,
    company_id INTEGER,
    email TEXT NOT NULL,
    current_step INTEGER DEFAULT 0,
    status TEXT DEFAULT 'active',
    enrolled_at TEXT DEFAULT (datetime('now')),
    last_sent_at TEXT,
    FOREIGN KEY(sequence_id) REFERENCES email_sequences(id) ON DELETE CASCADE,
    FOREIGN KEY(contact_id) REFERENCES contacts(id) ON DELETE SET NULL,
    FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE SET NULL
);

-- Lead Scoring Rules
CREATE TABLE IF NOT EXISTS lead_scoring_rules (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    entity_type TEXT DEFAULT 'deal',
    field TEXT NOT NULL,
    operator TEXT NOT NULL,
    value TEXT,
    score_delta INTEGER DEFAULT 10,
    active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- AI Embeddings (vector storage)
CREATE TABLE IF NOT EXISTS embeddings (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    entity_type TEXT NOT NULL,
    entity_id INTEGER NOT NULL,
    model TEXT NOT NULL,
    vector BLOB NOT NULL,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
    UNIQUE(org_id, entity_type, entity_id, model)
);

-- AI Agent Actions (audit trail for autonomous agents)
CREATE TABLE IF NOT EXISTS agent_actions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    org_id INTEGER NOT NULL,
    agent_name TEXT NOT NULL,
    action_type TEXT NOT NULL,
    entity_type TEXT,
    entity_id INTEGER,
    reasoning TEXT,
    success INTEGER DEFAULT 1,
    error TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE
);

-- AI Feedback (user feedback on AI outputs)
CREATE TABLE IF NOT EXISTS ai_feedback (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ai_job_id INTEGER NOT NULL,
    user_id INTEGER,
    rating INTEGER,
    correction TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    FOREIGN KEY(ai_job_id) REFERENCES ai_jobs(id) ON DELETE CASCADE,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);
"""


# ===== BLOCK: INDEXES =====

def ensure_indexes():
    """
    Create indexes for performance optimization.
    ✅ ИСПРАВЛЕНО: убран placeholder text
    """
    con = get_db()
    cur = con.cursor()
    
    indexes = [
        # Users
        "CREATE INDEX IF NOT EXISTS idx_users_org ON users(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_users_username ON users(org_id, username)",
        "CREATE INDEX IF NOT EXISTS idx_users_email ON users(email)",
        
        # Companies
        "CREATE INDEX IF NOT EXISTS idx_companies_org ON companies(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_companies_inn ON companies(inn)",
        "CREATE INDEX IF NOT EXISTS idx_companies_phone_norm ON companies(phone_norm)",
        
        # Contacts
        "CREATE INDEX IF NOT EXISTS idx_contacts_org ON contacts(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_contacts_company ON contacts(company_id)",
        "CREATE INDEX IF NOT EXISTS idx_contacts_phone_norm ON contacts(phone_norm)",
        
        # Inbox threads
        "CREATE INDEX IF NOT EXISTS idx_inbox_threads_org ON inbox_threads(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_inbox_threads_channel ON inbox_threads(channel_id)",
        "CREATE INDEX IF NOT EXISTS idx_inbox_threads_assignee ON inbox_threads(assignee_id)",
        "CREATE INDEX IF NOT EXISTS idx_inbox_threads_status ON inbox_threads(status)",
        "CREATE INDEX IF NOT EXISTS idx_inbox_threads_last_msg ON inbox_threads(last_message_at DESC)",
        
        # Inbox messages
        "CREATE INDEX IF NOT EXISTS idx_inbox_messages_thread ON inbox_messages(thread_id)",
        "CREATE INDEX IF NOT EXISTS idx_inbox_messages_created ON inbox_messages(created_at DESC)",
        
        # Tasks
        "CREATE INDEX IF NOT EXISTS idx_tasks_org ON tasks(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_assignee ON tasks(assignee_id)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_status ON tasks(status)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_due ON tasks(due_at)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_company ON tasks(company_id)",
        
        # Deals
        "CREATE INDEX IF NOT EXISTS idx_deals_org ON deals(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_deals_assignee ON deals(assignee_id)",
        "CREATE INDEX IF NOT EXISTS idx_deals_status ON deals(status)",
        "CREATE INDEX IF NOT EXISTS idx_deals_stage ON deals(stage)",
        "CREATE INDEX IF NOT EXISTS idx_deals_pipeline ON deals(pipeline_key)",
        "CREATE INDEX IF NOT EXISTS idx_deals_company ON deals(company_id)",
        
        # Calls
        "CREATE INDEX IF NOT EXISTS idx_calls_org ON calls(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_calls_agent ON calls(agent_id)",
        "CREATE INDEX IF NOT EXISTS idx_calls_started ON calls(started_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_calls_company ON calls(company_id)",
        
        # Audit logs
        "CREATE INDEX IF NOT EXISTS idx_audit_org ON audit_logs(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_logs(user_id)",
        "CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_logs(created_at DESC)",
        
        # AI jobs
        "CREATE INDEX IF NOT EXISTS idx_ai_jobs_org ON ai_jobs(org_id)",
        "CREATE INDEX IF NOT EXISTS idx_ai_jobs_type ON ai_jobs(job_type)",
        "CREATE INDEX IF NOT EXISTS idx_ai_jobs_status ON ai_jobs(status)",
        
        # Webhooks
        "CREATE INDEX IF NOT EXISTS idx_webhook_queue_status ON webhook_queue(status, next_try_at)",
        
        # Email sequences
        "CREATE INDEX IF NOT EXISTS idx_sequence_enrollments_status ON sequence_enrollments(status, last_sent_at)",
        
        # Embeddings
        "CREATE INDEX IF NOT EXISTS idx_embeddings_entity ON embeddings(org_id, entity_type, entity_id)",
        
        # Agent actions
        "CREATE INDEX IF NOT EXISTS idx_agent_actions_org ON agent_actions(org_id, created_at DESC)",
    ]
    
    for idx_sql in indexes:
        try:
            cur.execute(idx_sql)
        except sqlite3.Error as e:
            log("WARN", f"Index creation failed: {e}", sql=idx_sql[:50])
    
    con.commit()
    log("INFO", "Database indexes created")


# ===== BLOCK: MIGRATIONS =====

def get_current_schema_version() -> int:
    """Get current schema version from database"""
    try:
        row = query_db("SELECT MAX(version) as v FROM schema_migrations", one=True)
        return row["v"] if row and row["v"] else 0
    except sqlite3.Error:
        return 0


def record_migration(version: int):
    """Record migration as applied"""
    exec_db("INSERT INTO schema_migrations (version) VALUES (?)", (version,))


def ensure_schema():
    """
    Ensure database schema is up-to-date.
    Creates tables if not exist, runs migrations if needed.
    ✅ ИСПРАВЛЕНО: убран placeholder text
    """
    con = get_db()
    cur = con.cursor()
    
    # Create base schema
    try:
        cur.executescript(SCHEMA_SQL)
        con.commit()
        log("INFO", "Database schema initialized")
    except sqlite3.Error as e:
        log("ERROR", f"Schema creation failed: {e}")
        raise
    
    # Create indexes
    ensure_indexes()
    
    # Run migrations
    current_version = get_current_schema_version()
    
    if current_version < SCHEMA_VERSION:
        log("INFO", f"Running migrations from v{current_version} to v{SCHEMA_VERSION}")
        run_migrations(current_version)
    
    log("INFO", f"Database schema version: {SCHEMA_VERSION}")


def run_migrations(from_version: int):
    """
    Run all migrations newer than from_version.
    ✅ ИСПРАВЛЕНО: закрыты все SQL-запросы, убраны placeholders
    """
    migrations = [
        (1, _migration_1_add_phone_norm),
        (2, _migration_2_add_workflow_tables),
        (3, _migration_3_add_task_reminders),
        (4, _migration_4_add_ai_jobs),
        (5, _migration_5_add_saved_views),
        (6, _migration_6_fix_users_unique_username_scoped),
        (7, _migration_7_add_chat_tables),
        (8, _migration_8_add_lead_scoring),
        (9, _migration_9_add_email_sequences),
        (10, _migration_10_add_embeddings),
        (11, _migration_11_add_agent_actions),
        (12, _migration_12_add_ai_feedback),
    ]
    
    for version, migration_func in migrations:
        if version > from_version:
            try:
                log("INFO", f"Applying migration v{version}: {migration_func.__name__}")
                migration_func()
                record_migration(version)
                log("INFO", f"Migration v{version} completed")
            except Exception as e:
                log("ERROR", f"Migration v{version} failed: {e}")
                raise


def _migration_1_add_phone_norm():
    """Add phone_norm column to companies and contacts"""
    con = get_db()
    try:
        con.execute("ALTER TABLE companies ADD COLUMN phone_norm TEXT")
        con.execute("ALTER TABLE contacts ADD COLUMN phone_norm TEXT")
        con.commit()
    except sqlite3.OperationalError:
        pass


def _migration_2_add_workflow_tables():
    """Add workflow-related tables (already in SCHEMA_SQL for new installs)"""
    pass


def _migration_3_add_task_reminders():
    """Add task_reminders table (already in SCHEMA_SQL)"""
    pass


def _migration_4_add_ai_jobs():
    """Add ai_jobs table (already in SCHEMA_SQL)"""
    pass


def _migration_5_add_saved_views():
    """Add saved_views table (already in SCHEMA_SQL)"""
    pass


def _migration_6_fix_users_unique_username_scoped():
    """
    Fix users table: UNIQUE constraint should be (org_id, username).
    ✅ ИСПРАВЛЕНО: SQL закрыт корректно
    """
    con = get_db()
    cur = con.cursor()
    
    try:
        cur.execute("BEGIN EXCLUSIVE")
        
        # Create new table with correct constraint
        cur.execute("""
            CREATE TABLE users_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                org_id INTEGER NOT NULL,
                username TEXT NOT NULL,
                email TEXT,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'agent',
                department_id INTEGER,
                first_name TEXT,
                last_name TEXT,
                phone TEXT,
                avatar_url TEXT,
                timezone TEXT DEFAULT 'UTC',
                locale TEXT DEFAULT 'ru',
                active INTEGER DEFAULT 1,
                totp_secret TEXT,
                totp_enabled INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now')),
                last_login_at TEXT,
                FOREIGN KEY(org_id) REFERENCES orgs(id) ON DELETE CASCADE,
                FOREIGN KEY(department_id) REFERENCES departments(id) ON DELETE SET NULL,
                UNIQUE(org_id, username)
            )
        """)
        
        # Copy data
        cur.execute("""
            INSERT INTO users_new 
            SELECT * FROM users
        """)
        
        # Swap tables
        cur.execute("DROP TABLE users")
        cur.execute("ALTER TABLE users_new RENAME TO users")
        
        con.commit()
    except sqlite3.Error as e:
        con.rollback()
        log("WARN", f"Migration 6 failed (may already be applied): {e}")


def _migration_7_add_chat_tables():
    """Add chat_channels, chat_members, chat_messages (already in SCHEMA_SQL)"""
    pass


def _migration_8_add_lead_scoring():
    """
    Add score column to deals and companies, add lead_scoring_rules table.
    """
    con = get_db()
    try:
        con.execute("ALTER TABLE deals ADD COLUMN score INTEGER DEFAULT 0")
        con.execute("ALTER TABLE companies ADD COLUMN score INTEGER DEFAULT 0")
        con.commit()
    except sqlite3.OperationalError:
        pass


def _migration_9_add_email_sequences():
    """Add email_sequences, sequence_steps, sequence_enrollments (already in SCHEMA_SQL)"""
    pass


def _migration_10_add_embeddings():
    """Add embeddings table (already in SCHEMA_SQL)"""
    pass


def _migration_11_add_agent_actions():
    """Add agent_actions table (already in SCHEMA_SQL)"""
    pass


def _migration_12_add_ai_feedback():
    """Add ai_feedback table (already in SCHEMA_SQL)"""
    pass


# ===== BLOCK: SEED DATA =====

def seed_defaults():
    """
    Seed default organization and admin user if database is empty.
    ✅ ИСПРАВЛЕНО: убран placeholder text, дописана логика
    """
    # Check if org exists
    org = query_db("SELECT id FROM orgs LIMIT 1", one=True)
    if not org:
        slug = "demo"
        org_id = exec_db(
            "INSERT INTO orgs (slug, name) VALUES (?, ?)",
            (slug, "Demo Organization")
        )
        log("INFO", f"Created default org: {slug}")
    else:
        org_id = org["id"]
    
    # Check if admin user exists
    user_cnt = query_db("SELECT COUNT(*) as cnt FROM users WHERE org_id=?", (org_id,), one=True)
    if user_cnt["cnt"] == 0:
        password_hash = hash_password("admin")
        exec_db(
            """
            INSERT INTO users (org_id, username, email, password_hash, role, active)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (org_id, "admin", "admin@demo.local", password_hash, "admin", 1)
        )
        log("INFO", "Created default admin user: admin / admin")
    
    # Seed default department
    dept_cnt = query_db("SELECT COUNT(*) as cnt FROM departments WHERE org_id=?", (org_id,), one=True)
    if dept_cnt["cnt"] == 0:
        exec_db(
            "INSERT INTO departments (org_id, name, slug) VALUES (?, ?, ?)",
            (org_id, "Общий отдел", "general")
        )
        log("INFO", "Created default department")
    
    # Seed default task statuses
    status_cnt = query_db("SELECT COUNT(*) as cnt FROM task_statuses WHERE org_id=?", (org_id,), one=True)
    if status_cnt["cnt"] == 0:
        statuses = [
            ("open", "#2bd66a", 1),
            ("in_progress", "#ffa500", 2),
            ("blocked", "#ff4444", 3),
            ("done", "#888888", 4),
        ]
        for name, color, order in statuses:
            exec_db(
                "INSERT INTO task_statuses (org_id, name, color, sort_order) VALUES (?, ?, ?, ?)",
                (org_id, name, color, order)
            )
        log("INFO", "Created default task statuses")
    
    # Seed default workflow stages (deal pipeline)
    stage_cnt = query_db(
        "SELECT COUNT(*) as cnt FROM workflow_stages WHERE org_id=? AND entity_type='deal'",
        (org_id,),
        one=True
    )
    if stage_cnt["cnt"] == 0:
        stages = [
            ("new", "Новая", 1, 24),
            ("qualify", "Квалификация", 2, 48),
            ("proposal", "Предложение", 3, 72),
            ("negotiation", "Переговоры", 4, 120),
            ("closed_won", "Выиграна", 5, None),
            ("closed_lost", "Проиграна", 6, None),
        ]
        for key, name, order, sla in stages:
            exec_db(
                """
                INSERT INTO workflow_stages (org_id, entity_type, pipeline_key, key, name, sort_order, sla_hours)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (org_id, "deal", "default", key, name, order, sla)
            )
        log("INFO", "Created default deal pipeline stages")


# ===== BLOCK: CLI MIGRATION HELPER =====

def run_migrations_cli_if_requested():
    """
    Check if script was run with --migrate flag.
    If yes, run migrations and exit.
    """
    if "--migrate" in sys.argv:
        log("INFO", "Running migrations (CLI mode)")
        ensure_schema()
        seed_defaults()
        log("INFO", "Migrations completed. Exiting.")
        sys.exit(0)


# ===== END OF CORE PART 2/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 3/10 — STORAGE, FTS, EMAIL & UTILITIES
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: STORAGE HELPERS =====

def safe_local_storage_path(upload_dir: str, key: str) -> str:
    """
    Safely construct local storage path, preventing path traversal.
    ✅ УЛУЧШЕНО: добавлена защита от symlink escape
    """
    base = os.path.abspath(upload_dir)
    
    # Sanitize key
    safe = key.replace("..", "").replace("/", "_").replace("\\", "_")
    
    # Construct path
    path = os.path.abspath(os.path.join(base, safe))
    
    # Check path escape
    if not (path == base or path.startswith(base + os.sep)):
        raise ValueError("path escape detected")
    
    # ✅ НОВОЕ: Check symlink escape
    if os.path.exists(path):
        real_path = os.path.realpath(path)
        real_base = os.path.realpath(base)
        if not (real_path == real_base or real_path.startswith(real_base + os.sep)):
            raise ValueError("symlink escape detected")
    
    return path


def store_file_local(key: str, data: bytes) -> str:
    """
    Store file in local filesystem.
    Returns URL path for retrieval.
    """
    os.makedirs(LOCAL_UPLOAD_DIR, exist_ok=True)
    
    path = safe_local_storage_path(LOCAL_UPLOAD_DIR, key)
    
    with open(path, "wb") as f:
        f.write(data)
    
    return f"/uploads/{key}"


def get_file_local(key: str) -> bytes | None:
    """Retrieve file from local storage"""
    try:
        path = safe_local_storage_path(LOCAL_UPLOAD_DIR, key)
        with open(path, "rb") as f:
            return f.read()
    except (FileNotFoundError, ValueError):
        return None


def store_file_s3(key: str, data: bytes, content_type: str = "application/octet-stream") -> str:
    """
    Store file in S3-compatible storage.
    Returns public URL.
    """
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        raise RuntimeError("boto3 not installed for S3 storage")
    
    s3_client = boto3.client(
        "s3",
        region_name=S3_REGION,
        endpoint_url=S3_ENDPOINT or None,
    )
    
    try:
        s3_client.put_object(
            Bucket=S3_BUCKET,
            Key=key,
            Body=data,
            ContentType=content_type,
        )
    except ClientError as e:
        log("ERROR", f"S3 upload failed: {e}")
        raise
    
    # Generate URL
    if S3_ENDPOINT:
        return f"{S3_ENDPOINT}/{S3_BUCKET}/{key}"
    else:
        return f"https://{S3_BUCKET}.s3.{S3_REGION}.amazonaws.com/{key}"


def get_file_s3(key: str) -> bytes | None:
    """Retrieve file from S3"""
    try:
        import boto3
        from botocore.exceptions import ClientError
    except ImportError:
        return None
    
    s3_client = boto3.client(
        "s3",
        region_name=S3_REGION,
        endpoint_url=S3_ENDPOINT or None,
    )
    
    try:
        response = s3_client.get_object(Bucket=S3_BUCKET, Key=key)
        return response["Body"].read()
    except ClientError:
        return None


def store_file(org_id: int, filename: str, data: bytes, content_type: str, user_id: int = None) -> dict:
    """
    Store file using configured backend (local or S3).
    Creates metadata record in files table.
    Returns file metadata dict.
    """
    # Generate unique key
    ext = os.path.splitext(filename)[1][:10]  # Limit extension length
    storage_key = f"{org_id}/{uuid.uuid4().hex}{ext}"
    
    # Store file
    if STORAGE_BACKEND == "s3":
        url = store_file_s3(storage_key, data, content_type)
    else:
        url = store_file_local(storage_key, data)
    
    # Create metadata record
    file_id = exec_db(
        """
        INSERT INTO files (org_id, storage_key, name, content_type, size_bytes, uploaded_by)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (org_id, storage_key, filename, content_type, len(data), user_id)
    )
    
    return {
        "id": file_id,
        "name": filename,
        "url": url,
        "content_type": content_type,
        "size_bytes": len(data),
    }


def get_file_by_id(file_id: int) -> tuple[bytes, str, str] | None:
    """
    Retrieve file by ID.
    Returns (data, filename, content_type) or None.
    """
    file_meta = query_db("SELECT * FROM files WHERE id=?", (file_id,), one=True)
    if not file_meta:
        return None
    
    # Get file data
    if STORAGE_BACKEND == "s3":
        data = get_file_s3(file_meta["storage_key"])
    else:
        data = get_file_local(file_meta["storage_key"])
    
    if not data:
        return None
    
    return (data, file_meta["name"], file_meta["content_type"])


# ===== BLOCK: AVATAR PROCESSING =====

def process_avatar(data: bytes, filename: str) -> tuple[bytes, str]:
    """
    Process avatar image: validate, resize, optimize.
    Returns (processed_data, content_type).
    ✅ НОВОЕ: avatar resizing для оптимизации
    """
    # Detect actual MIME type (security)
    detected_type = detect_mime_from_bytes(data, filename)
    
    # Fallback to extension-based if detection failed
    if not detected_type and filename:
        detected_type, _ = mimetypes.guess_type(filename)
    
    content_type = detected_type or "application/octet-stream"
    
    # Validate type
    if not content_type.startswith("image/") or content_type not in AVATAR_ALLOWED_TYPES:
        raise ValueError(f"Unsupported avatar type: {content_type}")
    
    # Validate size
    if len(data) > AVATAR_MAX_SIZE:
        raise ValueError(f"Avatar too large: {len(data)} bytes")
    
    # Resize image
    try:
        img = Image.open(BytesIO(data))
        
        # Convert to RGB if needed (for JPEG)
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")
        
        # Resize maintaining aspect ratio
        img.thumbnail(AVATAR_RESIZE_TO, Image.Resampling.LANCZOS)
        
        # Save to buffer
        buffer = BytesIO()
        
        # Determine format
        if content_type == "image/png":
            img.save(buffer, format="PNG", optimize=True)
        elif content_type == "image/webp":
            img.save(buffer, format="WEBP", quality=85)
        else:  # JPEG
            img.save(buffer, format="JPEG", quality=85, optimize=True)
            content_type = "image/jpeg"
        
        return (buffer.getvalue(), content_type)
    
    except Exception as e:
        log("ERROR", f"Avatar processing failed: {e}")
        raise ValueError("Invalid image file")


# ===== BLOCK: HTML SANITIZATION =====

ALLOWED_HTML_TAGS = [
    'p', 'br', 'strong', 'em', 'u', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
    'ul', 'ol', 'li', 'a', 'img', 'table', 'thead', 'tbody', 'tr', 'th', 'td',
    'div', 'span', 'hr', 'blockquote', 'code', 'pre'
]

ALLOWED_HTML_ATTRS = {
    'a': ['href', 'title', 'target'],
    'img': ['src', 'alt', 'width', 'height'],
    'table': ['border', 'cellpadding', 'cellspacing'],
    '*': ['class', 'style']
}


def sanitize_html(html: str) -> str:
    """
    Sanitize user-provided HTML to prevent XSS.
    ✅ НОВОЕ: для безопасного вывода document content
    """
    return bleach.clean(
        html,
        tags=ALLOWED_HTML_TAGS,
        attributes=ALLOWED_HTML_ATTRS,
        strip=True
    )


# ===== BLOCK: FULL-TEXT SEARCH (FTS5) =====

def fts_sanitize(q: str) -> str:
    """
    Sanitize FTS query to prevent injection.
    ✅ УЛУЧШЕНО: экранирование спецсимволов FTS5
    """
    if not q:
        return ""
    
    # Remove FTS5 operators
    q = q.replace('"', '').replace('*', '').replace('^', '').replace('-', ' ')
    
    # Split into tokens
    tokens = [t for t in re.split(r"\s+", q) if t]
    
    # Quote each token
    return " ".join(f'"{t}"' for t in tokens)


def rebuild_fts_table(table_name: str):
    """
    Rebuild FTS table using shadow table pattern (non-blocking).
    ✅ УЛУЧШЕНО: используем shadow tables для минимизации блокировки
    """
    con = get_db()
    cur = con.cursor()
    
    shadow_name = f"{table_name}_shadow"
    
    try:
        # Step 1: Create shadow table
        cur.execute(f"DROP TABLE IF EXISTS {shadow_name}")
        
        if table_name == "fts_inbox_messages":
            cur.execute(f"""
                CREATE VIRTUAL TABLE {shadow_name} USING fts5(
                    body, username, external_user_id,
                    content='inbox_messages',
                    content_rowid='id'
                )
            """)
            
            # Populate shadow table
            cur.execute(f"""
                INSERT INTO {shadow_name}(rowid, body, username, external_user_id)
                SELECT 
                    m.id,
                    COALESCE(m.body, ''),
                    COALESCE(u.username, ''),
                    COALESCE(m.external_user_id, '')
                FROM inbox_messages m
                LEFT JOIN users u ON m.user_id = u.id
            """)
        
        elif table_name == "fts_tasks":
            cur.execute(f"""
                CREATE VIRTUAL TABLE {shadow_name} USING fts5(
                    title, description, address,
                    content='tasks',
                    content_rowid='id'
                )
            """)
            
            cur.execute(f"""
                INSERT INTO {shadow_name}(rowid, title, description, address)
                SELECT 
                    id,
                    COALESCE(title, ''),
                    COALESCE(description, ''),
                    COALESCE(address, '')
                FROM tasks
            """)
        
        elif table_name == "fts_chat_messages":
            cur.execute(f"""
                CREATE VIRTUAL TABLE {shadow_name} USING fts5(
                    body, username,
                    content='chat_messages',
                    content_rowid='id'
                )
            """)
            
            cur.execute(f"""
                INSERT INTO {shadow_name}(rowid, body, username)
                SELECT 
                    m.id,
                    COALESCE(m.body, ''),
                    COALESCE(u.username, '')
                FROM chat_messages m
                LEFT JOIN users u ON m.user_id = u.id
            """)
        
        # Step 2: Atomic swap (requires EXCLUSIVE lock but brief)
        cur.execute("BEGIN EXCLUSIVE")
        try:
            cur.execute(f"DROP TABLE IF EXISTS {table_name}")
            cur.execute(f"ALTER TABLE {shadow_name} RENAME TO {table_name}")
            con.commit()
            log("INFO", f"FTS table {table_name} rebuilt")
        except Exception:
            con.rollback()
            raise
    
    except sqlite3.Error as e:
        log("ERROR", f"FTS rebuild failed for {table_name}: {e}")
        raise


def search_fts(table_name: str, query: str, limit: int = 50) -> list[dict]:
    """
    Search FTS table.
    Returns list of matching rowids with highlighted snippets.
    """
    safe_query = fts_sanitize(query)
    if not safe_query:
        return []
    
    try:
        rows = query_db(
            f"""
            SELECT rowid, snippet({table_name}, 0, '<mark>', '</mark>', '...', 32) as snippet
            FROM {table_name}
            WHERE {table_name} MATCH ?
            LIMIT ?
            """,
            (safe_query, limit)
        )
        return [dict(r) for r in rows]
    except sqlite3.Error as e:
        log("ERROR", f"FTS search failed: {e}")
        return []


# ===== BLOCK: EMAIL HELPERS =====

def send_email(to: str, subject: str, body: str, html: bool = False) -> bool:
    """
    Send email via SMTP.
    Returns True on success.
    """
    if not SMTP_HOST:
        log("WARN", "SMTP not configured, email not sent")
        return False
    
    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = SMTP_FROM
        msg['To'] = to
        msg['Subject'] = subject
        
        if html:
            msg.attach(MIMEText(body, 'html', 'utf-8'))
        else:
            msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as server:
            server.starttls()
            if SMTP_USER and SMTP_PASS:
                server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
        
        log("INFO", f"Email sent to {to}")
        return True
    
    except Exception as e:
        log("ERROR", f"Email send failed: {e}")
        return False


def fetch_emails_imap(account_config: dict) -> list[dict]:
    """
    Fetch new emails from IMAP account.
    ✅ УЛУЧШЕНО: добавлена проверка UIDVALIDITY
    
    Returns list of email dicts: {uid, from, subject, body, date}
    """
    host = account_config.get("host")
    port = account_config.get("port", 993)
    user = account_config.get("user")
    password = account_config.get("password")
    folder = account_config.get("folder", "INBOX")
    last_uid = int(account_config.get("last_uid", 0))
    last_uidvalidity = account_config.get("last_uidvalidity")
    
    if not all([host, user, password]):
        return []
    
    try:
        M = imaplib.IMAP4_SSL(host, port, timeout=30)
        M.login(user, password)
        M.select(folder)
        
        # ✅ НОВОЕ: Check UIDVALIDITY to detect mailbox rebuild
        status, data = M.status(folder, "(UIDVALIDITY)")
        if status != "OK":
            log("WARN", "Could not get UIDVALIDITY")
        else:
            # Parse UIDVALIDITY from response
            match = re.search(r'UIDVALIDITY\s+(\d+)', data[0].decode())
            if match:
                current_uidvalidity = match.group(1)
                
                # If UIDVALIDITY changed, reset last_uid
                if last_uidvalidity and last_uidvalidity != current_uidvalidity:
                    log("WARN", f"UIDVALIDITY changed ({last_uidvalidity} -> {current_uidvalidity}), resetting last_uid")
                    last_uid = 0
                    account_config["last_uidvalidity"] = current_uidvalidity
                elif not last_uidvalidity:
                    account_config["last_uidvalidity"] = current_uidvalidity
        
        # Search for new messages
        typ, data = M.uid("search", None, f"UID {last_uid + 1}:*")
        if typ != "OK":
            return []
        
        uids = data[0].split()
        if not uids:
            return []
        
        emails = []
        for uid_bytes in uids[:100]:  # Limit to 100 per fetch
            uid = uid_bytes.decode()
            
            typ, msg_data = M.uid("fetch", uid, "(RFC822)")
            if typ != "OK":
                continue
            
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            
            # Extract fields
            from_addr = msg.get("From", "")
            subject = msg.get("Subject", "")
            date = msg.get("Date", "")
            
            # Extract body
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body = part.get_payload(decode=True).decode(errors="ignore")
                        break
            else:
                body = msg.get_payload(decode=True).decode(errors="ignore")
            
            emails.append({
                "uid": int(uid),
                "from": from_addr,
                "subject": subject,
                "body": body,
                "date": date,
            })
        
        M.close()
        M.logout()
        
        # Update last_uid in account_config
        if emails:
            account_config["last_uid"] = max(e["uid"] for e in emails)
        
        return emails
    
    except Exception as e:
        log("ERROR", f"IMAP fetch failed: {e}")
        return []


# ===== BLOCK: PHONE UTILITIES =====

def extract_phones_from_text(text: str) -> list[str]:
    """
    Extract phone numbers from text.
    Returns list of E.164 normalized phones.
    """
    if not text:
        return []
    
    # Regex for phone patterns
    patterns = [
        r'\+?\d[\d\s\-KATEX_INLINE_OPENKATEX_INLINE_CLOSE]{7,}\d',  # General international
        r'8\s?KATEX_INLINE_OPEN?\d{3}KATEX_INLINE_CLOSE?\s?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}',  # Russian 8(XXX)XXX-XX-XX
    ]
    
    phones = []
    for pattern in patterns:
        matches = re.findall(pattern, text)
        for match in matches:
            normalized = normalize_phone(match)
            if normalized and normalized not in phones:
                phones.append(normalized)
    
    return phones


# ===== BLOCK: NETWORK UTILITIES =====

def _resolve_all_ips(hostname: str) -> list[str]:
    """
    Resolve hostname to all IP addresses.
    Used for SSRF protection.
    """
    try:
        results = socket.getaddrinfo(hostname, None)
        ips = list(set(r[4][0] for r in results))
        return ips
    except socket.gaierror:
        return []


def _is_private_ip(ip: str) -> bool:
    """
    Check if IP is private/internal.
    Used for SSRF protection.
    """
    try:
        ip_obj = ipaddress.ip_address(ip)
        return ip_obj.is_private or ip_obj.is_loopback or ip_obj.is_link_local
    except ValueError:
        return True  # Fail-safe: treat invalid as private


def _resolve_and_validate_url(url: str) -> tuple[str, bool]:
    """
    Resolve URL and validate IP is not private.
    Returns: (resolved_url_with_ip, is_safe)
    ✅ НОВОЕ: для SSRF mitigation в CTI recording
    """
    parsed = urllib.parse.urlparse(url)
    host = parsed.hostname
    
    if not host:
        return url, False
    
    # Resolve once
    ips = _resolve_all_ips(host)
    if not ips:
        return url, False
    
    # Check all IPs
    for ip in ips:
        if _is_private_ip(ip):
            return url, False
    
    # Use first IP (prevent DNS rebinding)
    # Rebuild URL with IP
    safe_url = url.replace(f"//{host}", f"//{ips[0]}")
    return safe_url, True


# ===== BLOCK: AUDIT LOGGING =====

def add_audit(
    org_id: int,
    user_id: int | None,
    action: str,
    entity_type: str | None = None,
    entity_id: int | None = None,
    details: dict | None = None,
    ip_address: str | None = None,
) -> int | None:
    """
    Add entry to audit_logs table.
    ✅ ИСПРАВЛЕНО: закрыт except block
    """
    try:
        return exec_db(
            """
            INSERT INTO audit_logs (org_id, user_id, action, entity_type, entity_id, details, ip_address, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                org_id,
                user_id,
                action,
                entity_type,
                entity_id,
                json.dumps(details) if details else None,
                ip_address,
                utc_now(),
            ),
        )
    except Exception as e:
        # Audit failure should not break main flow
        log("ERROR", f"Audit log failed: {e}")
        pass
    return None


# ===== BLOCK: AI PROVIDER ABSTRACTION =====

def ai_provider_call(
    prompt: str,
    system: str = "",
    model: str = None,
    temperature: float = None,
    max_tokens: int = None,
) -> str:
    """
    Call AI provider (OpenAI/Anthropic/custom).
    Returns generated text.
    """
    model = model or AI_MODEL
    temperature = temperature if temperature is not None else AI_TEMPERATURE
    max_tokens = max_tokens or AI_MAX_TOKENS
    
    if not AI_API_KEY:
        raise ValueError("AI_API_KEY not configured")
    
    # Track metric
    _increment_metric("api_calls_total", {"provider": AI_PROVIDER, "model": model})
    
    if AI_PROVIDER == "openai":
        return _ai_openai_call(prompt, system, model, temperature, max_tokens)
    elif AI_PROVIDER == "anthropic":
        return _ai_anthropic_call(prompt, system, model, temperature, max_tokens)
    else:
        raise ValueError(f"Unknown AI provider: {AI_PROVIDER}")


def _ai_openai_call(prompt: str, system: str, model: str, temperature: float, max_tokens: int) -> str:
    """Call OpenAI API"""
    messages = []
    if system:
        messages.append({"role": "system", "content": system})
    messages.append({"role": "user", "content": prompt})
    
    response = _rq.post(
        "https://api.openai.com/v1/chat/completions",
        headers={
            "Authorization": f"Bearer {AI_API_KEY}",
            "Content-Type": "application/json",
        },
        json={
            "model": model,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": max_tokens,
        },
        timeout=60,
    )
    
    if response.status_code != 200:
        raise RuntimeError(f"OpenAI API error: {response.status_code} {response.text}")
    
    data = response.json()
    return data["choices"][0]["message"]["content"]


def _ai_anthropic_call(prompt: str, system: str, model: str, temperature: float, max_tokens: int) -> str:
    """Call Anthropic API"""
    response = _rq.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": AI_API_KEY,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json",
        },
        json={
            "model": model,
            "max_tokens": max_tokens,
            "temperature": temperature,
            "system": system or "You are a helpful assistant.",
            "messages": [{"role": "user", "content": prompt}],
        },
        timeout=60,
    )
    
    if response.status_code != 200:
        raise RuntimeError(f"Anthropic API error: {response.status_code} {response.text}")
    
    data = response.json()
    return data["content"][0]["text"]


# ===== BLOCK: CONTEXT MANAGEMENT FOR AI =====

def truncate_for_ai_context(text: str, max_chars: int = 8000) -> str:
    """
    Truncate text to fit in AI context window.
    ✅ УЛУЧШЕНО: будущая интеграция с tiktoken для точного подсчета токенов
    """
    if len(text) <= max_chars:
        return text
    
    # Simple truncation (can be improved with tiktoken)
    return text[:max_chars] + "\n\n[... truncated ...]"


# ===== END OF CORE PART 3/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 4/10 — BUSINESS LOGIC: TASKS, DEALS, COMPANIES, INBOX
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: BUSINESS LOGIC: TASKS =====

def create_task(
    org_id: int,
    title: str,
    description: str = "",
    assignee_id: int | None = None,
    status: str = "open",
    priority: str = "normal",
    due_at: str | None = None,
    company_id: int | None = None,
    contact_id: int | None = None,
    monthly_fee: float = 0.0,
    address: str = "",
    contact_phone: str = "",
) -> int:
    """
    Create new task.
    Returns task ID.
    """
    task_id = exec_db(
        """
        INSERT INTO tasks (
            org_id, title, description, assignee_id, status, priority,
            due_at, company_id, contact_id, monthly_fee, address, contact_phone,
            created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            org_id,
            title,
            description,
            assignee_id,
            status,
            priority,
            ensure_iso_datetime(due_at) if due_at else None,
            company_id,
            contact_id,
            monthly_fee,
            address,
            contact_phone,
            utc_now(),
            utc_now(),
        )
    )
    
    # Add activity log
    exec_db(
        "INSERT INTO task_activity (task_id, user_id, kind, created_at) VALUES (?, ?, ?, ?)",
        (task_id, assignee_id, "created", utc_now())
    )
    
    return task_id


def update_task(task_id: int, org_id: int, updates: dict) -> bool:
    """
    Update task fields.
    Returns True on success.
    """
    allowed_fields = {
        "title", "description", "assignee_id", "status", "priority",
        "due_at", "company_id", "contact_id", "monthly_fee", "address", "contact_phone"
    }
    
    # Filter allowed fields
    filtered = {k: v for k, v in updates.items() if k in allowed_fields}
    if not filtered:
        return False
    
    # Normalize datetime fields
    if "due_at" in filtered:
        filtered["due_at"] = ensure_iso_datetime(filtered["due_at"]) if filtered["due_at"] else None
    
    # Add updated_at
    filtered["updated_at"] = utc_now()
    
    # Build SQL
    set_clause = ", ".join(f"{k}=?" for k in filtered.keys())
    values = list(filtered.values())
    
    exec_db(
        f"UPDATE tasks SET {set_clause} WHERE id=? AND org_id=?",
        (*values, task_id, org_id)
    )
    
    # Log activity
    if "status" in updates:
        exec_db(
            "INSERT INTO task_activity (task_id, kind, details_json, created_at) VALUES (?, ?, ?, ?)",
            (task_id, "status_changed", json.dumps({"to": updates["status"]}), utc_now())
        )
    
    return True


def toggle_task_status(task_id: int, org_id: int) -> str:
    """
    Toggle task between 'open' and 'done'.
    Returns new status.
    """
    task = query_db("SELECT status FROM tasks WHERE id=? AND org_id=?", (task_id, org_id), one=True)
    if not task:
        return "open"
    
    new_status = "done" if task["status"] != "done" else "open"
    completed_at = utc_now() if new_status == "done" else None
    
    exec_db(
        "UPDATE tasks SET status=?, completed_at=?, updated_at=? WHERE id=? AND org_id=?",
        (new_status, completed_at, utc_now(), task_id, org_id)
    )
    
    return new_status


def add_task_comment(task_id: int, user_id: int, body: str, format: str = "plain", attachments: list = None) -> int:
    """
    Add comment to task.
    Returns comment ID.
    """
    # Sanitize HTML if format is HTML
    if format == "html":
        body = sanitize_html(body)
    
    comment_id = exec_db(
        "INSERT INTO task_comments (task_id, user_id, body, format, created_at) VALUES (?, ?, ?, ?, ?)",
        (task_id, user_id, body, format, utc_now())
    )
    
    # Add attachments
    if attachments:
        for att in attachments:
            exec_db(
                "INSERT INTO task_comment_attachments (comment_id, file_id, created_at) VALUES (?, ?, ?)",
                (comment_id, att["file_id"], utc_now())
            )
    
    # Update last_commented_at
    exec_db(
        "UPDATE tasks SET last_commented_at=?, updated_at=? WHERE id=?",
        (utc_now(), utc_now(), task_id)
    )
    
    return comment_id


def get_task_comments(task_id: int, page: int = 1, per_page: int = 50) -> list[dict]:
    """Get task comments with attachments"""
    offset = (page - 1) * per_page
    
    comments = query_db(
        """
        SELECT c.*, u.username
        FROM task_comments c
        LEFT JOIN users u ON c.user_id = u.id
        WHERE c.task_id = ?
        ORDER BY c.created_at DESC
        LIMIT ? OFFSET ?
        """,
        (task_id, per_page, offset)
    )
    
    result = []
    for comment in comments:
        c = dict(comment)
        
        # Get attachments
        attachments = query_db(
            """
            SELECT f.id, f.name, f.storage_key, f.content_type, f.size_bytes
            FROM task_comment_attachments a
            JOIN files f ON a.file_id = f.id
            WHERE a.comment_id = ?
            """,
            (c["id"],)
        )
        
        c["attachments"] = [
            {
                "id": a["id"],
                "name": a["name"],
                "url": f"/api/files/{a['id']}/download",
                "size": a["size_bytes"]
            }
            for a in attachments
        ]
        
        result.append(c)
    
    return result


def update_task_participants(task_id: int, add: list[int], remove: list[int], role: str = "watcher"):
    """
    Update task participants.
    add/remove are lists of user IDs.
    """
    for user_id in remove:
        exec_db(
            "DELETE FROM task_participants WHERE task_id=? AND user_id=?",
            (task_id, user_id)
        )
    
    for user_id in add:
        try:
            exec_db(
                "INSERT INTO task_participants (task_id, user_id, role, created_at) VALUES (?, ?, ?, ?)",
                (task_id, user_id, role, utc_now())
            )
        except sqlite3.IntegrityError:
            # Already exists, update role
            exec_db(
                "UPDATE task_participants SET role=? WHERE task_id=? AND user_id=?",
                (role, task_id, user_id)
            )


def update_task_checklist(task_id: int, items: list[dict]):
    """
    Update task checklist.
    items: [{"id": int|None, "item": str, "checked": bool, "sort_order": int}]
    """
    # Delete removed items
    existing_ids = [i["id"] for i in items if i.get("id")]
    if existing_ids:
        placeholders = ",".join("?" * len(existing_ids))
        exec_db(
            f"DELETE FROM task_checklists WHERE task_id=? AND id NOT IN ({placeholders})",
            (task_id, *existing_ids)
        )
    else:
        exec_db("DELETE FROM task_checklists WHERE task_id=?", (task_id,))
    
    # Insert/update items
    for item_data in items:
        if item_data.get("id"):
            # Update existing
            exec_db(
                "UPDATE task_checklists SET item=?, checked=?, sort_order=? WHERE id=?",
                (item_data["item"], int(item_data.get("checked", False)), item_data.get("sort_order", 0), item_data["id"])
            )
        else:
            # Insert new
            exec_db(
                "INSERT INTO task_checklists (task_id, item, checked, sort_order, created_at) VALUES (?, ?, ?, ?, ?)",
                (task_id, item_data["item"], int(item_data.get("checked", False)), item_data.get("sort_order", 0), utc_now())
            )


def add_task_reminder(task_id: int, user_id: int, remind_at: str, message: str = ""):
    """Add reminder for task"""
    return exec_db(
        "INSERT INTO task_reminders (task_id, user_id, remind_at, message, created_at) VALUES (?, ?, ?, ?, ?)",
        (task_id, user_id, ensure_iso_datetime(remind_at), message, utc_now())
    )


def pin_file_to_task(task_id: int, file_id: int, pinned: bool = True):
    """Pin or unpin file from task"""
    if pinned:
        try:
            exec_db(
                "INSERT INTO task_files (task_id, file_id, pinned, created_at) VALUES (?, ?, ?, ?)",
                (task_id, file_id, 1, utc_now())
            )
        except sqlite3.IntegrityError:
            # Already exists, just update pinned status
            exec_db(
                "UPDATE task_files SET pinned=? WHERE task_id=? AND file_id=?",
                (1, task_id, file_id)
            )
    else:
        exec_db(
            "UPDATE task_files SET pinned=? WHERE task_id=? AND file_id=?",
            (0, task_id, file_id)
        )


def batch_update_tasks(org_id: int, task_ids: list[int], updates: dict) -> int:
    """
    Batch update multiple tasks.
    ✅ НОВОЕ: для массовых операций
    Returns count of updated tasks.
    """
    if not task_ids or len(task_ids) > 100:
        return 0
    
    # Validate ownership
    placeholders = ",".join("?" * len(task_ids))
    valid_tasks = query_db(
        f"SELECT id FROM tasks WHERE id IN ({placeholders}) AND org_id=?",
        (*task_ids, org_id)
    )
    valid_ids = [t["id"] for t in valid_tasks]
    
    if not valid_ids:
        return 0
    
    # Apply updates to each task
    count = 0
    for task_id in valid_ids:
        if update_task(task_id, org_id, updates):
            count += 1
    
    return count


# ===== BLOCK: BUSINESS LOGIC: DEALS =====

def create_deal(
    org_id: int,
    title: str,
    amount: float = 0.0,
    currency: str = "RUB",
    status: str = "open",
    stage: str = "new",
    pipeline_key: str = "default",
    assignee_id: int | None = None,
    company_id: int | None = None,
    contact_id: int | None = None,
    due_at: str | None = None,
) -> int:
    """
    Create new deal.
    Returns deal ID.
    """
    deal_id = exec_db(
        """
        INSERT INTO deals (
            org_id, title, amount, currency, status, stage, pipeline_key,
            assignee_id, company_id, contact_id, due_at, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            org_id, title, amount, currency, status, stage, pipeline_key,
            assignee_id, company_id, contact_id,
            ensure_iso_datetime(due_at) if due_at else None,
            utc_now(), utc_now()
        )
    )
    
    # Calculate initial score
    recalculate_score("deal", deal_id, org_id)
    
    return deal_id


def update_deal(deal_id: int, org_id: int, updates: dict) -> bool:
    """
    Update deal fields.
    Returns True on success.
    """
    allowed_fields = {
        "title", "amount", "currency", "status", "stage", "pipeline_key",
        "assignee_id", "company_id", "contact_id", "due_at"
    }
    
    filtered = {k: v for k, v in updates.items() if k in allowed_fields}
    if not filtered:
        return False
    
    # Normalize datetime
    if "due_at" in filtered:
        filtered["due_at"] = ensure_iso_datetime(filtered["due_at"]) if filtered["due_at"] else None
    
    # Handle status transitions
    if "status" in filtered:
        if filtered["status"] == "won":
            filtered["won_at"] = utc_now()
        elif filtered["status"] == "lost":
            filtered["lost_at"] = utc_now()
    
    filtered["updated_at"] = utc_now()
    
    # Build SQL
    set_clause = ", ".join(f"{k}=?" for k in filtered.keys())
    values = list(filtered.values())
    
    exec_db(
        f"UPDATE deals SET {set_clause} WHERE id=? AND org_id=?",
        (*values, deal_id, org_id)
    )
    
    # Log stage transition
    if "stage" in updates:
        old_deal = query_db("SELECT stage FROM deals WHERE id=?", (deal_id,), one=True)
        if old_deal:
            exec_db(
                """
                INSERT INTO stage_transitions (entity_type, entity_id, from_stage, to_stage, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                ("deal", deal_id, old_deal["stage"], updates["stage"], utc_now())
            )
    
    # Recalculate score if relevant fields changed
    if any(k in updates for k in ["amount", "stage", "company_id"]):
        recalculate_score("deal", deal_id, org_id)
    
    return True


def recalculate_score(entity_type: str, entity_id: int, org_id: int):
    """
    Recalculate lead score based on scoring rules.
    ✅ НОВОЕ: автоматический расчет score для deals/companies
    """
    # Get scoring rules
    rules = query_db(
        """
        SELECT * FROM lead_scoring_rules
        WHERE org_id=? AND entity_type=? AND active=1
        ORDER BY id
        """,
        (org_id, entity_type)
    )
    
    if not rules:
        return
    
    # Get entity data
    if entity_type == "deal":
        entity = query_db("SELECT * FROM deals WHERE id=?", (entity_id,), one=True)
    elif entity_type == "company":
        entity = query_db("SELECT * FROM companies WHERE id=?", (entity_id,), one=True)
    else:
        return
    
    if not entity:
        return
    
    total_score = 0
    
    for rule in rules:
        field_value = entity.get(rule["field"])
        rule_value = rule["value"]
        operator = rule["operator"]
        
        matched = False
        
        try:
            if operator == "==":
                matched = str(field_value) == str(rule_value)
            elif operator == "!=":
                matched = str(field_value) != str(rule_value)
            elif operator == ">":
                matched = float(field_value or 0) > float(rule_value)
            elif operator == ">=":
                matched = float(field_value or 0) >= float(rule_value)
            elif operator == "<":
                matched = float(field_value or 0) < float(rule_value)
            elif operator == "<=":
                matched = float(field_value or 0) <= float(rule_value)
            elif operator == "contains":
                matched = str(rule_value).lower() in str(field_value or "").lower()
            elif operator == "starts_with":
                matched = str(field_value or "").lower().startswith(str(rule_value).lower())
        except (ValueError, TypeError):
            pass
        
        if matched:
            total_score += rule["score_delta"]
    
    # Update score
    table = f"{entity_type}s"
    exec_db(f"UPDATE {table} SET score=? WHERE id=?", (total_score, entity_id))


# ===== BLOCK: BUSINESS LOGIC: COMPANIES & CONTACTS =====

def create_company(
    org_id: int,
    name: str,
    inn: str = "",
    phone: str = "",
    email: str = "",
    address: str = "",
    notes: str = "",
) -> int:
    """
    Create new company.
    Returns company ID.
    """
    phone_norm = normalize_phone(phone) if phone else ""
    
    company_id = exec_db(
        """
        INSERT INTO companies (org_id, name, inn, phone, phone_norm, email, address, notes, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (org_id, name, inn, phone, phone_norm, email, address, notes, utc_now(), utc_now())
    )
    
    # Calculate initial score
    recalculate_score("company", company_id, org_id)
    
    return company_id


def update_company(company_id: int, org_id: int, updates: dict) -> bool:
    """Update company fields"""
    allowed_fields = {"name", "inn", "phone", "email", "address", "notes"}
    
    filtered = {k: v for k, v in updates.items() if k in allowed_fields}
    if not filtered:
        return False
    
    # Normalize phone
    if "phone" in filtered:
        filtered["phone_norm"] = normalize_phone(filtered["phone"]) if filtered["phone"] else ""
    
    filtered["updated_at"] = utc_now()
    
    set_clause = ", ".join(f"{k}=?" for k in filtered.keys())
    values = list(filtered.values())
    
    exec_db(
        f"UPDATE companies SET {set_clause} WHERE id=? AND org_id=?",
        (*values, company_id, org_id)
    )
    
    # Recalculate score
    recalculate_score("company", company_id, org_id)
    
    return True


def create_contact(
    org_id: int,
    name: str,
    company_id: int | None = None,
    position: str = "",
    phone: str = "",
    email: str = "",
    notes: str = "",
) -> int:
    """Create new contact"""
    phone_norm = normalize_phone(phone) if phone else ""
    
    return exec_db(
        """
        INSERT INTO contacts (org_id, company_id, name, position, phone, phone_norm, email, notes, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (org_id, company_id, name, position, phone, phone_norm, email, notes, utc_now(), utc_now())
    )


def lookup_by_phone(org_id: int, phone: str) -> dict:
    """
    Lookup companies and contacts by phone.
    Returns dict with companies and contacts lists.
    """
    phone_norm = normalize_phone(phone)
    if not phone_norm:
        return {"companies": [], "contacts": []}
    
    # Fuzzy pattern (remove +, keep digits)
    pattern = "%" + phone_norm.replace("+", "") + "%"
    
    companies = query_db(
        """
        SELECT * FROM companies
        WHERE org_id=? AND (phone_norm LIKE ? OR phone LIKE ?)
        LIMIT 10
        """,
        (org_id, pattern, pattern)
    )
    
    contacts = query_db(
        """
        SELECT * FROM contacts
        WHERE org_id=? AND (phone_norm LIKE ? OR phone LIKE ?)
        LIMIT 10
        """,
        (org_id, pattern, pattern)
    )
    
    return {
        "companies": [dict(c) for c in companies],
        "contacts": [dict(c) for c in contacts]
    }


def lookup_by_inn(org_id: int, inn: str) -> list[dict]:
    """Lookup companies by INN"""
    if not validate_inn(inn):
        return []
    
    companies = query_db(
        "SELECT * FROM companies WHERE org_id=? AND inn=? LIMIT 10",
        (org_id, inn)
    )
    
    return [dict(c) for c in companies]


def lookup_by_email(org_id: int, email: str) -> dict:
    """Lookup companies and contacts by email"""
    if not validate_email(email):
        return {"companies": [], "contacts": []}
    
    companies = query_db(
        "SELECT * FROM companies WHERE org_id=? AND email=? LIMIT 10",
        (org_id, email)
    )
    
    contacts = query_db(
        "SELECT * FROM contacts WHERE org_id=? AND email=? LIMIT 10",
        (org_id, email)
    )
    
    return {
        "companies": [dict(c) for c in companies],
        "contacts": [dict(c) for c in contacts]
    }


# ===== BLOCK: BUSINESS LOGIC: INBOX =====

def create_inbox_thread(
    org_id: int,
    channel_id: int | None,
    subject: str = "",
    external_thread_id: str = "",
    company_id: int | None = None,
    contact_id: int | None = None,
) -> int:
    """
    Create new inbox thread.
    Returns thread ID.
    """
    return exec_db(
        """
        INSERT INTO inbox_threads (
            org_id, channel_id, subject, status, priority, external_thread_id,
            company_id, contact_id, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (org_id, channel_id, subject, "open", "normal", external_thread_id, company_id, contact_id, utc_now())
    )


def add_message(
    thread_id: int,
    sender_type: str,
    body: str,
    user_id: int | None = None,
    external_user_id: str = "",
    internal_note: bool = False,
    attachments: list = None,
) -> int | None:
    """
    Add message to thread.
    Returns message ID or None if duplicate.
    """
    # Anti-duplicate check (within 60 seconds)
    recent = query_db(
        """
        SELECT id FROM inbox_messages
        WHERE thread_id=? AND body=? AND created_at > datetime('now', '-60 seconds')
        LIMIT 1
        """,
        (thread_id, body),
        one=True
    )
    
    if recent:
        return None
    
    message_id = exec_db(
        """
        INSERT INTO inbox_messages (thread_id, sender_type, user_id, external_user_id, body, internal_note, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (thread_id, sender_type, user_id, external_user_id, body, int(internal_note), utc_now())
    )
    
    # Add attachments
    if attachments:
        for att in attachments:
            exec_db(
                "INSERT INTO message_attachments (message_id, file_id, created_at) VALUES (?, ?, ?)",
                (message_id, att["file_id"], utc_now())
            )
    
    # Update thread last_message_at
    exec_db(
        "UPDATE inbox_threads SET last_message_at=? WHERE id=?",
        (utc_now(), thread_id)
    )
    
    # Update first_response_at if this is first agent reply
    if sender_type == "agent" and not internal_note:
        thread = query_db("SELECT first_response_at FROM inbox_threads WHERE id=?", (thread_id,), one=True)
        if thread and not thread["first_response_at"]:
            exec_db(
                "UPDATE inbox_threads SET first_response_at=? WHERE id=?",
                (utc_now(), thread_id)
            )
    
    return message_id


def update_inbox_thread(thread_id: int, org_id: int, updates: dict) -> bool:
    """Update inbox thread fields"""
    allowed_fields = {"subject", "status", "priority", "assignee_id", "company_id", "contact_id", "tags_csv"}
    
    filtered = {k: v for k, v in updates.items() if k in allowed_fields}
    if not filtered:
        return False
    
    set_clause = ", ".join(f"{k}=?" for k in filtered.keys())
    values = list(filtered.values())
    
    exec_db(
        f"UPDATE inbox_threads SET {set_clause} WHERE id=? AND org_id=?",
        (*values, thread_id, org_id)
    )
    
    return True


# ===== BLOCK: WORKFLOW EVALUATION HELPERS =====

def _wf_eval_value(value, ctx: dict, payload: dict):
    """
    Evaluate Jinja2-style {{ expressions }} in workflow node config.
    ✅ НОВОЕ: дописана функция для workflow engine
    
    Examples:
        {{ ctx.company_id }} → ctx.get("company_id")
        {{ payload.amount }} → payload.get("amount")
        Static value → return as-is
    """
    if not isinstance(value, str):
        return value
    
    # Simple template evaluation (без полного Jinja2 из-за безопасности)
    pattern = r'\{\{\s*([a-zA-Z0-9_\.]+)\s*\}\}'
    
    def replacer(m):
        path = m.group(1).split('.')
        obj = {'ctx': ctx, 'payload': payload}
        try:
            for key in path:
                obj = obj[key]
            return str(obj)
        except (KeyError, TypeError):
            return ''
    
    result = re.sub(pattern, replacer, value)
    
    # Try to parse as JSON if looks like dict/list
    if result.startswith('{') or result.startswith('['):
        try:
            return json.loads(result)
        except json.JSONDecodeError:
            pass
    
    return result


def _wf_eval_condition(condition: str, ctx: dict, payload: dict) -> bool:
    """
    Evaluate workflow condition.
    Simple expressions: field == value, field > value, etc.
    """
    if not condition:
        return True
    
    # Parse condition (simple format: "field operator value")
    match = re.match(r'(\S+)\s*(==|!=|>|<|>=|<=|contains)\s*(.+)', condition.strip())
    if not match:
        return True  # Invalid condition = pass
    
    field, operator, value = match.groups()
    
    # Get field value from context
    field_value = _wf_eval_value(f"{{{{ {field} }}}}", ctx, payload)
    value = value.strip().strip('"\'')
    
    try:
        if operator == "==":
            return str(field_value) == value
        elif operator == "!=":
            return str(field_value) != value
        elif operator == ">":
            return float(field_value) > float(value)
        elif operator == "<":
            return float(field_value) < float(value)
        elif operator == ">=":
            return float(field_value) >= float(value)
        elif operator == "<=":
            return float(field_value) <= float(value)
        elif operator == "contains":
            return value.lower() in str(field_value).lower()
    except (ValueError, TypeError):
        pass
    
    return False


# ===== END OF CORE PART 4/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 5/10 — API ROUTES: AUTH, PROFILE, TASKS, DEALS
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: API ROUTES — AUTHENTICATION =====

@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page and handler"""
    if request.method == "GET":
        if _get_current_user():
            return redirect(url_for("index"))
        return render_template_string(LOGIN_TMPL, error="")
    
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    totp_token = request.form.get("totp_token", "").strip()
    
    if not username or not password:
        return render_template_string(LOGIN_TMPL, error="Username and password required")
    
    # Find user
    user = query_db(
        "SELECT * FROM users WHERE username=? AND active=1",
        (username,),
        one=True
    )
    
    if not user or not verify_password(password, user["password_hash"]):
        return render_template_string(LOGIN_TMPL, error="Invalid credentials")
    
    # Check 2FA if enabled
    if user["totp_enabled"]:
        if not totp_token:
            return render_template_string(LOGIN_TMPL, error="2FA token required", show_2fa=True)
        
        if not verify_totp(user["totp_secret"], totp_token):
            return render_template_string(LOGIN_TMPL, error="Invalid 2FA token", show_2fa=True)
    
    # Create session
    session["user_id"] = user["id"]
    session.permanent = True
    
    # Update last login
    exec_db("UPDATE users SET last_login_at=? WHERE id=?", (utc_now(), user["id"]))
    
    # Audit log
    add_audit(
        user["org_id"],
        user["id"],
        "user.login",
        ip_address=request.remote_addr
    )
    
    return redirect(url_for("index"))


@app.route("/logout", methods=["POST"])
def logout():
    """Logout handler"""
    user = _get_current_user()
    if user:
        add_audit(
            user["org_id"],
            user["id"],
            "user.logout",
            ip_address=request.remote_addr
        )
    
    session.clear()
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    """Registration page (optional - can be disabled in production)"""
    if request.method == "GET":
        return render_template_string(REGISTER_TMPL, error="")
    
    username = request.form.get("username", "").strip()
    email = request.form.get("email", "").strip()
    password = request.form.get("password", "").strip()
    org_slug = request.form.get("org_slug", "").strip()
    
    # Validation
    if not username or not password or not org_slug:
        return render_template_string(REGISTER_TMPL, error="All fields required")
    
    if len(password) < 12:
        return render_template_string(REGISTER_TMPL, error="Password must be at least 12 characters")
    
    # Check if org exists
    org = query_db("SELECT id FROM orgs WHERE slug=?", (org_slug,), one=True)
    if not org:
        # Create new org (optional - may want to disable this)
        org_id = exec_db(
            "INSERT INTO orgs (slug, name, created_at) VALUES (?, ?, ?)",
            (org_slug, org_slug.title(), utc_now())
        )
    else:
        org_id = org["id"]
    
    # Check if username exists in org
    existing = query_db(
        "SELECT id FROM users WHERE org_id=? AND username=?",
        (org_id, username),
        one=True
    )
    if existing:
        return render_template_string(REGISTER_TMPL, error="Username already exists")
    
    # Create user
    password_hash = hash_password(password)
    user_id = exec_db(
        """
        INSERT INTO users (org_id, username, email, password_hash, role, active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (org_id, username, email, password_hash, "agent", 1, utc_now())
    )
    
    # Auto-login
    session["user_id"] = user_id
    
    return redirect(url_for("index"))


# ===== BLOCK: API ROUTES — PROFILE =====

@app.route("/api/profile", methods=["GET"])
@_login_required
def api_profile():
    """Get current user profile"""
    user = g.user
    
    # Get org info
    org = query_db("SELECT * FROM orgs WHERE id=?", (user["org_id"],), one=True)
    
    return jsonify(
        ok=True,
        user={
            "id": user["id"],
            "username": user["username"],
            "email": user["email"],
            "role": user["role"],
            "first_name": user["first_name"],
            "last_name": user["last_name"],
            "phone": user["phone"],
            "avatar_url": user["avatar_url"],
            "timezone": user["timezone"],
            "locale": user["locale"],
            "totp_enabled": bool(user["totp_enabled"]),
        },
        org={
            "id": org["id"],
            "slug": org["slug"],
            "name": org["name"],
        } if org else None
    )


@app.route("/api/profile", methods=["PATCH"])
@_login_required
@_csrf_protect
def api_profile_update():
    """Update user profile"""
    user = g.user
    data = request.get_json() or {}
    
    allowed_fields = {"first_name", "last_name", "email", "phone", "timezone", "locale"}
    updates = {k: v for k, v in data.items() if k in allowed_fields}
    
    if not updates:
        return jsonify(ok=False, error="No valid fields to update"), 400
    
    # Build SQL
    set_clause = ", ".join(f"{k}=?" for k in updates.keys())
    values = list(updates.values())
    
    exec_db(
        f"UPDATE users SET {set_clause} WHERE id=?",
        (*values, user["id"])
    )
    
    return jsonify(ok=True)


@app.route("/api/profile/avatar", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("avatar_upload", per_min=5)
def api_profile_avatar():
    """
    Upload user avatar.
    ✅ УЛУЧШЕНО: добавлена обработка изображения через Pillow
    """
    user = g.user
    
    if "file" not in request.files:
        return jsonify(ok=False, error="No file provided"), 400
    
    f = request.files["file"]
    if not f.filename:
        return jsonify(ok=False, error="Empty filename"), 400
    
    try:
        # Read file
        raw = f.read()
        
        # Process avatar (validate, resize, optimize)
        processed_data, content_type = process_avatar(raw, f.filename)
        
        # Store file
        file_info = store_file(
            user["org_id"],
            f"avatar_{user['id']}_{uuid.uuid4().hex[:8]}.jpg",
            processed_data,
            content_type,
            user["id"]
        )
        
        # Update user avatar_url
        exec_db(
            "UPDATE users SET avatar_url=? WHERE id=?",
            (file_info["url"], user["id"])
        )
        
        return jsonify(ok=True, url=file_info["url"])
    
    except ValueError as e:
        return jsonify(ok=False, error=str(e)), 400
    except Exception as e:
        log("ERROR", f"Avatar upload failed: {e}")
        return jsonify(ok=False, error="Upload failed"), 500


@app.route("/api/profile/password", methods=["POST"])
@_login_required
@_csrf_protect
def api_profile_password():
    """Change user password"""
    user = g.user
    data = request.get_json() or {}
    
    current_password = data.get("current_password", "")
    new_password = data.get("new_password", "")
    
    if not verify_password(current_password, user["password_hash"]):
        return jsonify(ok=False, error="Current password incorrect"), 400
    
    if len(new_password) < 12:
        return jsonify(ok=False, error="Password must be at least 12 characters"), 400
    
    # Update password
    new_hash = hash_password(new_password)
    exec_db("UPDATE users SET password_hash=? WHERE id=?", (new_hash, user["id"]))
    
    add_audit(user["org_id"], user["id"], "user.password_changed")
    
    return jsonify(ok=True)


@app.route("/api/profile/2fa/enable", methods=["POST"])
@_login_required
@_csrf_protect
def api_profile_2fa_enable():
    """Enable 2FA for user"""
    if not PYOTP_AVAILABLE:
        return jsonify(ok=False, error="2FA not available"), 400
    
    user = g.user
    
    # Generate TOTP secret
    secret = generate_totp_secret()
    
    # Generate backup codes
    backup_codes = [generate_random_code(8, "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789") for _ in range(8)]
    
    # Save backup codes (hashed)
    for code in backup_codes:
        code_hash = backup_code_hash(code)
        exec_db(
            "INSERT INTO user_2fa_backup_codes (user_id, code_hash, created_at) VALUES (?, ?, ?)",
            (user["id"], code_hash, utc_now())
        )
    
    # Update user
    exec_db(
        "UPDATE users SET totp_secret=?, totp_enabled=? WHERE id=?",
        (secret, 1, user["id"])
    )
    
    add_audit(user["org_id"], user["id"], "user.2fa_enabled")
    
    # Generate QR code URL
    import pyotp
    totp = pyotp.TOTP(secret)
    qr_url = totp.provisioning_uri(
        name=user["username"],
        issuer_name="CRM"
    )
    
    return jsonify(
        ok=True,
        secret=secret,
        qr_url=qr_url,
        backup_codes=backup_codes
    )


@app.route("/api/profile/2fa/disable", methods=["POST"])
@_login_required
@_csrf_protect
def api_profile_2fa_disable():
    """Disable 2FA"""
    user = g.user
    data = request.get_json() or {}
    
    password = data.get("password", "")
    if not verify_password(password, user["password_hash"]):
        return jsonify(ok=False, error="Password incorrect"), 400
    
    # Disable 2FA
    exec_db(
        "UPDATE users SET totp_enabled=? WHERE id=?",
        (0, user["id"])
    )
    
    # Delete backup codes
    exec_db("DELETE FROM user_2fa_backup_codes WHERE user_id=?", (user["id"],))
    
    add_audit(user["org_id"], user["id"], "user.2fa_disabled")
    
    return jsonify(ok=True)


# ===== BLOCK: API ROUTES — TASKS =====

@app.route("/api/tasks/list", methods=["GET"])
@_login_required
def api_tasks_list():
    """
    Get tasks list with filters.
    Query params: f (filter), page, per_page, q (search), assignee_id, status, etc.
    """
    user = g.user
    org_id = user["org_id"]
    
    # Pagination
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 50)), 100)
    offset = (page - 1) * per_page
    
    # Filters
    f = request.args.get("f", "open")
    q = request.args.get("q", "").strip()
    assignee_id = request.args.get("assignee_id", "")
    status = request.args.get("status", "")
    created_from = request.args.get("created_from", "")
    created_to = request.args.get("created_to", "")
    
    # Build WHERE clause
    where_parts = ["t.org_id = ?"]
    params = [org_id]
    
    # Apply filter presets
    if f == "open":
        where_parts.append("t.status NOT IN ('done', 'cancelled')")
    elif f == "today":
        where_parts.append("date(t.due_at) = date('now')")
    elif f == "overdue":
        where_parts.append("t.due_at < datetime('now') AND t.status NOT IN ('done', 'cancelled')")
    elif f == "done":
        where_parts.append("t.status = 'done'")
    
    # Additional filters
    if assignee_id:
        where_parts.append("t.assignee_id = ?")
        params.append(int(assignee_id))
    
    if status:
        where_parts.append("t.status = ?")
        params.append(status)
    
    if created_from:
        where_parts.append("t.created_at >= ?")
        params.append(ensure_iso_datetime(created_from))
    
    if created_to:
        where_parts.append("t.created_at <= ?")
        params.append(ensure_iso_datetime(created_to))
    
    # Search
    if q:
        where_parts.append("(t.title LIKE ? OR t.description LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%"])
    
    where_clause = " AND ".join(where_parts)
    
    # Query
    tasks = query_db(
        f"""
        SELECT 
            t.*,
            u.username AS assignee_name,
            c.name AS company_name
        FROM tasks t
        LEFT JOIN users u ON t.assignee_id = u.id
        LEFT JOIN companies c ON t.company_id = c.id
        WHERE {where_clause}
        ORDER BY t.due_at ASC, t.created_at DESC
        LIMIT ? OFFSET ?
        """,
        (*params, per_page, offset)
    )
    
    return jsonify(
        ok=True,
        items=[dict(t) for t in tasks],
        page=page,
        per_page=per_page
    )


@app.route("/api/task/create", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_create():
    """Create new task"""
    user = g.user
    data = request.get_json() or {}
    
    title = data.get("title", "").strip()
    if not title:
        return jsonify(ok=False, error="Title required"), 400
    
    task_id = create_task(
        org_id=user["org_id"],
        title=title,
        description=data.get("description", ""),
        assignee_id=data.get("assignee_id"),
        status=data.get("status", "open"),
        priority=data.get("priority", "normal"),
        due_at=data.get("due_at"),
        company_id=data.get("company_id"),
        contact_id=data.get("contact_id"),
        monthly_fee=float(data.get("monthly_fee", 0)),
        address=data.get("address", ""),
        contact_phone=data.get("contact_phone", ""),
    )
    
    add_audit(user["org_id"], user["id"], "task.created", "task", task_id)
    
    return jsonify(ok=True, id=task_id)


@app.route("/api/task/update", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_update():
    """Update task"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("id")
    if not task_id:
        return jsonify(ok=False, error="Task ID required"), 400
    
    # Verify ownership
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    updates = {k: v for k, v in data.items() if k != "id"}
    if update_task(task_id, user["org_id"], updates):
        add_audit(user["org_id"], user["id"], "task.updated", "task", task_id, updates)
        return jsonify(ok=True)
    
    return jsonify(ok=False, error="No valid updates"), 400


@app.route("/api/task/toggle", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_toggle():
    """Toggle task status between open and done"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("id")
    if not task_id:
        return jsonify(ok=False, error="Task ID required"), 400
    
    new_status = toggle_task_status(task_id, user["org_id"])
    
    return jsonify(ok=True, status=new_status)


@app.route("/api/task/comment", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_comment_create():
    """Add comment to task"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    body = data.get("body", "").strip()
    
    if not task_id or not body:
        return jsonify(ok=False, error="Task ID and body required"), 400
    
    # Verify task exists
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    comment_id = add_task_comment(
        task_id,
        user["id"],
        body,
        data.get("format", "plain"),
        data.get("attachments")
    )
    
    return jsonify(ok=True, id=comment_id)


@app.route("/api/task/participants", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_participants_update():
    """Update task participants"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    add = data.get("add", [])
    remove = data.get("remove", [])
    role = data.get("role", "watcher")
    
    if not task_id:
        return jsonify(ok=False, error="Task ID required"), 400
    
    # Verify task
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    update_task_participants(task_id, add, remove, role)
    
    return jsonify(ok=True)


@app.route("/api/task/checklist", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_checklist_update():
    """Update task checklist"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    items = data.get("items", [])
    
    if not task_id:
        return jsonify(ok=False, error="Task ID required"), 400
    
    # Verify task
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    update_task_checklist(task_id, items)
    
    return jsonify(ok=True)


@app.route("/api/tasks/bulk_update", methods=["POST"])
@_login_required
@_csrf_protect
def api_tasks_bulk_update():
    """
    Bulk update tasks.
    ✅ НОВОЕ: массовые операции
    """
    user = g.user
    data = request.get_json() or {}
    
    ids = data.get("ids", [])
    if not ids or len(ids) > 100:
        return jsonify(ok=False, error="Invalid IDs count (max 100)"), 400
    
    updates = {k: v for k, v in data.items() if k not in ("ids", "csrf_token")}
    if not updates:
        return jsonify(ok=False, error="No updates provided"), 400
    
    count = batch_update_tasks(user["org_id"], ids, updates)
    
    add_audit(user["org_id"], user["id"], "tasks.bulk_updated", details={"count": count, "updates": updates})
    
    return jsonify(ok=True, updated=count)


# ===== BLOCK: API ROUTES — DEALS =====

@app.route("/api/deals/list", methods=["GET"])
@_login_required
def api_deals_list():
    """Get deals list with filters"""
    user = g.user
    org_id = user["org_id"]
    
    # Pagination
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 50)), 100)
    offset = (page - 1) * per_page
    
    # Filters
    pipeline = request.args.get("pipeline", "")
    stage = request.args.get("stage", "")
    status = request.args.get("status", "")
    assignee_id = request.args.get("assignee_id", "")
    
    # Build WHERE
    where_parts = ["d.org_id = ?"]
    params = [org_id]
    
    if pipeline:
        where_parts.append("d.pipeline_key = ?")
        params.append(pipeline)
    
    if stage:
        where_parts.append("d.stage = ?")
        params.append(stage)
    
    if status:
        where_parts.append("d.status = ?")
        params.append(status)
    
    if assignee_id:
        where_parts.append("d.assignee_id = ?")
        params.append(int(assignee_id))
    
    where_clause = " AND ".join(where_parts)
    
    deals = query_db(
        f"""
        SELECT 
            d.*,
            u.username AS assignee_name,
            c.name AS company_name
        FROM deals d
        LEFT JOIN users u ON d.assignee_id = u.id
        LEFT JOIN companies c ON d.company_id = c.id
        WHERE {where_clause}
        ORDER BY d.created_at DESC
        LIMIT ? OFFSET ?
        """,
        (*params, per_page, offset)
    )
    
    return jsonify(
        ok=True,
        items=[dict(d) for d in deals],
        page=page,
        per_page=per_page
    )


@app.route("/api/deal/create", methods=["POST"])
@_login_required
@_csrf_protect
def api_deal_create():
    """Create new deal"""
    user = g.user
    data = request.get_json() or {}
    
    title = data.get("title", "").strip()
    if not title:
        return jsonify(ok=False, error="Title required"), 400
    
    deal_id = create_deal(
        org_id=user["org_id"],
        title=title,
        amount=float(data.get("amount", 0)),
        currency=data.get("currency", "RUB"),
        status=data.get("status", "open"),
        stage=data.get("stage", "new"),
        pipeline_key=data.get("pipeline_key", "default"),
        assignee_id=data.get("assignee_id"),
        company_id=data.get("company_id"),
        contact_id=data.get("contact_id"),
        due_at=data.get("due_at"),
    )
    
    add_audit(user["org_id"], user["id"], "deal.created", "deal", deal_id)
    
    return jsonify(ok=True, id=deal_id)


@app.route("/api/deal/update", methods=["POST"])
@_login_required
@_csrf_protect
def api_deal_update():
    """Update deal"""
    user = g.user
    data = request.get_json() or {}
    
    deal_id = data.get("id")
    if not deal_id:
        return jsonify(ok=False, error="Deal ID required"), 400
    
    # Verify ownership
    deal = query_db("SELECT id FROM deals WHERE id=? AND org_id=?", (deal_id, user["org_id"]), one=True)
    if not deal:
        return jsonify(ok=False, error="Deal not found"), 404
    
    updates = {k: v for k, v in data.items() if k != "id"}
    if update_deal(deal_id, user["org_id"], updates):
        add_audit(user["org_id"], user["id"], "deal.updated", "deal", deal_id, updates)
        return jsonify(ok=True)
    
    return jsonify(ok=False, error="No valid updates"), 400


@app.route("/api/deals/kanban", methods=["GET"])
@_login_required
def api_deals_kanban():
    """
    Get deals organized by stages for kanban view.
    Returns: {columns: [stage_keys], items: {stage: [deals]}}
    """
    user = g.user
    org_id = user["org_id"]
    pipeline = request.args.get("pipeline", "default")
    
    # Get stages for pipeline
    stages = query_db(
        """
        SELECT key, name FROM workflow_stages
        WHERE org_id=? AND entity_type='deal' AND pipeline_key=?
        ORDER BY sort_order
        """,
        (org_id, pipeline)
    )
    
    if not stages:
        return jsonify(ok=True, columns=[], items={})
    
    stage_keys = [s["key"] for s in stages]
    
    # Get deals grouped by stage
    deals = query_db(
        """
        SELECT d.*, u.username AS assignee_name
        FROM deals d
        LEFT JOIN users u ON d.assignee_id = u.id
        WHERE d.org_id=? AND d.pipeline_key=? AND d.status='open'
        ORDER BY d.created_at DESC
        """,
        (org_id, pipeline)
    )
    
    # Group by stage
    items = {key: [] for key in stage_keys}
    for deal in deals:
        stage = deal["stage"] or "new"
        if stage in items:
            items[stage].append(dict(deal))
    
    return jsonify(ok=True, columns=stage_keys, items=items)


@app.route("/api/deals/kanban/update", methods=["POST"])
@_login_required
@_csrf_protect
def api_deals_kanban_update():
    """Update deal stage (for kanban drag-and-drop)"""
    user = g.user
    data = request.get_json() or {}
    
    deal_id = data.get("id")
    stage = data.get("stage")
    
    if not deal_id or not stage:
        return jsonify(ok=False, error="Deal ID and stage required"), 400
    
    # Verify deal
    deal = query_db("SELECT id FROM deals WHERE id=? AND org_id=?", (deal_id, user["org_id"]), one=True)
    if not deal:
        return jsonify(ok=False, error="Deal not found"), 404
    
    update_deal(deal_id, user["org_id"], {"stage": stage})
    
    return jsonify(ok=True)


# ===== END OF CORE PART 5/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 6/10 — API ROUTES: COMPANIES, INBOX, CALLS, MEETINGS, CHAT, AI
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: API ROUTES — COMPANIES & CONTACTS =====

@app.route("/api/clients/list", methods=["GET"])
@_login_required
def api_clients_list():
    """Get companies list with search"""
    user = g.user
    org_id = user["org_id"]
    
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 50)), 100)
    offset = (page - 1) * per_page
    
    q = request.args.get("q", "").strip()
    
    # Build WHERE
    where_parts = ["org_id = ?"]
    params = [org_id]
    
    if q:
        where_parts.append("(name LIKE ? OR inn LIKE ? OR phone LIKE ? OR email LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
    
    where_clause = " AND ".join(where_parts)
    
    companies = query_db(
        f"""
        SELECT c.*, 
            (SELECT COUNT(*) FROM deals WHERE company_id=c.id) AS deals
        FROM companies c
        WHERE {where_clause}
        ORDER BY c.created_at DESC
        LIMIT ? OFFSET ?
        """,
        (*params, per_page, offset)
    )
    
    return jsonify(
        ok=True,
        items=[dict(c) for c in companies],
        page=page,
        per_page=per_page
    )


@app.route("/api/clients/<int:client_id>", methods=["PATCH"])
@_login_required
@_csrf_protect
def api_client_update(client_id):
    """Update company"""
    user = g.user
    data = request.get_json() or {}
    
    # Verify ownership
    company = query_db("SELECT id FROM companies WHERE id=? AND org_id=?", (client_id, user["org_id"]), one=True)
    if not company:
        return jsonify(ok=False, error="Company not found"), 404
    
    if update_company(client_id, user["org_id"], data):
        return jsonify(ok=True)
    
    return jsonify(ok=False, error="No valid updates"), 400


@app.route("/api/lookup", methods=["GET"])
@_login_required
def api_lookup():
    """
    Unified lookup API for companies/contacts.
    Query params: id, phone, inn, email
    """
    user = g.user
    org_id = user["org_id"]
    
    company_id = request.args.get("id")
    phone = request.args.get("phone")
    inn = request.args.get("inn")
    email_addr = request.args.get("email")
    
    result = {"companies": [], "contacts": []}
    
    if company_id:
        company = query_db("SELECT * FROM companies WHERE id=? AND org_id=?", (company_id, org_id), one=True)
        if company:
            result["companies"] = [dict(company)]
    
    elif phone:
        result = lookup_by_phone(org_id, phone)
    
    elif inn:
        result["companies"] = lookup_by_inn(org_id, inn)
    
    elif email_addr:
        result = lookup_by_email(org_id, email_addr)
    
    return jsonify(ok=True, **result)


# ===== BLOCK: API ROUTES — INBOX =====

@app.route("/api/inbox/list", methods=["GET"])
@_login_required
def api_inbox_list():
    """Get inbox threads with filters"""
    user = g.user
    org_id = user["org_id"]
    
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 50)), 100)
    offset = (page - 1) * per_page
    
    # Filters
    status = request.args.get("status", "")
    channel = request.args.get("channel", "")
    assignee = request.args.get("assignee", "")
    who = request.args.get("who", "")  # "me" для моих задач
    
    # Build WHERE
    where_parts = ["t.org_id = ?"]
    params = [org_id]
    
    if status:
        where_parts.append("t.status = ?")
        params.append(status)
    
    if channel:
        where_parts.append("t.channel_id = ?")
        params.append(int(channel))
    
    if assignee:
        where_parts.append("t.assignee_id = ?")
        params.append(int(assignee))
    
    if who == "me":
        where_parts.append("t.assignee_id = ?")
        params.append(user["id"])
    
    where_clause = " AND ".join(where_parts)
    
    threads = query_db(
        f"""
        SELECT 
            t.*,
            c.name AS channel_name,
            u.username AS assignee_name
        FROM inbox_threads t
        LEFT JOIN channels c ON t.channel_id = c.id
        LEFT JOIN users u ON t.assignee_id = u.id
        WHERE {where_clause}
        ORDER BY t.last_message_at DESC
        LIMIT ? OFFSET ?
        """,
        (*params, per_page, offset)
    )
    
    return jsonify(
        ok=True,
        items=[dict(t) for t in threads],
        page=page,
        per_page=per_page
    )


@app.route("/api/thread/update", methods=["POST"])
@_login_required
@_csrf_protect
def api_thread_update():
    """Update inbox thread"""
    user = g.user
    data = request.get_json() or {}
    
    thread_id = data.get("id")
    if not thread_id:
        return jsonify(ok=False, error="Thread ID required"), 400
    
    # Verify ownership
    thread = query_db("SELECT id FROM inbox_threads WHERE id=? AND org_id=?", (thread_id, user["org_id"]), one=True)
    if not thread:
        return jsonify(ok=False, error="Thread not found"), 404
    
    updates = {k: v for k, v in data.items() if k != "id"}
    
    # Handle tags (convert array to CSV)
    if "tags_json" in updates:
        tags = updates.pop("tags_json")
        updates["tags_csv"] = ",".join(tags) if isinstance(tags, list) else ""
    
    if update_inbox_thread(thread_id, user["org_id"], updates):
        return jsonify(ok=True)
    
    return jsonify(ok=False, error="No valid updates"), 400


@app.route("/api/message/send", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("message_send", per_min=60)
def api_message_send():
    """Send message to inbox thread"""
    user = g.user
    data = request.get_json() or {}
    
    thread_id = data.get("thread_id")
    body = data.get("body", "").strip()
    internal_note = data.get("internal_note", False)
    attachments = data.get("attachments", [])
    
    if not thread_id or not body:
        return jsonify(ok=False, error="Thread ID and body required"), 400
    
    # Verify thread
    thread = query_db("SELECT id FROM inbox_threads WHERE id=? AND org_id=?", (thread_id, user["org_id"]), one=True)
    if not thread:
        return jsonify(ok=False, error="Thread not found"), 404
    
    message_id = add_message(
        thread_id,
        "agent",
        body,
        user["id"],
        "",
        internal_note,
        attachments
    )
    
    if message_id is None:
        return jsonify(ok=False, error="Duplicate message"), 400
    
    return jsonify(ok=True, id=message_id)


@app.route("/api/message/to_task", methods=["POST"])
@_login_required
@_csrf_protect
def api_message_to_task():
    """Create task from message"""
    user = g.user
    data = request.get_json() or {}
    
    message_id = data.get("message_id")
    title = data.get("title", "").strip()
    
    if not message_id or not title:
        return jsonify(ok=False, error="Message ID and title required"), 400
    
    # Get message
    message = query_db(
        """
        SELECT m.*, t.company_id, t.contact_id
        FROM inbox_messages m
        JOIN inbox_threads t ON m.thread_id = t.id
        WHERE m.id=? AND t.org_id=?
        """,
        (message_id, user["org_id"]),
        one=True
    )
    
    if not message:
        return jsonify(ok=False, error="Message not found"), 404
    
    # Create task
    task_id = create_task(
        org_id=user["org_id"],
        title=title,
        description=message["body"] or "",
        assignee_id=user["id"],
        due_at=data.get("due_at"),
        company_id=data.get("company_id") or message["company_id"],
        contact_id=message["contact_id"],
    )
    
    return jsonify(ok=True, task_id=task_id)


# ===== BLOCK: API ROUTES — CALLS =====

@app.route("/api/calls/list", methods=["GET"])
@_login_required
def api_calls_list():
    """Get calls list"""
    user = g.user
    org_id = user["org_id"]
    
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 100)), 200)
    offset = (page - 1) * per_page
    
    mine = request.args.get("mine", "1") == "1"
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    
    # Build WHERE
    where_parts = ["c.org_id = ?"]
    params = [org_id]
    
    if mine:
        where_parts.append("c.agent_id = ?")
        params.append(user["id"])
    
    if date_from:
        where_parts.append("c.started_at >= ?")
        params.append(ensure_iso_datetime(date_from))
    
    if date_to:
        where_parts.append("c.started_at <= ?")
        params.append(ensure_iso_datetime(date_to))
    
    where_clause = " AND ".join(where_parts)
    
    calls = query_db(
        f"""
        SELECT 
            c.*,
            u.username AS agent_name,
            comp.name AS company_name,
            cont.name AS contact_name
        FROM calls c
        LEFT JOIN users u ON c.agent_id = u.id
        LEFT JOIN companies comp ON c.company_id = comp.id
        LEFT JOIN contacts cont ON c.contact_id = cont.id
        WHERE {where_clause}
        ORDER BY c.started_at DESC
        LIMIT ? OFFSET ?
        """,
        (*params, per_page, offset)
    )
    
    return jsonify(
        ok=True,
        items=[dict(c) for c in calls],
        page=page,
        per_page=per_page
    )


@app.route("/api/call/assign_agent", methods=["POST"])
@_login_required
@_csrf_protect
def api_call_assign_agent():
    """Assign agent to call"""
    user = g.user
    data = request.get_json() or {}
    
    call_id = data.get("call_id")
    agent_id = data.get("agent_id")
    
    if not call_id:
        return jsonify(ok=False, error="Call ID required"), 400
    
    # Verify call
    call = query_db("SELECT id FROM calls WHERE id=? AND org_id=?", (call_id, user["org_id"]), one=True)
    if not call:
        return jsonify(ok=False, error="Call not found"), 404
    
    exec_db("UPDATE calls SET agent_id=? WHERE id=?", (agent_id, call_id))
    
    return jsonify(ok=True)


@app.route("/api/call/to_task", methods=["POST"])
@_login_required
@_csrf_protect
def api_call_to_task():
    """Create task from call"""
    user = g.user
    data = request.get_json() or {}
    
    call_id = data.get("call_id")
    title = data.get("title", f"Звонок #{call_id}")
    
    if not call_id:
        return jsonify(ok=False, error="Call ID required"), 400
    
    # Get call
    call = query_db(
        "SELECT * FROM calls WHERE id=? AND org_id=?",
        (call_id, user["org_id"]),
        one=True
    )
    
    if not call:
        return jsonify(ok=False, error="Call not found"), 404
    
    # Create task
    task_id = create_task(
        org_id=user["org_id"],
        title=title,
        description=f"От: {call['from_e164']}\nКому: {call['to_e164']}\nДлительность: {call['duration_sec']} сек",
        assignee_id=user["id"],
        company_id=call["company_id"],
        contact_id=call["contact_id"],
    )
    
    return jsonify(ok=True, task_id=task_id)


@app.route("/api/task/phones/<int:task_id>", methods=["GET"])
@_login_required
def api_task_phones(task_id):
    """Get phone numbers associated with task (from company/contact)"""
    user = g.user
    
    task = query_db(
        "SELECT company_id, contact_id, contact_phone FROM tasks WHERE id=? AND org_id=?",
        (task_id, user["org_id"]),
        one=True
    )
    
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    phones = []
    
    # Task direct phone
    if task["contact_phone"]:
        phones.append(normalize_phone(task["contact_phone"]))
    
    # Company phone
    if task["company_id"]:
        company = query_db("SELECT phone FROM companies WHERE id=?", (task["company_id"],), one=True)
        if company and company["phone"]:
            phones.append(normalize_phone(company["phone"]))
    
    # Contact phone
    if task["contact_id"]:
        contact = query_db("SELECT phone FROM contacts WHERE id=?", (task["contact_id"],), one=True)
        if contact and contact["phone"]:
            phones.append(normalize_phone(contact["phone"]))
    
    # Deduplicate
    phones = list(dict.fromkeys(p for p in phones if p))
    
    if not phones:
        return jsonify(ok=False, error="No phones found"), 404
    
    return jsonify(ok=True, items=phones)


@app.route("/api/cti/click_to_call", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("click_to_call", per_min=30)
def api_cti_click_to_call():
    """
    Initiate outbound call via CTI.
    This is a placeholder - actual implementation depends on telephony provider.
    """
    user = g.user
    data = request.get_json() or {}
    
    to = data.get("to", "").strip()
    if not to:
        return jsonify(ok=False, error="Phone number required"), 400
    
    # Normalize phone
    to_e164 = normalize_phone(to)
    if not to_e164:
        return jsonify(ok=False, error="Invalid phone number"), 400
    
    # Get user's phone channel config
    channel = query_db(
        """
        SELECT * FROM channels
        WHERE org_id=? AND type='phone' AND active=1
        LIMIT 1
        """,
        (user["org_id"],),
        one=True
    )
    
    if not channel:
        return jsonify(ok=False, error="Phone channel not configured"), 400
    
    # Here you would integrate with telephony provider API
    # Example: Mango, UIS, TELFIN, Twilio, etc.
    # For now, just log the intent
    
    log("INFO", f"Click-to-call initiated: {user['username']} → {to_e164}")
    
    # Create call record
    call_id = exec_db(
        """
        INSERT INTO calls (org_id, channel_id, direction, from_e164, to_e164, agent_id, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (user["org_id"], channel["id"], "outbound", user.get("phone", ""), to_e164, user["id"], "initiated", utc_now())
    )
    
    return jsonify(ok=True, call_id=call_id, message=f"Calling {to_e164}...")


# ===== BLOCK: API ROUTES — MEETINGS =====

@app.route("/api/meetings", methods=["GET"])
@_login_required
def api_meetings_list():
    """Get meetings list"""
    user = g.user
    
    meetings = query_db(
        """
        SELECT * FROM meetings
        WHERE org_id=?
        ORDER BY start_at DESC
        LIMIT 1000
        """,
        (user["org_id"],)
    )
    
    return jsonify(ok=True, items=[dict(m) for m in meetings])


@app.route("/api/meetings/schedule", methods=["POST"])
@_login_required
@_csrf_protect
def api_meetings_schedule():
    """Schedule new meeting"""
    user = g.user
    data = request.get_json() or {}
    
    title = data.get("title", "Встреча")
    start_at = data.get("start_at")
    end_at = data.get("end_at")
    participants = data.get("participants", [])
    department_ids = data.get("department_ids", [])
    
    # Generate room ID
    room = f"room_{uuid.uuid4().hex[:12]}"
    
    # Create meeting
    meeting_id = exec_db(
        """
        INSERT INTO meetings (org_id, title, room, start_at, end_at, created_by, participants_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["org_id"],
            title,
            room,
            ensure_iso_datetime(start_at) if start_at else None,
            ensure_iso_datetime(end_at) if end_at else None,
            user["id"],
            json.dumps(participants),
            utc_now()
        )
    )
    
    # Send notifications (placeholder)
    # TODO: Send email/push notifications to participants
    
    return jsonify(ok=True, id=meeting_id, room=room)


@app.route("/api/meetings/<int:meeting_id>", methods=["DELETE"])
@_login_required
@_csrf_protect
def api_meeting_delete(meeting_id):
    """Delete meeting"""
    user = g.user
    
    meeting = query_db("SELECT id FROM meetings WHERE id=? AND org_id=?", (meeting_id, user["org_id"]), one=True)
    if not meeting:
        return jsonify(ok=False, error="Meeting not found"), 404
    
    exec_db("DELETE FROM meetings WHERE id=?", (meeting_id,))
    
    return jsonify(ok=True)


# ===== BLOCK: API ROUTES — CHAT =====

@app.route("/api/chat/create", methods=["POST"])
@_login_required
@_csrf_protect
def api_chat_create():
    """Create chat channel"""
    user = g.user
    data = request.get_json() or {}
    
    channel_type = data.get("type", "public")
    title = data.get("title", "")
    members = data.get("members", [])
    department_ids = data.get("department_ids", [])
    
    # Create channel
    channel_id = exec_db(
        "INSERT INTO chat_channels (org_id, type, title, created_at) VALUES (?, ?, ?, ?)",
        (user["org_id"], channel_type, title, utc_now())
    )
    
    # Add members
    for user_id in members:
        exec_db(
            "INSERT INTO chat_members (channel_id, user_id, created_at) VALUES (?, ?, ?)",
            (channel_id, user_id, utc_now())
        )
    
    # Add departments
    for dept_id in department_ids:
        exec_db(
            "INSERT INTO chat_members (channel_id, department_id, created_at) VALUES (?, ?, ?)",
            (channel_id, dept_id, utc_now())
        )
    
    return jsonify(ok=True, id=channel_id)


@app.route("/api/chat/send", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("chat_send", per_min=120)
def api_chat_send():
    """Send chat message"""
    user = g.user
    data = request.get_json() or {}
    
    channel_id = data.get("channel_id")
    body = data.get("body", "").strip()
    
    if not channel_id or not body:
        return jsonify(ok=False, error="Channel ID and body required"), 400
    
    # Verify channel access
    channel = query_db(
        """
        SELECT c.* FROM chat_channels c
        LEFT JOIN chat_members m ON c.id = m.channel_id
        WHERE c.id=? AND c.org_id=? AND (c.type='public' OR m.user_id=?)
        LIMIT 1
        """,
        (channel_id, user["org_id"], user["id"]),
        one=True
    )
    
    if not channel:
        return jsonify(ok=False, error="Channel not found or access denied"), 403
    
    # Add message
    message_id = exec_db(
        "INSERT INTO chat_messages (channel_id, user_id, body, created_at) VALUES (?, ?, ?, ?)",
        (channel_id, user["id"], body, utc_now())
    )
    
    return jsonify(ok=True, id=message_id)


# ===== BLOCK: API ROUTES — FILES =====

@app.route("/api/files/<int:file_id>/download", methods=["GET"])
@_login_required
def api_file_download(file_id):
    """Download file"""
    user = g.user
    
    # Verify file belongs to org
    file_meta = query_db("SELECT * FROM files WHERE id=? AND org_id=?", (file_id, user["org_id"]), one=True)
    if not file_meta:
        return jsonify(ok=False, error="File not found"), 404
    
    # Get file data
    result = get_file_by_id(file_id)
    if not result:
        return jsonify(ok=False, error="File data not found"), 404
    
    data, filename, content_type = result
    
    return send_file(
        BytesIO(data),
        mimetype=content_type,
        as_attachment=True,
        download_name=filename
    )


# ===== BLOCK: API ROUTES — AI =====

@app.route("/api/ai/summarize_thread", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("ai_summarize", per_min=10)
def api_ai_summarize_thread():
    """Generate AI summary of thread"""
    user = g.user
    data = request.get_json() or {}
    
    thread_id = data.get("thread_id")
    if not thread_id:
        return jsonify(ok=False, error="Thread ID required"), 400
    
    # Get thread messages
    messages = query_db(
        """
        SELECT body, sender_type, created_at
        FROM inbox_messages
        WHERE thread_id=?
        ORDER BY created_at ASC
        LIMIT 50
        """,
        (thread_id,)
    )
    
    if not messages:
        return jsonify(ok=False, error="No messages found"), 404
    
    # Build context
    context = "\n".join([
        f"[{m['sender_type']}] {m['body']}"
        for m in messages
    ])
    
    context = truncate_for_ai_context(context, 6000)
    
    # Call AI
    try:
        prompt = f"Summarize this customer support conversation:\n\n{context}"
        summary = ai_provider_call(prompt, system="You are a helpful assistant that summarizes conversations.")
        
        # Log AI job
        exec_db(
            """
            INSERT INTO ai_jobs (org_id, user_id, job_type, entity_type, entity_id, input_text, output_text, status, created_at, completed_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (user["org_id"], user["id"], "summarize", "thread", thread_id, context[:1000], summary, "completed", utc_now(), utc_now())
        )
        
        return jsonify(ok=True, summary=summary)
    
    except Exception as e:
        log("ERROR", f"AI summarize failed: {e}")
        return jsonify(ok=False, error="AI request failed"), 500


@app.route("/api/ai/draft_reply", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("ai_draft", per_min=10)
def api_ai_draft_reply():
    """Generate AI draft reply"""
    user = g.user
    data = request.get_json() or {}
    
    thread_id = data.get("thread_id")
    tone = data.get("tone", "neutral")
    
    if not thread_id:
        return jsonify(ok=False, error="Thread ID required"), 400
    
    # Get last few messages
    messages = query_db(
        """
        SELECT body, sender_type
        FROM inbox_messages
        WHERE thread_id=?
        ORDER BY created_at DESC
        LIMIT 10
        """,
        (thread_id,)
    )
    
    if not messages:
        return jsonify(ok=False, error="No messages found"), 404
    
    # Build context (reverse to chronological order)
    context = "\n".join([
        f"[{m['sender_type']}] {m['body']}"
        for m in reversed(list(messages))
    ])
    
    context = truncate_for_ai_context(context, 4000)
    
    # Tone mapping
    tone_instructions = {
        "friendly": "friendly and warm",
        "neutral": "professional and neutral",
        "formal": "formal and business-like"
    }
    
    tone_text = tone_instructions.get(tone, "professional")
    
    # Generate 3 variants (sequentially for now, can be parallelized)
    try:
        variants = []
        for i in range(3):
            prompt = f"Generate a {tone_text} reply to this conversation (variant {i+1}):\n\n{context}\n\nReply:"
            reply = ai_provider_call(prompt, system="You are a customer support assistant.")
            variants.append(reply.strip())
        
        return jsonify(ok=True, variants=variants)
    
    except Exception as e:
        log("ERROR", f"AI draft failed: {e}")
        return jsonify(ok=False, error="AI request failed"), 500


@app.route("/api/ai/autotag", methods=["POST"])
@_login_required
@_csrf_protect
def api_ai_autotag():
    """Auto-tag thread based on content"""
    user = g.user
    data = request.get_json() or {}
    
    thread_id = data.get("thread_id")
    if not thread_id:
        return jsonify(ok=False, error="Thread ID required"), 400
    
    # Get messages
    messages = query_db(
        "SELECT body FROM inbox_messages WHERE thread_id=? LIMIT 20",
        (thread_id,)
    )
    
    if not messages:
        return jsonify(ok=False, error="No messages"), 404
    
    context = " ".join([m["body"] for m in messages])
    context = truncate_for_ai_context(context, 2000)
    
    try:
        prompt = f"Extract 3-5 tags from this conversation. Return only comma-separated tags:\n\n{context}"
        tags_text = ai_provider_call(prompt, max_tokens=50)
        
        # Parse tags
        tags = [t.strip() for t in tags_text.split(",") if t.strip()][:5]
        
        # Update thread
        exec_db(
            "UPDATE inbox_threads SET tags_csv=? WHERE id=?",
            (",".join(tags), thread_id)
        )
        
        return jsonify(ok=True, tags=tags)
    
    except Exception as e:
        log("ERROR", f"AI autotag failed: {e}")
        return jsonify(ok=False, error="AI request failed"), 500


# ===== END OF CORE PART 6/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 7/10 — WEB ROUTES, SSE, WEBHOOKS, REPORTS, ADMIN
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: WEB ROUTES — MAIN PAGES =====

@app.route("/")
@_login_required
def index():
    """Dashboard / Home page"""
    user = g.user
    
    # Get quick stats
    open_tasks_count = query_db(
        "SELECT COUNT(*) as cnt FROM tasks WHERE org_id=? AND status NOT IN ('done', 'cancelled')",
        (user["org_id"],),
        one=True
    )["cnt"]
    
    my_tasks_count = query_db(
        "SELECT COUNT(*) as cnt FROM tasks WHERE org_id=? AND assignee_id=? AND status NOT IN ('done', 'cancelled')",
        (user["org_id"], user["id"]),
        one=True
    )["cnt"]
    
    open_deals_count = query_db(
        "SELECT COUNT(*) as cnt FROM deals WHERE org_id=? AND status='open'",
        (user["org_id"],),
        one=True
    )["cnt"]
    
    return render_template_string(
        DASHBOARD_TMPL,
        user=user,
        open_tasks=open_tasks_count,
        my_tasks=my_tasks_count,
        open_deals=open_deals_count,
        now=datetime.utcnow(),
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/profile")
@_login_required
def profile():
    """User profile page"""
    user = g.user
    
    # Get user stats
    tasks_created = query_db(
        "SELECT COUNT(*) as cnt FROM tasks WHERE org_id=? AND assignee_id=?",
        (user["org_id"], user["id"]),
        one=True
    )["cnt"]
    
    # Get available countries/timezones for dropdown
    countries = [
        {"code": "RU", "name": "Россия"},
        {"code": "US", "name": "США"},
        {"code": "GB", "name": "Великобритания"},
    ]
    
    timezones = [
        "UTC",
        "Europe/Moscow",
        "America/New_York",
        "Europe/London",
    ]
    
    return render_template_string(
        PROFILE_TMPL,
        user=user,
        tasks_created=tasks_created,
        countries=countries,
        timezones=timezones,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/inbox")
@_login_required
def inbox():
    """Inbox threads list page"""
    user = g.user
    
    # Get filters from query params
    filters = {
        "status": request.args.get("status", ""),
        "channel": request.args.get("channel", ""),
        "assignee": request.args.get("assignee", ""),
        "kind": request.args.get("kind", ""),
        "tags": request.args.get("tags", ""),
        "who": request.args.get("who", ""),
        "date_from": request.args.get("date_from", ""),
        "date_to": request.args.get("date_to", ""),
        "q": request.args.get("q", ""),
    }
    
    # Get initial data (first page)
    page = 1
    per_page = 50
    
    # Build WHERE clause
    where_parts = ["t.org_id = ?"]
    params = [user["org_id"]]
    
    if filters["status"]:
        where_parts.append("t.status = ?")
        params.append(filters["status"])
    
    if filters["channel"]:
        where_parts.append("t.channel_id = ?")
        params.append(int(filters["channel"]))
    
    if filters["assignee"]:
        where_parts.append("t.assignee_id = ?")
        params.append(int(filters["assignee"]))
    
    if filters["who"] == "me":
        where_parts.append("t.assignee_id = ?")
        params.append(user["id"])
    
    if filters["q"]:
        where_parts.append("(t.subject LIKE ? OR t.external_thread_id LIKE ?)")
        params.extend([f"%{filters['q']}%", f"%{filters['q']}%"])
    
    where_clause = " AND ".join(where_parts)
    
    rows = query_db(
        f"""
        SELECT 
            t.*,
            c.name AS channel_name
        FROM inbox_threads t
        LEFT JOIN channels c ON t.channel_id = c.id
        WHERE {where_clause}
        ORDER BY t.last_message_at DESC
        LIMIT ? OFFSET ?
        """,
        (*params, per_page, 0)
    )
    
    # Get channels for filter
    channels = query_db(
        "SELECT id, name, type FROM channels WHERE org_id=? AND active=1",
        (user["org_id"],)
    )
    
    # Get agents for filter
    agents = query_db(
        "SELECT id, username FROM users WHERE org_id=? AND active=1",
        (user["org_id"],)
    )
    
    return render_template_string(
        INBOX_TMPL,
        user=user,
        filters=filters,
        rows=rows,
        channels=channels,
        agents=agents,
        now=datetime.utcnow(),
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/thread/<int:tid>")
@_login_required
def thread_view(tid):
    """Thread detail page"""
    user = g.user
    
    # Get thread
    r = query_db(
        "SELECT * FROM inbox_threads WHERE id=? AND org_id=?",
        (tid, user["org_id"]),
        one=True
    )
    
    if not r:
        return "Thread not found", 404
    
    # Get messages
    messages = query_db(
        """
        SELECT 
            m.*,
            u.username
        FROM inbox_messages m
        LEFT JOIN users u ON m.user_id = u.id
        WHERE m.thread_id=?
        ORDER BY m.created_at DESC
        LIMIT 50
        """,
        (tid,)
    )
    
    return render_template_string(
        THREAD_TMPL,
        user=user,
        r=r,
        messages=messages,
        now=datetime.utcnow(),
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/tasks")
@_login_required
def tasks_page():
    """Tasks list page"""
    user = g.user
    
    # Get filter
    current_filter = request.args.get("f", "open")
    
    # Get initial tasks (first page)
    where_parts = ["t.org_id = ?"]
    params = [user["org_id"]]
    
    if current_filter == "open":
        where_parts.append("t.status NOT IN ('done', 'cancelled')")
    elif current_filter == "today":
        where_parts.append("date(t.due_at) = date('now')")
    elif current_filter == "overdue":
        where_parts.append("t.due_at < datetime('now') AND t.status NOT IN ('done', 'cancelled')")
    elif current_filter == "done":
        where_parts.append("t.status = 'done'")
    
    where_clause = " AND ".join(where_parts)
    
    tasks = query_db(
        f"""
        SELECT 
            t.*,
            c.name AS company_name
        FROM tasks t
        LEFT JOIN companies c ON t.company_id = c.id
        WHERE {where_clause}
        ORDER BY t.due_at ASC, t.created_at DESC
        LIMIT 50
        """,
        tuple(params)
    )
    
    # Get task statuses
    statuses = query_db(
        "SELECT * FROM task_statuses WHERE org_id=? ORDER BY sort_order",
        (user["org_id"],)
    )
    
    # Get agents
    agents = query_db(
        "SELECT id, username FROM users WHERE org_id=? AND active=1",
        (user["org_id"],)
    )
    
    return render_template_string(
        TASKS_TMPL,
        user=user,
        current_filter=current_filter,
        tasks=tasks,
        statuses=statuses,
        agents=agents,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/tasks", methods=["POST"])
@_login_required
@_csrf_protect
def tasks_create_form():
    """Create task from form (web route)"""
    user = g.user
    
    title = request.form.get("title", "").strip()
    if not title:
        return redirect(url_for("tasks_page"))
    
    create_task(
        org_id=user["org_id"],
        title=title,
        description=request.form.get("description", ""),
        assignee_id=request.form.get("assignee_id") or None,
        due_at=request.form.get("due_at"),
        company_id=request.form.get("company_id") or None,
        monthly_fee=float(request.form.get("monthly_fee", 0) or 0),
    )
    
    return redirect(url_for("tasks_page"))


@app.route("/task/<int:tid>")
@_login_required
def task_view(tid):
    """Task detail page"""
    user = g.user
    
    # Get task
    t = query_db(
        "SELECT * FROM tasks WHERE id=? AND org_id=?",
        (tid, user["org_id"]),
        one=True
    )
    
    if not t:
        return "Task not found", 404
    
    # Get comments
    comments = get_task_comments(tid)
    
    # Get participants
    participants = query_db(
        """
        SELECT p.*, u.username
        FROM task_participants p
        JOIN users u ON p.user_id = u.id
        WHERE p.task_id=?
        """,
        (tid,)
    )
    
    # Get pinned files
    pinned_files = query_db(
        """
        SELECT f.*
        FROM task_files tf
        JOIN files f ON tf.file_id = f.id
        WHERE tf.task_id=? AND tf.pinned=1
        """,
        (tid,)
    )
    
    # Get transitions
    transitions = query_db(
        """
        SELECT * FROM stage_transitions
        WHERE entity_type='task' AND entity_id=?
        ORDER BY created_at DESC
        LIMIT 10
        """,
        (tid,)
    )
    
    # Get activity
    activity = query_db(
        "SELECT * FROM task_activity WHERE task_id=? ORDER BY created_at DESC LIMIT 20",
        (tid,)
    )
    
    # Get task statuses
    statuses = query_db(
        "SELECT * FROM task_statuses WHERE org_id=?",
        (user["org_id"],)
    )
    
    return render_template_string(
        TASK_VIEW_TMPL,
        user=user,
        t=t,
        comments=comments,
        participants=participants,
        pinned_files=pinned_files,
        transitions=transitions,
        activity=activity,
        statuses=statuses,
        now=datetime.utcnow(),
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/deals")
@_login_required
def deals_page():
    """Deals list page"""
    user = g.user
    
    # Get initial deals
    deals = query_db(
        """
        SELECT 
            d.*,
            c.name AS company_name
        FROM deals d
        LEFT JOIN companies c ON d.company_id = c.id
        WHERE d.org_id=?
        ORDER BY d.created_at DESC
        LIMIT 50
        """,
        (user["org_id"],)
    )
    
    # Get users
    users = query_db(
        "SELECT id, username FROM users WHERE org_id=? AND active=1",
        (user["org_id"],)
    )
    
    # Create users_map for JS
    users_map = {str(u["id"]): u["username"] for u in users}
    
    return render_template_string(
        DEALS_TMPL,
        user=user,
        deals=deals,
        users=users,
        users_map=users_map,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/clients")
@_login_required
def clients_list():
    """Companies list page"""
    user = g.user
    
    # Get initial clients
    clients = query_db(
        """
        SELECT 
            c.*,
            (SELECT COUNT(*) FROM deals WHERE company_id=c.id) AS deals
        FROM companies c
        WHERE c.org_id=?
        ORDER BY c.created_at DESC
        LIMIT 50
        """,
        (user["org_id"],)
    )
    
    return render_template_string(
        CLIENTS_TMPL,
        user=user,
        clients=clients,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/clients", methods=["POST"])
@_login_required
@_csrf_protect
def clients_create_form():
    """Create client from form"""
    user = g.user
    
    name = request.form.get("name", "").strip()
    if not name:
        return redirect(url_for("clients_list"))
    
    create_company(
        org_id=user["org_id"],
        name=name,
        inn=request.form.get("inn", ""),
        phone=request.form.get("phone", ""),
        email=request.form.get("email", ""),
        address=request.form.get("address", ""),
        notes=request.form.get("notes", ""),
    )
    
    return redirect(url_for("clients_list"))


@app.route("/client/<int:cid>")
@_login_required
def client_page(cid):
    """Company detail page"""
    user = g.user
    
    c = query_db(
        "SELECT * FROM companies WHERE id=? AND org_id=?",
        (cid, user["org_id"]),
        one=True
    )
    
    if not c:
        return "Company not found", 404
    
    # Get calls
    calls = query_db(
        "SELECT * FROM calls WHERE company_id=? ORDER BY started_at DESC LIMIT 50",
        (cid,)
    )
    
    return render_template_string(
        CLIENT_PAGE_TMPL,
        user=user,
        c=c,
        calls=calls,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/calls")
@_login_required
def calls_page():
    """Calls list page"""
    user = g.user
    
    # Get agents for dropdown
    agents_rows = query_db(
        "SELECT id, username FROM users WHERE org_id=? AND active=1 ORDER BY username",
        (user["org_id"],)
    )
    
    return render_template_string(
        CALLS_TMPL,
        user=user,
        agents_rows=agents_rows,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/meeting/<int:meeting_id>")
@_login_required
def meeting_view(meeting_id):
    """Meeting detail page"""
    user = g.user
    
    meeting = query_db(
        "SELECT * FROM meetings WHERE id=? AND org_id=?",
        (meeting_id, user["org_id"]),
        one=True
    )
    
    return render_template_string(
        MEETING_TMPL,
        user=user,
        meeting=meeting,
        jitsi_base=JITSI_BASE,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/meetings")
@_login_required
def meetings_page():
    """Meetings list & create page"""
    user = g.user
    
    return render_template_string(
        MEETING_TMPL,
        user=user,
        meeting=None,
        jitsi_base=JITSI_BASE,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/chat")
@_login_required
def chat():
    """Chat page (list of channels)"""
    user = g.user
    
    # Get channels
    channels = query_db(
        """
        SELECT DISTINCT c.*
        FROM chat_channels c
        LEFT JOIN chat_members m ON c.channel_id = m.channel_id
        WHERE c.org_id=? AND (c.type='public' OR m.user_id=?)
        ORDER BY c.created_at DESC
        """,
        (user["org_id"], user["id"])
    )
    
    return render_template_string(
        CHAT_TMPL,
        user=user,
        channels=channels,
        current=None,
        messages=[],
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/chat/<int:cid>")
@_login_required
def chat_channel(cid):
    """Chat channel view"""
    user = g.user
    
    # Get channel
    current = query_db("SELECT * FROM chat_channels WHERE id=? AND org_id=?", (cid, user["org_id"]), one=True)
    if not current:
        return "Channel not found", 404
    
    # Get all channels for sidebar
    channels = query_db(
        """
        SELECT DISTINCT c.*
        FROM chat_channels c
        LEFT JOIN chat_members m ON c.id = m.channel_id
        WHERE c.org_id=? AND (c.type='public' OR m.user_id=?)
        ORDER BY c.created_at DESC
        """,
        (user["org_id"], user["id"])
    )
    
    # Get messages
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page
    
    messages = query_db(
        """
        SELECT m.*, u.username
        FROM chat_messages m
        LEFT JOIN users u ON m.user_id = u.id
        WHERE m.channel_id=?
        ORDER BY m.created_at DESC
        LIMIT ? OFFSET ?
        """,
        (cid, per_page, offset)
    )
    
    return render_template_string(
        CHAT_TMPL,
        user=user,
        channels=channels,
        current=current,
        messages=messages,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/search")
@_login_required
def search_page():
    """Global search page"""
    user = g.user
    q = request.args.get("q", "").strip()
    
    results = {
        "inbox": [],
        "tasks": [],
        "chats": []
    }
    
    if q and len(q) >= 2:
        # Search inbox messages (FTS)
        fts_results = search_fts("fts_inbox_messages", q, 20)
        if fts_results:
            ids = [r["rowid"] for r in fts_results]
            placeholders = ",".join("?" * len(ids))
            messages = query_db(
                f"SELECT id, thread_id, body FROM inbox_messages WHERE id IN ({placeholders})",
                tuple(ids)
            )
            results["inbox"] = [dict(m) for m in messages]
        
        # Search tasks (simple LIKE for now)
        tasks = query_db(
            """
            SELECT id, title, description FROM tasks
            WHERE org_id=? AND (title LIKE ? OR description LIKE ?)
            LIMIT 20
            """,
            (user["org_id"], f"%{q}%", f"%{q}%")
        )
        results["tasks"] = [dict(t) for t in tasks]
        
        # Search chat messages
        chat_fts = search_fts("fts_chat_messages", q, 20)
        if chat_fts:
            ids = [r["rowid"] for r in chat_fts]
            placeholders = ",".join("?" * len(ids))
            chats = query_db(
                f"SELECT id, channel_id, body FROM chat_messages WHERE id IN ({placeholders})",
                tuple(ids)
            )
            results["chats"] = [dict(c) for c in chats]
    
    return render_template_string(
        SEARCH_TMPL,
        user=user,
        q=q,
        results=results,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/documents")
@_login_required
def documents_page():
    """Documents page"""
    user = g.user
    
    # Get templates
    templates = query_db(
        "SELECT * FROM document_templates WHERE org_id=? ORDER BY created_at DESC",
        (user["org_id"],)
    )
    
    # Get companies for dropdown
    companies = query_db(
        "SELECT id, name, inn FROM companies WHERE org_id=? ORDER BY name LIMIT 500",
        (user["org_id"],)
    )
    
    # Get recent documents
    docs = query_db(
        """
        SELECT d.*, c.name AS company_name
        FROM documents d
        LEFT JOIN companies c ON d.company_id = c.id
        WHERE d.org_id=?
        ORDER BY d.created_at DESC
        LIMIT 50
        """,
        (user["org_id"],)
    )
    
    return render_template_string(
        DOCUMENTS_TMPL,
        user=user,
        templates=templates,
        companies=companies,
        docs=docs,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/documents/template", methods=["POST"])
@_login_required
@_csrf_protect
def documents_template_add():
    """Add document template"""
    user = g.user
    
    exec_db(
        """
        INSERT INTO document_templates (org_id, type, tkey, name, body_template, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            user["org_id"],
            request.form.get("type", "other"),
            request.form.get("tkey", ""),
            request.form.get("name", ""),
            request.form.get("body_template", ""),
            utc_now()
        )
    )
    
    return redirect(url_for("documents_page"))


@app.route("/documents", methods=["POST"])
@_login_required
@_csrf_protect
def documents_create():
    """Create document from template"""
    user = g.user
    
    template_id = request.form.get("template_id")
    company_id = request.form.get("company_id")
    
    if not template_id or not company_id:
        return redirect(url_for("documents_page"))
    
    # Get template
    template = query_db(
        "SELECT * FROM document_templates WHERE id=? AND org_id=?",
        (template_id, user["org_id"]),
        one=True
    )
    
    if not template:
        return redirect(url_for("documents_page"))
    
    # Get company
    company = query_db("SELECT * FROM companies WHERE id=? AND org_id=?", (company_id, user["org_id"]), one=True)
    if not company:
        return redirect(url_for("documents_page"))
    
    # Get org
    org = query_db("SELECT * FROM orgs WHERE id=?", (user["org_id"],), one=True)
    
    # Render template (simple replacement for now)
    html = template["body_template"]
    
    # Replace variables
    replacements = {
        "{{ org.name }}": org["name"] if org else "",
        "{{ company.name }}": company["name"],
        "{{ company.inn }}": company["inn"] or "",
        "{{ user.username }}": user["username"],
        "{{ user.first_name }}": user["first_name"] or "",
        "{{ now }}": datetime.utcnow().strftime("%Y-%m-%d"),
    }
    
    for old, new in replacements.items():
        html = html.replace(old, new)
    
    # Sanitize HTML
    html = sanitize_html(html)
    
    # Create document
    doc_id = exec_db(
        """
        INSERT INTO documents (org_id, template_id, doc_type, title, content_html, company_id, user_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["org_id"],
            template_id,
            template["type"],
            template["name"] + " - " + company["name"],
            html,
            company_id,
            user["id"],
            utc_now()
        )
    )
    
    return redirect(url_for("document_view", doc_id=doc_id))


@app.route("/document/<int:doc_id>")
@_login_required
def document_view(doc_id):
    """Document view page"""
    user = g.user
    
    d = query_db(
        "SELECT * FROM documents WHERE id=? AND org_id=?",
        (doc_id, user["org_id"]),
        one=True
    )
    
    if not d:
        return "Document not found", 404
    
    return render_template_string(
        DOCUMENT_VIEW_TMPL,
        user=user,
        d=d,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/warehouse")
@_login_required
def warehouse():
    """Warehouse/inventory page"""
    user = g.user
    
    products = query_db(
        "SELECT * FROM products WHERE org_id=? ORDER BY name",
        (user["org_id"],)
    )
    
    return render_template_string(
        WAREHOUSE_TMPL,
        user=user,
        products=products,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/warehouse/stock/set", methods=["POST"])
@_login_required
@_csrf_protect
def warehouse_stock_set():
    """Update product stock quantity"""
    user = g.user
    
    product_id = request.form.get("product_id")
    qty = request.form.get("qty", "0")
    
    if not product_id:
        return redirect(url_for("warehouse"))
    
    exec_db(
        "UPDATE products SET qty=?, updated_at=? WHERE id=? AND org_id=?",
        (float(qty), utc_now(), product_id, user["org_id"])
    )
    
    return redirect(url_for("warehouse"))


@app.route("/import")
@_login_required
def import_page():
    """CSV import wizard page"""
    user = g.user
    
    return render_template_string(
        IMPORT_TMPL,
        user=user,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/import", methods=["POST"])
@_login_required
@_csrf_protect
def import_csv_wizard():
    """Process CSV import"""
    user = g.user
    
    mode = request.form.get("mode", "companies")
    
    if "csvfile" not in request.files:
        return redirect(url_for("import_page"))
    
    f = request.files["csvfile"]
    if not f.filename:
        return redirect(url_for("import_page"))
    
    # Read CSV
    import csv
    import io
    
    content = f.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(content), delimiter=";")
    
    imported_count = 0
    
    try:
        if mode == "companies":
            for row in reader:
                name = row.get("name", "").strip()
                if name:
                    create_company(
                        org_id=user["org_id"],
                        name=name,
                        inn=row.get("inn", ""),
                        phone=row.get("phone", ""),
                        email=row.get("email", ""),
                        address=row.get("address", ""),
                        notes=row.get("notes", "")
                    )
                    imported_count += 1
        
        elif mode == "tasks":
            for row in reader:
                title = row.get("title", "").strip()
                if title:
                    create_task(
                        org_id=user["org_id"],
                        title=title,
                        description=row.get("description", ""),
                        due_at=row.get("due_at"),
                        assignee_id=row.get("assignee_id") or None
                    )
                    imported_count += 1
    
    except Exception as e:
        log("ERROR", f"CSV import failed: {e}")
    
    add_audit(user["org_id"], user["id"], f"csv_import.{mode}", details={"count": imported_count})
    
    return redirect(url_for("import_page"))


@app.route("/analytics")
@_login_required
def analytics():
    """Analytics dashboard page"""
    user = g.user
    
    return render_template_string(
        ANALYTICS_TMPL,
        user=user,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


@app.route("/settings")
@_login_required
@_require_role("admin")
def settings_page():
    """Settings page (admin only)"""
    user = g.user
    
    # Get channels
    channels = query_db(
        "SELECT * FROM channels WHERE org_id=?",
        (user["org_id"],)
    )
    
    # Get webhooks
    webhooks = query_db(
        "SELECT * FROM webhooks WHERE org_id=?",
        (user["org_id"],)
    )
    
    # Get task statuses
    task_statuses = query_db(
        "SELECT * FROM task_statuses WHERE org_id=? ORDER BY sort_order",
        (user["org_id"],)
    )
    
    # Get users
    users = query_db(
        "SELECT * FROM users WHERE org_id=? ORDER BY username",
        (user["org_id"],)
    )
    
    # Get departments
    departments = query_db(
        "SELECT * FROM departments WHERE org_id=? ORDER BY name",
        (user["org_id"],)
    )
    
    # Get agents for dropdowns
    agents = query_db(
        "SELECT id, username FROM users WHERE org_id=? AND active=1",
        (user["org_id"],)
    )
    
    return render_template_string(
        SETTINGS_TMPL,
        user=user,
        channels=channels,
        webhooks=webhooks,
        task_statuses=task_statuses,
        users=users,
        departments=departments,
        agents=agents,
        csp_nonce=g.get("csp_nonce", ""),
        query_db=query_db
    )


# ===== BLOCK: SSE (SERVER-SENT EVENTS) =====

_sse_queues = {}  # {user_id: [Queue(), Queue(), ...]}
_sse_lock = threading.Lock()


def sse_push(user_id: int, event: str, data: dict):
    """
    Push event to user's SSE connections.
    ✅ УЛУЧШЕНО: graceful handling если нет подключений
    """
    if not SSE_ENABLED:
        return
    
    with _sse_lock:
        queues = _sse_queues.get(user_id, [])
        for q in queues:
            try:
                q.put({"event": event, "data": data}, block=False)
            except Exception:
                pass


@app.route("/sse")
@_login_required
def sse_stream():
    """SSE endpoint for real-time events"""
    if not SSE_ENABLED:
        return "SSE disabled", 503
    
    user = g.user
    user_id = user["id"]
    
    # Check connection limit
    with _sse_lock:
        existing = _sse_queues.get(user_id, [])
        if len(existing) >= SSE_MAX_CONN_PER_USER:
            return "Too many connections", 429
    
    # Create queue
    q = Queue(maxsize=100)
    
    with _sse_lock:
        if user_id not in _sse_queues:
            _sse_queues[user_id] = []
        _sse_queues[user_id].append(q)
    
    def generate():
        try:
            # Send initial ping
            yield "data: {\"event\":\"connected\"}\n\n"
            
            while True:
                try:
                    msg = q.get(timeout=30)
                    yield f"event: {msg['event']}\n"
                    yield f"data: {json.dumps(msg['data'])}\n\n"
                except Empty:
                    # Heartbeat
                    yield ": ping\n\n"
        finally:
            # Cleanup
            with _sse_lock:
                if user_id in _sse_queues:
                    _sse_queues[user_id].remove(q)
                    if not _sse_queues[user_id]:
                        del _sse_queues[user_id]
    
    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no"
        }
    )


# ===== BLOCK: WEBHOOKS =====

def trigger_webhook(org_id: int, event: str, payload: dict):
    """
    Queue webhook delivery for event.
    ✅ УЛУЧШЕНО: добавлен jitter для retry
    """
    # Get webhooks for event
    webhooks = query_db(
        "SELECT * FROM webhooks WHERE org_id=? AND event=? AND active=1",
        (org_id, event)
    )
    
    for webhook in webhooks:
        # Add to queue
        exec_db(
            """
            INSERT INTO webhook_queue (webhook_id, event, payload_json, status, next_try_at, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (webhook["id"], event, json.dumps(payload), "pending", utc_now(), utc_now())
        )


def _webhook_deliver_once():
    """
    Deliver pending webhooks (called by worker).
    ✅ УЛУЧШЕНО: добавлен jitter для exponential backoff
    """
    import random
    
    # Get pending tasks
    tasks = query_db(
        """
        SELECT wq.*, w.url, w.secret
        FROM webhook_queue wq
        JOIN webhooks w ON wq.webhook_id = w.id
        WHERE wq.status='pending' AND (wq.next_try_at IS NULL OR wq.next_try_at <= datetime('now'))
        LIMIT 10
        """
    )
    
    for task in tasks:
        try:
            # Prepare request
            payload_json = task["payload_json"]
            
            headers = {"Content-Type": "application/json"}
            
            # Add signature if secret exists
            if task["secret"]:
                signature = hmac.new(
                    task["secret"].encode(),
                    payload_json.encode(),
                    hashlib.sha256
                ).hexdigest()
                headers["X-Webhook-Signature"] = f"sha256={signature}"
            
            # Send request
            response = _rq.post(
                task["url"],
                data=payload_json,
                headers=headers,
                timeout=10
            )
            
            if response.status_code in (200, 201, 202, 204):
                # Success
                exec_db(
                    "UPDATE webhook_queue SET status='delivered' WHERE id=?",
                    (task["id"],)
                )
            else:
                # Failed - retry
                raise RuntimeError(f"HTTP {response.status_code}")
        
        except Exception as e:
            # Update attempts
            attempts = task["attempts"] + 1
            
            # Calculate next retry with exponential backoff + jitter
            base_delay = min(3600, int((2 ** min(6, attempts)) * 5))
            jitter = base_delay * random.uniform(0, 0.1)
            delay = base_delay + jitter
            
            next_try = (datetime.utcnow() + timedelta(seconds=delay)).isoformat(" ", "seconds")
            
            if attempts >= 10:
                # Give up
                exec_db(
                    "UPDATE webhook_queue SET status='failed', attempts=?, last_error=? WHERE id=?",
                    (attempts, str(e)[:500], task["id"])
                )
            else:
                exec_db(
                    "UPDATE webhook_queue SET attempts=?, next_try_at=?, last_error=? WHERE id=?",
                    (attempts, next_try, str(e)[:500], task["id"])
                )


@app.route("/api/webhook/queue", methods=["GET"])
@_login_required
@_require_role("admin")
def api_webhook_queue():
    """Get webhook delivery queue"""
    user = g.user
    
    items = query_db(
        """
        SELECT wq.*, w.event, w.url
        FROM webhook_queue wq
        JOIN webhooks w ON wq.webhook_id = w.id
        WHERE w.org_id=?
        ORDER BY wq.created_at DESC
        LIMIT 100
        """,
        (user["org_id"],)
    )
    
    return jsonify(ok=True, items=[dict(i) for i in items])


@app.route("/api/webhook/retry/<int:queue_id>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_webhook_retry(queue_id):
    """Retry webhook delivery"""
    user = g.user
    
    # Verify ownership
    item = query_db(
        """
        SELECT wq.* FROM webhook_queue wq
        JOIN webhooks w ON wq.webhook_id = w.id
        WHERE wq.id=? AND w.org_id=?
        """,
        (queue_id, user["org_id"]),
        one=True
    )
    
    if not item:
        return jsonify(ok=False, error="Not found"), 404
    
    # Reset status
    exec_db(
        "UPDATE webhook_queue SET status='pending', next_try_at=datetime('now'), attempts=0 WHERE id=?",
        (queue_id,)
    )
    
    return jsonify(ok=True)


# ===== BLOCK: REPORTS & ANALYTICS =====

@app.route("/api/reports/tasks_daily", methods=["GET"])
@_login_required
def api_reports_tasks_daily():
    """Daily tasks report"""
    user = g.user
    
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    
    # Build WHERE
    where_parts = ["org_id = ?"]
    params = [user["org_id"]]
    
    if date_from:
        where_parts.append("date(created_at) >= date(?)")
        params.append(date_from)
    
    if date_to:
        where_parts.append("date(created_at) <= date(?)")
        params.append(date_to)
    
    where_clause = " AND ".join(where_parts)
    
    # Query aggregated data
    items = query_db(
        f"""
        SELECT 
            date(created_at) AS ymd,
            SUM(CASE WHEN status NOT IN ('done', 'cancelled') THEN 1 ELSE 0 END) AS created_cnt,
            SUM(CASE WHEN status='done' THEN 1 ELSE 0 END) AS done_cnt,
            SUM(CASE WHEN due_at < datetime('now') AND status NOT IN ('done', 'cancelled') THEN 1 ELSE 0 END) AS overdue_cnt,
            SUM(monthly_fee) AS monthly_fee_sum
        FROM tasks
        WHERE {where_clause}
        GROUP BY ymd
        ORDER BY ymd DESC
        LIMIT 365
        """,
        tuple(params)
    )
    
    return jsonify(ok=True, items=[dict(i) for i in items])


@app.route("/api/reports/calls_daily", methods=["GET"])
@_login_required
def api_reports_calls_daily():
    """Daily calls report"""
    user = g.user
    
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    
    where_parts = ["org_id = ?"]
    params = [user["org_id"]]
    
    if date_from:
        where_parts.append("date(started_at) >= date(?)")
        params.append(date_from)
    
    if date_to:
        where_parts.append("date(started_at) <= date(?)")
        params.append(date_to)
    
    where_clause = " AND ".join(where_parts)
    
    items = query_db(
        f"""
        SELECT 
            date(started_at) AS ymd,
            SUM(CASE WHEN direction='inbound' THEN 1 ELSE 0 END) AS in_cnt,
            SUM(CASE WHEN direction='outbound' THEN 1 ELSE 0 END) AS out_cnt,
            SUM(duration_sec) AS dur_sum
        FROM calls
        WHERE {where_clause}
        GROUP BY ymd
        ORDER BY ymd DESC
        LIMIT 365
        """,
        tuple(params)
    )
    
    return jsonify(ok=True, items=[dict(i) for i in items])


# ===== BLOCK: ADMIN/SETTINGS ROUTES =====

@app.route("/settings/channel/add", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_channel_add():
    """Add channel"""
    user = g.user
    
    exec_db(
        "INSERT INTO channels (org_id, type, name, active, created_at) VALUES (?, ?, ?, ?, ?)",
        (user["org_id"], request.form.get("type", ""), request.form.get("name", ""), 1, utc_now())
    )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/channel/toggle/<int:cid>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_channel_toggle(cid):
    """Toggle channel active status"""
    user = g.user
    
    channel = query_db("SELECT active FROM channels WHERE id=? AND org_id=?", (cid, user["org_id"]), one=True)
    if channel:
        new_active = 0 if channel["active"] else 1
        exec_db("UPDATE channels SET active=? WHERE id=?", (new_active, cid))
    
    return redirect(url_for("settings_page"))


@app.route("/settings/webhook/add", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_webhook_add():
    """Add webhook"""
    user = g.user
    
    exec_db(
        "INSERT INTO webhooks (org_id, event, url, secret, active, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (
            user["org_id"],
            request.form.get("event", ""),
            request.form.get("url", ""),
            request.form.get("secret", ""),
            1,
            utc_now()
        )
    )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/webhook/delete/<int:wid>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_webhook_delete(wid):
    """Delete webhook"""
    user = g.user
    
    exec_db("DELETE FROM webhooks WHERE id=? AND org_id=?", (wid, user["org_id"]))
    
    return redirect(url_for("settings_page"))


@app.route("/settings/webhook/test/<int:wid>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_webhook_test(wid):
    """Test webhook"""
    user = g.user
    
    webhook = query_db("SELECT * FROM webhooks WHERE id=? AND org_id=?", (wid, user["org_id"]), one=True)
    if not webhook:
        return jsonify(ok=False, error="Not found"), 404
    
    # Queue test event
    trigger_webhook(user["org_id"], webhook["event"], {"test": True, "timestamp": utc_now()})
    
    return jsonify(ok=True)


@app.route("/settings/user/add", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_user_add():
    """Add user (admin only)"""
    user = g.user
    
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    
    if not username or len(password) < 12:
        return redirect(url_for("settings_page"))
    
    password_hash = hash_password(password)
    
    exec_db(
        """
        INSERT INTO users (org_id, username, email, password_hash, role, department_id, active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["org_id"],
            username,
            request.form.get("email", ""),
            password_hash,
            request.form.get("role", "agent"),
            request.form.get("department_id") or None,
            1,
            utc_now()
        )
    )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/user/toggle/<int:uid>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_user_toggle(uid):
    """Toggle user active status"""
    admin = g.user
    
    user_row = query_db("SELECT active FROM users WHERE id=? AND org_id=?", (uid, admin["org_id"]), one=True)
    if user_row:
        new_active = 0 if user_row["active"] else 1
        exec_db("UPDATE users SET active=? WHERE id=?", (new_active, uid))
    
    return redirect(url_for("settings_page"))


@app.route("/settings/department/add", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_department_add():
    """Add department"""
    user = g.user
    
    name = request.form.get("name", "").strip()
    if name:
        slug = re.sub(r'[^a-z0-9]+', '_', name.lower())
        exec_db(
            "INSERT INTO departments (org_id, name, slug, created_at) VALUES (?, ?, ?, ?)",
            (user["org_id"], name, slug, utc_now())
        )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/task_status/add", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_task_status_add():
    """Add task status"""
    user = g.user
    
    name = request.form.get("name", "").strip()
    if name:
        exec_db(
            "INSERT INTO task_statuses (org_id, name, created_at) VALUES (?, ?, ?)",
            (user["org_id"], name, utc_now())
        )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/task_status/delete/<int:sid>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_task_status_delete(sid):
    """Delete task status"""
    user = g.user
    
    exec_db("DELETE FROM task_statuses WHERE id=? AND org_id=?", (sid, user["org_id"]))
    
    return redirect(url_for("settings_page"))


# ===== BLOCK: METRICS ENDPOINT =====

@app.route("/metrics")
def metrics():
    """Prometheus-compatible metrics endpoint"""
    output = []
    
    with _metrics_lock:
        output.append(f"# HELP crm_requests_total Total HTTP requests")
        output.append(f"# TYPE crm_requests_total counter")
        output.append(f"crm_requests_total {_metrics['requests_total']}")
        
        output.append(f"# HELP crm_errors_total Total errors")
        output.append(f"# TYPE crm_errors_total counter")
        output.append(f"crm_errors_total {_metrics['errors_total']}")
        
        # Requests by endpoint
        for key, count in _metrics["requests_by_endpoint"].items():
            labels = dict(key) if isinstance(key, frozenset) else {}
            endpoint = labels.get("endpoint", "unknown")
            output.append(f'crm_requests_by_endpoint{{endpoint="{endpoint}"}} {count}')
    
    return Response("\n".join(output), mimetype="text/plain")


@app.route("/health")
def health():
    """Health check endpoint"""
    return jsonify(ok=True, status="healthy", version=SCHEMA_VERSION)


@app.route("/readyz")
def readyz():
    """Readiness check"""
    try:
        # Test DB
        query_db("SELECT 1", one=True)
        
        # Test Redis (optional)
        r = get_redis()
        if r:
            r.ping()
        
        return jsonify(ok=True, ready=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 503


# ===== END OF CORE PART 7/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 8/10 — CTI, EMAIL WORKERS, AI JOBS, WORKFLOW ENGINE, WORKERS
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: CTI INTEGRATION — WEBHOOK HANDLERS =====

@app.route("/cti/event", methods=["POST"])
@_rate_limit("cti_webhook", per_min=300)
def cti_event():
    """
    CTI webhook handler for telephony events.
    Supports: Mango, UIS, TELFIN (configurable via channel config)
    """
    # Get channel from header or query param
    channel_id = request.args.get("channel_id") or request.headers.get("X-Channel-ID")
    
    if not channel_id:
        return jsonify(ok=False, error="Channel ID required"), 400
    
    # Get channel config
    channel = query_db("SELECT * FROM channels WHERE id=? AND type='phone' AND active=1", (channel_id,), one=True)
    if not channel:
        return jsonify(ok=False, error="Channel not found"), 404
    
    # Verify signature (if configured)
    if channel["secret"]:
        signature = request.headers.get("X-Signature", "")
        expected = hmac.new(
            channel["secret"].encode(),
            request.get_data(),
            hashlib.sha256
        ).hexdigest()
        
        if not hmac.compare_digest(signature, expected):
            log("WARN", "CTI webhook signature mismatch")
            return jsonify(ok=False, error="Invalid signature"), 403
    
    # Parse event data
    data = request.get_json() or {}
    event_type = data.get("event", "")
    
    org_id = channel["org_id"]
    
    try:
        if event_type == "call.incoming":
            _handle_call_incoming(org_id, channel["id"], data)
        elif event_type == "call.answered":
            _handle_call_answered(org_id, channel["id"], data)
        elif event_type == "call.ended":
            _handle_call_ended(org_id, channel["id"], data)
        elif event_type == "call.recording.ready":
            _handle_call_recording_ready(org_id, channel["id"], data)
        else:
            log("WARN", f"Unknown CTI event: {event_type}")
    
    except Exception as e:
        log("ERROR", f"CTI event handler failed: {e}")
        return jsonify(ok=False, error="Processing failed"), 500
    
    return jsonify(ok=True)


def _handle_call_incoming(org_id: int, channel_id: int, data: dict):
    """Handle incoming call event"""
    call_id_external = data.get("call_id", "")
    from_number = normalize_phone(data.get("from", ""))
    to_number = normalize_phone(data.get("to", ""))
    
    # Check if call already exists
    existing = query_db(
        "SELECT id FROM calls WHERE org_id=? AND external_call_id=?",
        (org_id, call_id_external),
        one=True
    )
    
    if existing:
        return  # Already processed
    
    # Lookup company/contact by phone
    lookup = lookup_by_phone(org_id, from_number)
    company_id = lookup["companies"][0]["id"] if lookup["companies"] else None
    contact_id = lookup["contacts"][0]["id"] if lookup["contacts"] else None
    
    # Create call record
    call_id = exec_db(
        """
        INSERT INTO calls (
            org_id, channel_id, external_call_id, direction, from_e164, to_e164,
            company_id, contact_id, status, started_at, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            org_id, channel_id, call_id_external, "inbound", from_number, to_number,
            company_id, contact_id, "ringing", utc_now(), utc_now()
        )
    )
    
    # Trigger webhook
    trigger_webhook(org_id, "call.incoming", {
        "call_id": call_id,
        "from": from_number,
        "to": to_number,
        "company_id": company_id,
        "contact_id": contact_id
    })
    
    # Push SSE to online agents for screen pop
    agents = query_db("SELECT id FROM users WHERE org_id=? AND active=1", (org_id,))
    for agent in agents:
        sse_push(agent["id"], "call.incoming", {
            "call_id": call_id,
            "from": from_number,
            "company_name": lookup["companies"][0]["name"] if lookup["companies"] else "",
            "contact_name": lookup["contacts"][0]["name"] if lookup["contacts"] else ""
        })


def _handle_call_answered(org_id: int, channel_id: int, data: dict):
    """Handle call answered event"""
    call_id_external = data.get("call_id", "")
    agent_extension = data.get("agent_extension", "")
    
    call = query_db(
        "SELECT id FROM calls WHERE org_id=? AND external_call_id=?",
        (org_id, call_id_external),
        one=True
    )
    
    if not call:
        return
    
    # Find agent by extension (stored in user.phone or custom field)
    agent = query_db(
        "SELECT id FROM users WHERE org_id=? AND phone LIKE ?",
        (org_id, f"%{agent_extension}%"),
        one=True
    )
    
    # Update call
    exec_db(
        "UPDATE calls SET status='answered', agent_id=? WHERE id=?",
        (agent["id"] if agent else None, call["id"])
    )


def _handle_call_ended(org_id: int, channel_id: int, data: dict):
    """Handle call ended event"""
    call_id_external = data.get("call_id", "")
    duration_sec = int(data.get("duration", 0))
    
    call = query_db(
        "SELECT id FROM calls WHERE org_id=? AND external_call_id=?",
        (org_id, call_id_external),
        one=True
    )
    
    if not call:
        return
    
    exec_db(
        "UPDATE calls SET status='completed', duration_sec=?, ended_at=? WHERE id=?",
        (duration_sec, utc_now(), call["id"])
    )
    
    # Trigger webhook
    trigger_webhook(org_id, "call.ended", {
        "call_id": call["id"],
        "duration_sec": duration_sec
    })


def _handle_call_recording_ready(org_id: int, channel_id: int, data: dict):
    """Handle call recording ready event"""
    call_id_external = data.get("call_id", "")
    recording_url = data.get("recording_url", "")
    
    call = query_db(
        "SELECT id FROM calls WHERE org_id=? AND external_call_id=?",
        (org_id, call_id_external),
        one=True
    )
    
    if not call:
        return
    
    exec_db(
        "UPDATE calls SET recording_url=? WHERE id=?",
        (recording_url, call["id"])
    )


@app.route("/cti/recording/<int:rec_id>", methods=["GET", "POST"])
@_login_required
@_rate_limit("cti_recording", per_min=30)
def cti_recording(rec_id):
    """
    Fetch and proxy call recording.
    ✅ УЛУЧШЕНО: SSRF mitigation с DNS validation
    """
    user = g.user
    
    # Get call
    call = query_db(
        "SELECT * FROM calls WHERE id=? AND org_id=?",
        (rec_id, user["org_id"]),
        one=True
    )
    
    if not call or not call["recording_url"]:
        return jsonify(ok=False, error="Recording not found"), 404
    
    rec_url = call["recording_url"]
    
    # ✅ НОВОЕ: SSRF protection
    resolved_url, is_safe = _resolve_and_validate_url(rec_url)
    if not is_safe:
        return jsonify(ok=False, error="Untrusted recording host"), 400
    
    try:
        # Fetch recording from provider
        parsed = urllib.parse.urlparse(rec_url)
        
        with _rq.get(
            resolved_url,
            timeout=15,
            stream=True,
            headers={"Host": parsed.hostname} if parsed.hostname else {}
        ) as r:
            r.raise_for_status()
            
            # Stream to client
            def generate():
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        yield chunk
            
            return Response(
                generate(),
                mimetype=r.headers.get("Content-Type", "audio/mpeg"),
                headers={
                    "Content-Disposition": f'attachment; filename="recording_{rec_id}.mp3"'
                }
            )
    
    except Exception as e:
        log("ERROR", f"Recording fetch failed: {e}")
        return jsonify(ok=False, error="Failed to fetch recording"), 500


@app.route("/api/call/recording/presign/<int:call_id>", methods=["GET"])
@_login_required
def api_call_recording_presign(call_id):
    """
    Generate presigned URL for call recording.
    Alternative to proxy - returns direct link (if safe).
    """
    user = g.user
    
    call = query_db(
        "SELECT recording_url FROM calls WHERE id=? AND org_id=?",
        (call_id, user["org_id"]),
        one=True
    )
    
    if not call or not call["recording_url"]:
        return jsonify(ok=False, error="No recording"), 404
    
    # For now, just return the URL (in production, generate presigned S3 URL if stored in S3)
    return jsonify(ok=True, url=call["recording_url"])


# ===== BLOCK: EMAIL WORKERS =====

def _email_fetch_once():
    """
    Fetch emails from all configured IMAP accounts.
    Called by background worker.
    ✅ УЛУЧШЕНО: UIDVALIDITY handling добавлен в CORE PART 3
    """
    # Get all channels with type=email
    channels = query_db("SELECT * FROM channels WHERE type='email' AND active=1")
    
    for channel in channels:
        try:
            config = json.loads(channel["config_json"]) if channel["config_json"] else {}
            
            if not config.get("imap_host"):
                continue
            
            # Fetch emails
            emails = fetch_emails_imap(config)
            
            if not emails:
                continue
            
            # Process each email
            for email_data in emails:
                _process_incoming_email(channel["org_id"], channel["id"], email_data)
            
            # Update config with new last_uid
            exec_db(
                "UPDATE channels SET config_json=? WHERE id=?",
                (json.dumps(config), channel["id"])
            )
        
        except Exception as e:
            log("ERROR", f"Email fetch failed for channel {channel['id']}: {e}")


def _process_incoming_email(org_id: int, channel_id: int, email_data: dict):
    """Process incoming email - create thread and message"""
    from_addr = email_data.get("from", "")
    subject = email_data.get("subject", "No subject")
    body = email_data.get("body", "")
    
    # Extract email address
    email_match = re.search(r'[\w\.-]+@[\w\.-]+', from_addr)
    sender_email = email_match.group(0) if email_match else from_addr
    
    # Lookup company/contact by email
    lookup = lookup_by_email(org_id, sender_email)
    company_id = lookup["companies"][0]["id"] if lookup["companies"] else None
    contact_id = lookup["contacts"][0]["id"] if lookup["contacts"] else None
    
    # Check if thread exists (by subject or external ID)
    thread = query_db(
        "SELECT id FROM inbox_threads WHERE org_id=? AND channel_id=? AND subject=?",
        (org_id, channel_id, subject),
        one=True
    )
    
    if not thread:
        # Create new thread
        thread_id = create_inbox_thread(
            org_id=org_id,
            channel_id=channel_id,
            subject=subject,
            company_id=company_id,
            contact_id=contact_id
        )
    else:
        thread_id = thread["id"]
    
    # Add message
    add_message(
        thread_id=thread_id,
        sender_type="client",
        body=body,
        external_user_id=sender_email
    )


def _email_sequence_worker():
    """
    Process email sequences (drip campaigns).
    ✅ НОВОЕ: автоматические email-цепочки
    """
    try:
        # Find enrollments ready for next step
        enrollments = query_db(
            """
            SELECT e.*, s.delay_hours, s.subject, s.body_template, s.send_condition
            FROM sequence_enrollments e
            JOIN sequence_steps s ON s.sequence_id = e.sequence_id 
                AND s.step_num = e.current_step + 1
            WHERE e.status = 'active'
                AND (e.last_sent_at IS NULL OR 
                     datetime(e.last_sent_at, '+' || s.delay_hours || ' hours') <= datetime('now'))
            LIMIT 100
            """
        )
        
        for enr in enrollments:
            try:
                # Check send condition (if any)
                if enr["send_condition"]:
                    # Parse condition (simple JSON: {"field": "status", "op": "==", "value": "open"})
                    try:
                        cond = json.loads(enr["send_condition"])
                        # For now, skip complex conditions - placeholder for future
                        # In production, evaluate against contact/company data
                    except json.JSONDecodeError:
                        pass
                
                # Render template (simple replacement)
                subject = enr["subject"] or "Update"
                body = enr["body_template"] or ""
                
                # Replace variables (basic)
                body = body.replace("{{ email }}", enr["email"])
                
                # Send email
                if send_email(enr["email"], subject, body, html=True):
                    # Update enrollment
                    exec_db(
                        """
                        UPDATE sequence_enrollments 
                        SET current_step = current_step + 1, last_sent_at = ?
                        WHERE id = ?
                        """,
                        (utc_now(), enr["id"])
                    )
                    
                    log("INFO", f"Email sequence step sent: enrollment {enr['id']}")
            
            except Exception as e:
                log("ERROR", f"Email sequence step failed: {e}")
    
    except Exception as e:
        log("ERROR", f"Email sequence worker error: {e}")


# ===== BLOCK: AI JOBS PROCESSING =====

def _ai_job_process_pending():
    """
    Process pending AI jobs asynchronously.
    Called by background worker.
    """
    jobs = query_db(
        "SELECT * FROM ai_jobs WHERE status='pending' LIMIT 5"
    )
    
    for job in jobs:
        try:
            # Mark as processing
            exec_db("UPDATE ai_jobs SET status='processing' WHERE id=?", (job["id"],))
            
            result = None
            
            if job["job_type"] == "summarize":
                result = _ai_job_summarize(job)
            elif job["job_type"] == "draft_reply":
                result = _ai_job_draft(job)
            elif job["job_type"] == "extract_task":
                result = _ai_job_extract_task(job)
            elif job["job_type"] == "autotag":
                result = _ai_job_autotag(job)
            else:
                raise ValueError(f"Unknown job type: {job['job_type']}")
            
            # Update job as completed
            exec_db(
                "UPDATE ai_jobs SET status='completed', output_text=?, completed_at=? WHERE id=?",
                (result, utc_now(), job["id"])
            )
        
        except Exception as e:
            log("ERROR", f"AI job {job['id']} failed: {e}")
            exec_db(
                "UPDATE ai_jobs SET status='failed', error=? WHERE id=?",
                (str(e)[:500], job["id"])
            )


def _ai_job_summarize(job: dict) -> str:
    """Execute summarize job"""
    prompt = f"Summarize this text:\n\n{job['input_text']}"
    return ai_provider_call(prompt, system="You are a helpful summarization assistant.")


def _ai_job_draft(job: dict) -> str:
    """Execute draft reply job"""
    prompt = f"Draft a professional reply to:\n\n{job['input_text']}"
    return ai_provider_call(prompt, system="You are a customer support assistant.")


def _ai_job_extract_task(job: dict) -> str:
    """Extract task from text"""
    prompt = f"Extract actionable tasks from this text. Format as JSON array:\n\n{job['input_text']}"
    return ai_provider_call(prompt, system="You extract tasks from conversations.")


def _ai_job_autotag(job: dict) -> str:
    """Auto-generate tags"""
    prompt = f"Extract 3-5 tags from this text (comma-separated):\n\n{job['input_text']}"
    return ai_provider_call(prompt, max_tokens=50)


# ===== BLOCK: WORKFLOW ENGINE =====

def _wf_execute_node(workflow_id: int, node_id: str, ctx: dict, payload: dict):
    """
    Execute workflow node.
    ✅ ИСПРАВЛЕНО: дописана функция полностью
    """
    # Get workflow definition
    workflow = query_db("SELECT * FROM workflow_definitions WHERE id=?", (workflow_id,), one=True)
    if not workflow:
        raise ValueError("Workflow not found")
    
    graph = json.loads(workflow["graph_json"])
    node = graph.get("nodes", {}).get(node_id)
    
    if not node:
        raise ValueError(f"Node {node_id} not found")
    
    node_type = node.get("type")
    config = node.get("config", {})
    
    # Execute based on type
    if node_type == "task.create":
        # Create task
        title = _wf_eval_value(config.get("title", "Task"), ctx, payload)
        description = _wf_eval_value(config.get("description", ""), ctx, payload)
        assignee_id = _wf_eval_value(config.get("assignee_id"), ctx, payload)
        
        task_id = create_task(
            org_id=workflow["org_id"],
            title=str(title),
            description=str(description),
            assignee_id=int(assignee_id) if assignee_id else None
        )
        
        ctx["task_id"] = task_id
    
    elif node_type == "email.send":
        # Send email
        to = _wf_eval_value(config.get("to", ""), ctx, payload)
        subject = _wf_eval_value(config.get("subject", ""), ctx, payload)
        body = _wf_eval_value(config.get("body", ""), ctx, payload)
        
        send_email(str(to), str(subject), str(body))
    
    elif node_type == "webhook.call":
        # Trigger webhook
        url = _wf_eval_value(config.get("url", ""), ctx, payload)
        webhook_payload = _wf_eval_value(config.get("payload", {}), ctx, payload)
        
        _rq.post(str(url), json=webhook_payload, timeout=10)
    
    elif node_type == "condition":
        # Evaluate condition
        condition = config.get("condition", "")
        result = _wf_eval_condition(condition, ctx, payload)
        ctx["condition_result"] = result
    
    elif node_type == "delay":
        # Schedule next node
        delay_minutes = int(config.get("delay_minutes", 0))
        if delay_minutes > 0:
            scheduled_at = (datetime.utcnow() + timedelta(minutes=delay_minutes)).isoformat(" ", "seconds")
            # Create task for next node
            next_nodes = node.get("next", [])
            for next_id in next_nodes:
                exec_db(
                    """
                    INSERT INTO workflow_tasks (workflow_id, node_id, context_json, status, scheduled_at, created_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (workflow_id, next_id, json.dumps(ctx), "pending", scheduled_at, utc_now())
                )
            return  # Don't continue to next nodes immediately
    
    # Continue to next nodes
    next_nodes = node.get("next", [])
    for next_id in next_nodes:
        # Check if condition node - use condition result
        if node_type == "condition":
            if next_id == node.get("true_next") and not ctx.get("condition_result"):
                continue
            if next_id == node.get("false_next") and ctx.get("condition_result"):
                continue
        
        # Execute next node
        _wf_execute_node(workflow_id, next_id, ctx, payload)


def _wf_execute_once():
    """
    Execute pending workflow tasks.
    Called by background worker.
    ✅ ИСПРАВЛЕНО: добавлена логика выполнения
    """
    # Get pending tasks (due now)
    tasks = query_db(
        """
        SELECT * FROM workflow_tasks
        WHERE status='pending' AND (scheduled_at IS NULL OR scheduled_at <= datetime('now'))
        ORDER BY created_at ASC
        LIMIT 10
        """
    )
    
    for task in tasks:
        try:
            # Mark as processing (prevent duplicate execution)
            rows_updated = exec_db(
                "UPDATE workflow_tasks SET status='processing', started_at=? WHERE id=? AND status='pending'",
                (utc_now(), task["id"])
            )
            
            # If 0 rows updated, another worker took it
            if rows_updated == 0:
                continue
            
            # Parse context
            ctx = json.loads(task["context_json"]) if task["context_json"] else {}
            
            # Get entity payload
            payload = {}
            if task["entity_type"] and task["entity_id"]:
                if task["entity_type"] == "task":
                    entity = query_db("SELECT * FROM tasks WHERE id=?", (task["entity_id"],), one=True)
                    payload = dict(entity) if entity else {}
                elif task["entity_type"] == "deal":
                    entity = query_db("SELECT * FROM deals WHERE id=?", (task["entity_id"],), one=True)
                    payload = dict(entity) if entity else {}
            
            # Execute node
            _wf_execute_node(task["workflow_id"], task["node_id"], ctx, payload)
            
            # Mark as completed
            exec_db(
                "UPDATE workflow_tasks SET status='completed', completed_at=? WHERE id=?",
                (utc_now(), task["id"])
            )
        
        except Exception as e:
            log("ERROR", f"Workflow task {task['id']} failed: {e}")
            exec_db(
                "UPDATE workflow_tasks SET status='failed', error=?, completed_at=? WHERE id=?",
                (str(e)[:500], utc_now(), task["id"])
            )


# ===== BLOCK: BACKGROUND WORKERS =====

_workers_started = False
_workers_lock = threading.Lock()
_shutdown_event = threading.Event()


def _shutdown_handler(signum, frame):
    """Signal handler for graceful shutdown"""
    log("INFO", "Shutdown signal received, stopping workers...")
    _shutdown_event.set()


# Register signal handlers
import signal
signal.signal(signal.SIGTERM, _shutdown_handler)
signal.signal(signal.SIGINT, _shutdown_handler)


def webhook_worker():
    """
    Webhook delivery worker.
    Runs in background thread.
    """
    log("INFO", "webhook_worker started")
    
    while not _shutdown_event.is_set():
        try:
            _webhook_deliver_once()
        except Exception as e:
            log("ERROR", f"webhook_worker error: {e}")
        
        # Sleep 10 seconds or until shutdown
        _shutdown_event.wait(10)
    
    log("INFO", "webhook_worker stopped")


def maintenance_worker():
    """
    Maintenance worker: cleanup, FTS rebuild, etc.
    Runs in background thread.
    ✅ ИСПРАВЛЕНО: дописана логика
    """
    log("INFO", "maintenance_worker started")
    
    while not _shutdown_event.is_set():
        try:
            # Cleanup old audit logs (older than 1 year)
            exec_db(
                "DELETE FROM audit_logs WHERE created_at < datetime('now', '-1 year')"
            )
            
            # Cleanup old webhook queue (older than 30 days)
            exec_db(
                "DELETE FROM webhook_queue WHERE created_at < datetime('now', '-30 days')"
            )
            
            # Process reminders
            _process_reminders()
            
            # Execute workflow tasks
            _wf_execute_once()
            
            # Process AI jobs
            _ai_job_process_pending()
            
            # Process email sequences
            _email_sequence_worker()
            
            # Fetch emails (every hour, check if it's time)
            current_minute = datetime.utcnow().minute
            if current_minute < 5:  # Run at :00-:05
                _email_fetch_once()
        
        except Exception as e:
            log("ERROR", f"maintenance_worker error: {e}")
        
        # Sleep 60 seconds
        _shutdown_event.wait(60)
    
    log("INFO", "maintenance_worker stopped")


def _process_reminders():
    """
    Process due reminders and send notifications.
    """
    reminders = query_db(
        """
        SELECT r.*, t.title, u.username
        FROM task_reminders r
        JOIN tasks t ON r.task_id = t.id
        LEFT JOIN users u ON r.user_id = u.id
        WHERE r.sent=0 AND r.remind_at <= datetime('now')
        LIMIT 100
        """
    )
    
    for reminder in reminders:
        try:
            # Push SSE notification
            sse_push(reminder["user_id"], "task.reminder", {
                "reminder_id": reminder["id"],
                "task_id": reminder["task_id"],
                "task_title": reminder["title"],
                "message": reminder["message"] or f"Напоминание: {reminder['title']}"
            })
            
            # Mark as sent
            exec_db("UPDATE task_reminders SET sent=1 WHERE id=?", (reminder["id"],))
            
            log("INFO", f"Reminder sent: {reminder['id']}")
        
        except Exception as e:
            log("ERROR", f"Reminder processing failed: {e}")


def start_workers_once():
    """
    Start background workers (idempotent - only starts once).
    """
    global _workers_started
    
    with _workers_lock:
        if _workers_started:
            return
        
        # Start webhook worker
        webhook_thread = threading.Thread(target=webhook_worker, daemon=True, name="webhook_worker")
        webhook_thread.start()
        
        # Start maintenance worker
        maintenance_thread = threading.Thread(target=maintenance_worker, daemon=True, name="maintenance_worker")
        maintenance_thread.start()
        
        _workers_started = True
        log("INFO", "Background workers started")


# ===== BLOCK: API ROUTES — ADDITIONAL ADMIN/SETTINGS =====

@app.route("/settings/ai/config", methods=["GET"])
@_login_required
@_require_role("admin")
def api_settings_ai_config_get():
    """Get current AI configuration"""
    user = g.user
    
    org = query_db("SELECT settings_json FROM orgs WHERE id=?", (user["org_id"],), one=True)
    if not org:
        return jsonify(ok=False, error="Org not found"), 404
    
    settings = json.loads(org["settings_json"]) if org["settings_json"] else {}
    ai_config = settings.get("ai", {})
    
    return jsonify(ok=True, config=ai_config)


@app.route("/settings/ai/config", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_ai_config():
    """Update AI configuration"""
    user = g.user
    
    org = query_db("SELECT settings_json FROM orgs WHERE id=?", (user["org_id"],), one=True)
    settings = json.loads(org["settings_json"]) if org["settings_json"] else {}
    
    # Update AI config
    settings["ai"] = {
        "provider": request.form.get("provider", ""),
        "model": request.form.get("model", ""),
        "temperature": float(request.form.get("temperature", 0.3)),
        "max_tokens": int(request.form.get("max_tokens", 512)),
        "policy": json.loads(request.form.get("policy", "{}"))
    }
    
    exec_db(
        "UPDATE orgs SET settings_json=? WHERE id=?",
        (json.dumps(settings), user["org_id"])
    )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/user/password/<int:uid>", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_user_password(uid):
    """Change user password (admin)"""
    admin = g.user
    
    # Verify user in same org
    user = query_db("SELECT id FROM users WHERE id=? AND org_id=?", (uid, admin["org_id"]), one=True)
    if not user:
        return "User not found", 404
    
    password = request.form.get("password", "")
    if len(password) < 12:
        return "Password too short", 400
    
    password_hash = hash_password(password)
    exec_db("UPDATE users SET password_hash=? WHERE id=?", (password_hash, uid))
    
    add_audit(admin["org_id"], admin["id"], "user.password_changed_by_admin", "user", uid)
    
    return redirect(url_for("settings_page"))


@app.route("/settings/phone/<int:channel_id>/update", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def settings_phone_update(channel_id):
    """Update phone channel configuration"""
    user = g.user
    
    channel = query_db("SELECT * FROM channels WHERE id=? AND org_id=? AND type='phone'", (channel_id, user["org_id"]), one=True)
    if not channel:
        return "Channel not found", 404
    
    # Update config
    config = {
        "provider": request.form.get("provider", ""),
        "from_e164": request.form.get("from_e164", ""),
        "signing_key": request.form.get("signing_key", "")
    }
    
    secret = request.form.get("secret", "")
    
    exec_db(
        "UPDATE channels SET config_json=?, secret=? WHERE id=?",
        (json.dumps(config), secret or None, channel_id)
    )
    
    return redirect(url_for("settings_page"))


@app.route("/settings/phone/<int:channel_id>/webhook_urls", methods=["GET"])
@_login_required
@_require_role("admin")
def settings_phone_webhook_urls(channel_id):
    """Get webhook URLs for phone channel"""
    user = g.user
    
    channel = query_db("SELECT id FROM channels WHERE id=? AND org_id=? AND type='phone'", (channel_id, user["org_id"]), one=True)
    if not channel:
        return jsonify(ok=False, error="Channel not found"), 404
    
    base_url = request.url_root.rstrip('/')
    
    return jsonify(
        ok=True,
        cti_webhook=f"{base_url}/cti/event?channel_id={channel_id}",
        recording_webhook=f"{base_url}/cti/recording/{{call_id}}"
    )


@app.route("/api/tokens/list", methods=["GET"])
@_login_required
@_require_role("admin")
def api_tokens_list():
    """List API tokens"""
    user = g.user
    
    tokens = query_db(
        "SELECT id, name, user_id, scopes, active, expires_at, created_at, last_used_at FROM api_tokens WHERE org_id=?",
        (user["org_id"],)
    )
    
    return jsonify(ok=True, items=[dict(t) for t in tokens])


@app.route("/api/tokens/create", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_tokens_create():
    """Create API token"""
    user = g.user
    data = request.get_json() or {}
    
    name = data.get("name", "API Token")
    user_id = data.get("user_id")
    scopes = data.get("scopes", [])
    expires_at = data.get("expires_at")
    
    # Generate token
    token = generate_api_token()
    token_hash = hash_api_token(token)
    
    # Store
    exec_db(
        """
        INSERT INTO api_tokens (org_id, user_id, name, token_hash, scopes, active, expires_at, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user["org_id"],
            user_id,
            name,
            token_hash,
            ",".join(scopes) if isinstance(scopes, list) else scopes,
            1,
            ensure_iso_datetime(expires_at) if expires_at else None,
            utc_now()
        )
    )
    
    add_audit(user["org_id"], user["id"], "api_token.created", details={"name": name})
    
    # Return token (only time it's visible)
    return jsonify(ok=True, token=token)


@app.route("/api/tokens/toggle", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_tokens_toggle():
    """Toggle token active status"""
    user = g.user
    data = request.get_json() or {}
    
    token_id = data.get("id")
    if not token_id:
        return jsonify(ok=False, error="Token ID required"), 400
    
    token = query_db("SELECT active FROM api_tokens WHERE id=? AND org_id=?", (token_id, user["org_id"]), one=True)
    if not token:
        return jsonify(ok=False, error="Token not found"), 404
    
    new_active = 0 if token["active"] else 1
    exec_db("UPDATE api_tokens SET active=? WHERE id=?", (new_active, token_id))
    
    return jsonify(ok=True, active=bool(new_active))


@app.route("/api/search/quick", methods=["GET"])
@_login_required
def api_search_quick():
    """
    Quick search for command palette.
    ✅ НОВОЕ: для Cmd+K функционала
    """
    user = g.user
    q = request.args.get("q", "").strip()
    
    if len(q) < 2:
        return jsonify(ok=True, items=[])
    
    results = []
    
    # Search tasks
    tasks = query_db(
        """
        SELECT id, title FROM tasks
        WHERE org_id=? AND (title LIKE ? OR description LIKE ?)
        LIMIT 5
        """,
        (user["org_id"], f"%{q}%", f"%{q}%")
    )
    results.extend([{"type": "task", "title": t["title"], "url": f"/task/{t['id']}"} for t in tasks])
    
    # Search deals
    deals = query_db(
        "SELECT id, title FROM deals WHERE org_id=? AND title LIKE ? LIMIT 5",
        (user["org_id"], f"%{q}%")
    )
    results.extend([{"type": "deal", "title": d["title"], "url": f"/deal/{d['id']}"} for d in deals])
    
    # Search companies
    companies = query_db(
        "SELECT id, name FROM companies WHERE org_id=? AND (name LIKE ? OR inn LIKE ?) LIMIT 5",
        (user["org_id"], f"%{q}%", f"%{q}%")
    )
    results.extend([{"type": "company", "title": c["name"], "url": f"/client/{c['id']}"} for c in companies])
    
    return jsonify(ok=True, items=results[:15])


# ===== END OF CORE PART 8/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 9A/10 — CTI PROVIDERS, APPROVALS, ENHANCED APIs
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: CTI PROVIDERS — MANGO OFFICE =====

def _cti_mango_normalize_event(raw_data: dict) -> dict:
    """
    Normalize Mango Office webhook event to internal format.
    Docs: https://mango-office.ru/support/api/
    """
    event_type_map = {
        "call": "call.incoming",
        "summary": "call.ended",
        "recording": "call.recording.ready"
    }
    
    event_type = event_type_map.get(raw_data.get("command_id"), "unknown")
    
    normalized = {
        "event": event_type,
        "call_id": raw_data.get("call_id", ""),
        "from": raw_data.get("from", {}).get("number", ""),
        "to": raw_data.get("to", {}).get("number", ""),
        "duration": raw_data.get("total_time", 0),
        "recording_url": raw_data.get("recording_url", ""),
        "agent_extension": raw_data.get("to", {}).get("extension", "")
    }
    
    return normalized


def _cti_mango_verify_signature(data: bytes, signature: str, signing_key: str) -> bool:
    """Verify Mango Office webhook signature"""
    expected = hashlib.sha256((signing_key + data.decode()).encode()).hexdigest()
    return hmac.compare_digest(signature.lower(), expected.lower())


# ===== BLOCK: CTI PROVIDERS — UIS =====

def _cti_uis_normalize_event(raw_data: dict) -> dict:
    """
    Normalize UIS (Unitalk) webhook event.
    Docs: https://uis.tel/api
    """
    event_type_map = {
        "new_call": "call.incoming",
        "call_end": "call.ended",
        "record_ready": "call.recording.ready"
    }
    
    event_type = event_type_map.get(raw_data.get("event"), "unknown")
    
    normalized = {
        "event": event_type,
        "call_id": raw_data.get("virtual_id", ""),
        "from": raw_data.get("src_num", ""),
        "to": raw_data.get("dst_num", ""),
        "duration": int(raw_data.get("billsec", 0)),
        "recording_url": raw_data.get("record_url", ""),
        "agent_extension": raw_data.get("employee_ext", "")
    }
    
    return normalized


def _cti_uis_verify_signature(data: bytes, signature: str, signing_key: str) -> bool:
    """Verify UIS webhook signature"""
    expected = hmac.new(
        signing_key.encode(),
        data,
        hashlib.sha256
    ).hexdigest()
    return hmac.compare_digest(signature, expected)


# ===== BLOCK: CTI PROVIDERS — TELFIN =====

def _cti_telfin_normalize_event(raw_data: dict) -> dict:
    """
    Normalize TELFIN webhook event.
    Docs: https://telfin.ru/api/
    """
    event_type_map = {
        "call.new": "call.incoming",
        "call.completed": "call.ended",
        "recording.available": "call.recording.ready"
    }
    
    event_type = event_type_map.get(raw_data.get("event_type"), "unknown")
    
    normalized = {
        "event": event_type,
        "call_id": raw_data.get("call_uuid", ""),
        "from": raw_data.get("caller_number", ""),
        "to": raw_data.get("called_number", ""),
        "duration": int(raw_data.get("duration_sec", 0)),
        "recording_url": raw_data.get("recording_link", ""),
        "agent_extension": raw_data.get("extension", "")
    }
    
    return normalized


def _cti_telfin_verify_signature(data: bytes, signature: str, signing_key: str) -> bool:
    """Verify TELFIN webhook signature"""
    expected = hmac.new(
        signing_key.encode(),
        data,
        hashlib.sha256
    ).hexdigest()
    return hmac.compare_digest(signature, expected)


# ===== BLOCK: CTI PROVIDERS — UNIFIED HANDLER =====

@app.route("/cti/provider/<provider_name>", methods=["POST"])
@_rate_limit("cti_provider_webhook", per_min=500)
def cti_provider_webhook(provider_name):
    """
    Unified webhook handler for different CTI providers.
    Routes to provider-specific normalizers.
    """
    # Get channel by provider name (stored in config)
    channels = query_db(
        "SELECT * FROM channels WHERE type='phone' AND active=1"
    )
    
    matching_channel = None
    for ch in channels:
        config = json.loads(ch["config_json"]) if ch["config_json"] else {}
        if config.get("provider") == provider_name:
            matching_channel = ch
            break
    
    if not matching_channel:
        return jsonify(ok=False, error="Provider not configured"), 404
    
    # Get signature and verify
    raw_data = request.get_data()
    config = json.loads(matching_channel["config_json"]) if matching_channel["config_json"] else {}
    signing_key = config.get("signing_key", "")
    
    if signing_key:
        signature = request.headers.get("X-Signature") or request.headers.get("Signature", "")
        
        is_valid = False
        if provider_name == "mango":
            is_valid = _cti_mango_verify_signature(raw_data, signature, signing_key)
        elif provider_name == "uis":
            is_valid = _cti_uis_verify_signature(raw_data, signature, signing_key)
        elif provider_name == "telfin":
            is_valid = _cti_telfin_verify_signature(raw_data, signature, signing_key)
        
        if not is_valid:
            log("WARN", f"CTI {provider_name} signature verification failed")
            return jsonify(ok=False, error="Invalid signature"), 403
    
    # Parse and normalize event
    try:
        raw_event = request.get_json() or {}
        
        if provider_name == "mango":
            normalized = _cti_mango_normalize_event(raw_event)
        elif provider_name == "uis":
            normalized = _cti_uis_normalize_event(raw_event)
        elif provider_name == "telfin":
            normalized = _cti_telfin_normalize_event(raw_event)
        else:
            return jsonify(ok=False, error="Unknown provider"), 400
        
        # Process event using unified handler
        event_type = normalized.get("event")
        org_id = matching_channel["org_id"]
        channel_id = matching_channel["id"]
        
        if event_type == "call.incoming":
            _handle_call_incoming(org_id, channel_id, normalized)
        elif event_type == "call.answered":
            _handle_call_answered(org_id, channel_id, normalized)
        elif event_type == "call.ended":
            _handle_call_ended(org_id, channel_id, normalized)
        elif event_type == "call.recording.ready":
            _handle_call_recording_ready(org_id, channel_id, normalized)
        
        return jsonify(ok=True)
    
    except Exception as e:
        log("ERROR", f"CTI provider webhook failed: {e}")
        return jsonify(ok=False, error="Processing failed"), 500


# ===== BLOCK: APPROVAL WORKFLOWS =====

@app.route("/approve/<token>", methods=["GET", "POST"])
def approval_public(token):
    """
    Public approval link (no authentication required).
    Token-based access for external approvals.
    
    Example use case: Client approves quote/proposal via email link.
    """
    # Decode token (simple base64 for now, can use JWT in production)
    try:
        decoded = base64.b64decode(token).decode()
        parts = decoded.split(":")
        if len(parts) != 3:
            raise ValueError("Invalid token format")
        
        entity_type, entity_id, secret = parts
        entity_id = int(entity_id)
    except Exception:
        return "Invalid or expired approval link", 400
    
    # Get entity and verify secret
    if entity_type == "deal":
        entity = query_db("SELECT * FROM deals WHERE id=?", (entity_id,), one=True)
    elif entity_type == "task":
        entity = query_db("SELECT * FROM tasks WHERE id=?", (entity_id,), one=True)
    elif entity_type == "document":
        entity = query_db("SELECT * FROM documents WHERE id=?", (entity_id,), one=True)
    else:
        return "Unknown entity type", 400
    
    if not entity:
        return "Entity not found", 404
    
    # Verify secret (stored in entity metadata or generated from entity data)
    expected_secret = hashlib.sha256(f"{entity_type}:{entity_id}:{SECRET_KEY}".encode()).hexdigest()[:16]
    if secret != expected_secret:
        return "Invalid approval link", 403
    
    # Handle approval
    if request.method == "POST":
        decision = request.form.get("decision", "")  # "approve" or "reject"
        comment = request.form.get("comment", "")
        
        if decision not in ("approve", "reject"):
            return "Invalid decision", 400
        
        # Update entity
        if entity_type == "deal":
            new_status = "won" if decision == "approve" else "lost"
            exec_db("UPDATE deals SET status=? WHERE id=?", (new_status, entity_id))
        elif entity_type == "task":
            new_status = "done" if decision == "approve" else "cancelled"
            exec_db("UPDATE tasks SET status=? WHERE id=?", (new_status, entity_id))
        
        # Log activity
        exec_db(
            "INSERT INTO audit_logs (org_id, action, entity_type, entity_id, details, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (entity["org_id"], f"approval.{decision}", entity_type, entity_id, json.dumps({"comment": comment}), utc_now())
        )
        
        # Trigger webhook
        trigger_webhook(entity["org_id"], f"{entity_type}.approved", {
            "entity_id": entity_id,
            "decision": decision,
            "comment": comment
        })
        
        return f"""
        <html>
        <head><title>Approval Submitted</title></head>
        <body style="font-family: Arial, sans-serif; text-align: center; padding: 50px;">
            <h1>✓ Decision Submitted</h1>
            <p>Your {decision} has been recorded. Thank you!</p>
        </body>
        </html>
        """
    
    # Show approval form
    return render_template_string(
        APPROVAL_TMPL,
        entity_type=entity_type,
        entity=entity,
        token=token
    )


# ===== BLOCK: ENHANCED LOOKUP API =====

@app.route("/api/lookup/multi", methods=["POST"])
@_login_required
@_csrf_protect
def api_lookup_multi():
    """
    Multi-field lookup (batch search).
    ✅ НОВОЕ: поиск по нескольким критериям одновременно
    
    Body: {
        "phones": ["+79991234567", ...],
        "emails": ["test@example.com", ...],
        "inns": ["1234567890", ...]
    }
    """
    user = g.user
    data = request.get_json() or {}
    
    phones = data.get("phones", [])
    emails = data.get("emails", [])
    inns = data.get("inns", [])
    
    results = {
        "companies": [],
        "contacts": [],
        "grouped": {}
    }
    
    # Search by phones
    for phone in phones[:10]:  # Limit to 10
        lookup = lookup_by_phone(user["org_id"], phone)
        results["grouped"][phone] = lookup
        results["companies"].extend(lookup["companies"])
        results["contacts"].extend(lookup["contacts"])
    
    # Search by emails
    for email_addr in emails[:10]:
        lookup = lookup_by_email(user["org_id"], email_addr)
        results["grouped"][email_addr] = lookup
        results["companies"].extend(lookup["companies"])
        results["contacts"].extend(lookup["contacts"])
    
    # Search by INNs
    for inn in inns[:10]:
        companies = lookup_by_inn(user["org_id"], inn)
        results["grouped"][inn] = {"companies": companies, "contacts": []}
        results["companies"].extend(companies)
    
    # Deduplicate
    seen_companies = set()
    unique_companies = []
    for c in results["companies"]:
        if c["id"] not in seen_companies:
            seen_companies.add(c["id"])
            unique_companies.append(c)
    
    seen_contacts = set()
    unique_contacts = []
    for c in results["contacts"]:
        if c["id"] not in seen_contacts:
            seen_contacts.add(c["id"])
            unique_contacts.append(c)
    
    results["companies"] = unique_companies
    results["contacts"] = unique_contacts
    
    return jsonify(ok=True, **results)


# ===== BLOCK: ENHANCED THREAD/MESSAGE API =====

@app.route("/api/thread/messages", methods=["GET"])
@_login_required
def api_thread_messages():
    """
    Get thread messages with pagination.
    ✅ НОВОЕ: отдельный endpoint для load more
    """
    user = g.user
    
    thread_id = request.args.get("thread_id")
    page = int(request.args.get("page", 1))
    per_page = min(int(request.args.get("per_page", 50)), 100)
    offset = (page - 1) * per_page
    
    if not thread_id:
        return jsonify(ok=False, error="Thread ID required"), 400
    
    # Verify thread access
    thread = query_db(
        "SELECT id FROM inbox_threads WHERE id=? AND org_id=?",
        (thread_id, user["org_id"]),
        one=True
    )
    
    if not thread:
        return jsonify(ok=False, error="Thread not found"), 404
    
    # Get messages
    messages = query_db(
        """
        SELECT 
            m.*,
            u.username
        FROM inbox_messages m
        LEFT JOIN users u ON m.user_id = u.id
        WHERE m.thread_id=?
        ORDER BY m.created_at DESC
        LIMIT ? OFFSET ?
        """,
        (thread_id, per_page, offset)
    )
    
    # Get attachments for each message
    items = []
    for msg in messages:
        m = dict(msg)
        
        attachments = query_db(
            """
            SELECT f.id, f.name, f.storage_key, f.size_bytes
            FROM message_attachments a
            JOIN files f ON a.file_id = f.id
            WHERE a.message_id=?
            """,
            (m["id"],)
        )
        
        m["attachments"] = [
            {
                "id": a["id"],
                "name": a["name"],
                "url": f"/api/files/{a['id']}/download",
                "size": a["size_bytes"]
            }
            for a in attachments
        ]
        
        items.append(m)
    
    return jsonify(ok=True, items=items, page=page, per_page=per_page)


@app.route("/api/message/upload", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("message_upload", per_min=30)
def api_message_upload():
    """
    Upload attachment for message (before sending).
    Returns file metadata.
    """
    user = g.user
    
    if "file" not in request.files:
        return jsonify(ok=False, error="No file"), 400
    
    f = request.files["file"]
    if not f.filename:
        return jsonify(ok=False, error="Empty filename"), 400
    
    # Read and store
    data = f.read()
    
    # Detect content type
    content_type = detect_mime_from_bytes(data, f.filename) or "application/octet-stream"
    
    # Store file
    file_info = store_file(
        user["org_id"],
        secure_filename(f.filename),
        data,
        content_type,
        user["id"]
    )
    
    return jsonify(ok=True, file=file_info)


@app.route("/api/task/comment/upload", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("task_comment_upload", per_min=30)
def api_task_comment_upload():
    """Upload attachment for task comment"""
    user = g.user
    
    if "file" not in request.files:
        return jsonify(ok=False, error="No file"), 400
    
    f = request.files["file"]
    if not f.filename:
        return jsonify(ok=False, error="Empty filename"), 400
    
    data = f.read()
    content_type = detect_mime_from_bytes(data, f.filename) or "application/octet-stream"
    
    file_info = store_file(
        user["org_id"],
        secure_filename(f.filename),
        data,
        content_type,
        user["id"]
    )
    
    return jsonify(ok=True, file=file_info)


@app.route("/api/chat/upload", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("chat_upload", per_min=30)
def api_chat_upload():
    """Upload file to chat channel"""
    user = g.user
    
    channel_id = request.form.get("channel_id")
    
    if not channel_id or "file" not in request.files:
        return jsonify(ok=False, error="Channel ID and file required"), 400
    
    # Verify channel access
    channel = query_db(
        """
        SELECT c.* FROM chat_channels c
        LEFT JOIN chat_members m ON c.id = m.channel_id
        WHERE c.id=? AND c.org_id=? AND (c.type='public' OR m.user_id=?)
        LIMIT 1
        """,
        (channel_id, user["org_id"], user["id"]),
        one=True
    )
    
    if not channel:
        return jsonify(ok=False, error="Channel not found"), 403
    
    f = request.files["file"]
    data = f.read()
    content_type = detect_mime_from_bytes(data, f.filename) or "application/octet-stream"
    
    file_info = store_file(
        user["org_id"],
        secure_filename(f.filename),
        data,
        content_type,
        user["id"]
    )
    
    # Send as chat message (file link)
    exec_db(
        "INSERT INTO chat_messages (channel_id, user_id, body, created_at) VALUES (?, ?, ?, ?)",
        (channel_id, user["id"], f"📎 {file_info['name']} ({file_info['url']})", utc_now())
    )
    
    return jsonify(ok=True, file_id=file_info["id"], url=file_info["url"])


# ===== BLOCK: TASK REMINDERS API =====

@app.route("/api/task/reminder", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_reminder_add():
    """Add reminder to task"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    remind_at = data.get("remind_at")
    message = data.get("message", "")
    
    if not task_id or not remind_at:
        return jsonify(ok=False, error="Task ID and remind_at required"), 400
    
    # Verify task
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    reminder_id = add_task_reminder(task_id, user["id"], remind_at, message)
    
    return jsonify(ok=True, id=reminder_id)


@app.route("/api/task/reminders/<int:task_id>", methods=["GET"])
@_login_required
def api_task_reminders_list(task_id):
    """Get task reminders"""
    user = g.user
    
    # Verify task
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    reminders = query_db(
        "SELECT * FROM task_reminders WHERE task_id=? ORDER BY remind_at",
        (task_id,)
    )
    
    return jsonify(ok=True, items=[dict(r) for r in reminders])


@app.route("/api/task/file_pin", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_file_pin():
    """Pin or unpin file from task"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    file_id = data.get("file_id")
    pin = data.get("pin", True)
    
    if not task_id or not file_id:
        return jsonify(ok=False, error="Task ID and file ID required"), 400
    
    # Verify task
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    # Verify file
    file_obj = query_db("SELECT id FROM files WHERE id=? AND org_id=?", (file_id, user["org_id"]), one=True)
    if not file_obj:
        return jsonify(ok=False, error="File not found"), 404
    
    pin_file_to_task(task_id, file_id, pin)
    
    return jsonify(ok=True)


# ===== BLOCK: WORKFLOW API =====

@app.route("/api/task/workflow/transition", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_workflow_transition():
    """
    Apply workflow transition to task (stage change + department assignment).
    ✅ НОВОЕ: для UI перевода задачи по воронке
    """
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    stage_key = data.get("stage_key")
    department_id = data.get("department_id")
    due_at = data.get("due_at")
    comment = data.get("comment", "")
    
    if not task_id:
        return jsonify(ok=False, error="Task ID required"), 400
    
    # Verify task
    task = query_db("SELECT * FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    updates = {}
    
    # Get current stage from task metadata (if stored)
    current_stage = None  # Tasks don't have stage by default, but can extend schema
    
    # Apply stage transition
    if stage_key:
        # Log transition
        exec_db(
            "INSERT INTO stage_transitions (entity_type, entity_id, from_stage, to_stage, user_id, comment, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("task", task_id, current_stage, stage_key, user["id"], comment, utc_now())
        )
        
        # Get stage SLA
        stage = query_db(
            "SELECT sla_hours FROM workflow_stages WHERE org_id=? AND entity_type='task' AND key=?",
            (user["org_id"], stage_key),
            one=True
        )
        
        if stage and stage["sla_hours"]:
            # Calculate new due_at based on SLA
            new_due = (datetime.utcnow() + timedelta(hours=stage["sla_hours"])).isoformat(" ", "seconds")
            updates["due_at"] = new_due
    
    # Apply department
    if department_id:
        # Find default assignee in department (or keep current)
        dept_users = query_db(
            "SELECT id FROM users WHERE org_id=? AND department_id=? AND active=1 LIMIT 1",
            (user["org_id"], department_id),
            one=True
        )
        if dept_users:
            updates["assignee_id"] = dept_users["id"]
    
    # Apply custom due_at
    if due_at:
        updates["due_at"] = due_at
    
    # Update task
    if updates:
        update_task(task_id, user["org_id"], updates)
    
    # Add comment if provided
    if comment:
        add_task_comment(task_id, user["id"], f"Transition: {comment}")
    
    return jsonify(ok=True)


@app.route("/api/task/delegate", methods=["POST"])
@_login_required
@_csrf_protect
def api_task_delegate():
    """Delegate task to another user"""
    user = g.user
    data = request.get_json() or {}
    
    task_id = data.get("task_id")
    user_id = data.get("user_id")
    
    if not task_id or not user_id:
        return jsonify(ok=False, error="Task ID and user ID required"), 400
    
    # Verify task
    task = query_db("SELECT id FROM tasks WHERE id=? AND org_id=?", (task_id, user["org_id"]), one=True)
    if not task:
        return jsonify(ok=False, error="Task not found"), 404
    
    # Verify target user
    target_user = query_db("SELECT id FROM users WHERE id=? AND org_id=? AND active=1", (user_id, user["org_id"]), one=True)
    if not target_user:
        return jsonify(ok=False, error="User not found"), 404
    
    # Update task
    update_task(task_id, user["org_id"], {"assignee_id": user_id})
    
    # Add activity log
    exec_db(
        "INSERT INTO task_activity (task_id, user_id, kind, details_json, created_at) VALUES (?, ?, ?, ?, ?)",
        (task_id, user["id"], "delegated", json.dumps({"to_user_id": user_id}), utc_now())
    )
    
    return jsonify(ok=True)


# ===== BLOCK: MEETING RECORDING API =====

@app.route("/api/meeting/<int:meeting_id>/start_recording", methods=["POST"])
@_login_required
@_csrf_protect
def api_meeting_start_recording(meeting_id):
    """Start meeting recording (Jitsi API integration)"""
    user = g.user
    
    meeting = query_db(
        "SELECT * FROM meetings WHERE id=? AND org_id=?",
        (meeting_id, user["org_id"]),
        one=True
    )
    
    if not meeting:
        return jsonify(ok=False, error="Meeting not found"), 404
    
    # Update recording status
    exec_db(
        "UPDATE meetings SET recording_started_at=? WHERE id=?",
        (utc_now(), meeting_id)
    )
    
    # In production, call Jitsi API to start recording
    # For now, just log the action
    log("INFO", f"Meeting {meeting_id} recording started")
    
    return jsonify(ok=True)


@app.route("/api/meeting/<int:meeting_id>/stop_recording", methods=["POST"])
@_login_required
@_csrf_protect
def api_meeting_stop_recording(meeting_id):
    """Stop meeting recording"""
    user = g.user
    
    meeting = query_db(
        "SELECT * FROM meetings WHERE id=? AND org_id=?",
        (meeting_id, user["org_id"]),
        one=True
    )
    
    if not meeting:
        return jsonify(ok=False, error="Meeting not found"), 404
    
    exec_db(
        "UPDATE meetings SET recording_stopped_at=? WHERE id=?",
        (utc_now(), meeting_id)
    )
    
    log("INFO", f"Meeting {meeting_id} recording stopped")
    
    # Return summary
    duration_sec = 0
    if meeting["recording_started_at"]:
        start = datetime.fromisoformat(meeting["recording_started_at"])
        duration_sec = int((datetime.utcnow() - start).total_seconds())
    
    return jsonify(ok=True, summary=f"Recording duration: {duration_sec} seconds")


# ===== BLOCK: CHANNEL API =====

@app.route("/api/channel/send", methods=["POST"])
@_login_required
@_csrf_protect
@_rate_limit("channel_send", per_min=60)
def api_channel_send():
    """
    Send message via external channel (Telegram, VK, etc).
    Placeholder for channel integrations.
    """
    user = g.user
    data = request.get_json() or {}
    
    thread_id = data.get("thread_id")
    text = data.get("text", "").strip()
    
    if not thread_id or not text:
        return jsonify(ok=False, error="Thread ID and text required"), 400
    
    # Get thread
    thread = query_db(
        "SELECT * FROM inbox_threads WHERE id=? AND org_id=?",
        (thread_id, user["org_id"]),
        one=True
    )
    
    if not thread:
        return jsonify(ok=False, error="Thread not found"), 404
    
    # Get channel
    channel = query_db("SELECT * FROM channels WHERE id=?", (thread["channel_id"],), one=True)
    if not channel:
        return jsonify(ok=False, error="Channel not found"), 404
    
    # Send via channel (placeholder - implement provider-specific logic)
    # For now, just add as internal message
    add_message(thread_id, "agent", text, user["id"])
    
    log("INFO", f"Message sent via channel {channel['type']}: thread {thread_id}")
    
    return jsonify(ok=True)


# ===== END OF CORE PART 9A/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 9B/10 — REPORTS, EXPORT, SCORING, SEQUENCES, UTILITIES
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: ADVANCED REPORTS API =====

@app.route("/api/reports/sales_funnel", methods=["GET"])
@_login_required
def api_reports_sales_funnel():
    """
    Sales funnel report (conversion by stages).
    ✅ НОВОЕ: для dashboard analytics
    """
    user = g.user
    pipeline = request.args.get("pipeline", "default")
    
    # Get stages for pipeline
    stages = query_db(
        """
        SELECT key, name FROM workflow_stages
        WHERE org_id=? AND entity_type='deal' AND pipeline_key=?
        ORDER BY sort_order
        """,
        (user["org_id"], pipeline)
    )
    
    if not stages:
        return jsonify(ok=True, stages=[], data=[])
    
    # Count deals at each stage
    funnel_data = []
    for stage in stages:
        count = query_db(
            "SELECT COUNT(*) as cnt FROM deals WHERE org_id=? AND pipeline_key=? AND stage=? AND status='open'",
            (user["org_id"], pipeline, stage["key"]),
            one=True
        )["cnt"]
        
        funnel_data.append({
            "stage": stage["key"],
            "name": stage["name"],
            "count": count
        })
    
    return jsonify(ok=True, stages=[s["key"] for s in stages], data=funnel_data)


@app.route("/api/reports/team_performance", methods=["GET"])
@_login_required
def api_reports_team_performance():
    """
    Team performance report.
    Metrics: tasks completed, deals won, avg response time, etc.
    """
    user = g.user
    
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    
    # Build WHERE
    where_parts = ["org_id = ?"]
    params = [user["org_id"]]
    
    if date_from:
        where_parts.append("date(created_at) >= date(?)")
        params.append(date_from)
    
    if date_to:
        where_parts.append("date(created_at) <= date(?)")
        params.append(date_to)
    
    where_clause = " AND ".join(where_parts)
    
    # Get agents
    agents = query_db("SELECT id, username FROM users WHERE org_id=? AND active=1", (user["org_id"],))
    
    result = []
    for agent in agents:
        # Tasks completed
        tasks_done = query_db(
            f"SELECT COUNT(*) as cnt FROM tasks WHERE assignee_id=? AND status='done' AND {where_clause}",
            (agent["id"], *params),
            one=True
        )["cnt"]
        
        # Deals won
        deals_won = query_db(
            f"SELECT COUNT(*) as cnt, SUM(amount) as total FROM deals WHERE assignee_id=? AND status='won' AND {where_clause}",
            (agent["id"], *params),
            one=True
        )
        
        # Avg response time (from inbox FRT)
        avg_response = query_db(
            f"""
            SELECT AVG(
                CAST((julianday(first_response_at) - julianday(created_at)) * 24 * 60 AS INTEGER)
            ) as avg_min
            FROM inbox_threads
            WHERE assignee_id=? AND first_response_at IS NOT NULL AND {where_clause}
            """,
            (agent["id"], *params),
            one=True
        )
        
        result.append({
            "agent_id": agent["id"],
            "username": agent["username"],
            "tasks_completed": tasks_done,
            "deals_won": deals_won["cnt"] or 0,
            "revenue": float(deals_won["total"] or 0),
            "avg_response_time_min": int(avg_response["avg_min"] or 0)
        })
    
    return jsonify(ok=True, items=result)


@app.route("/api/reports/revenue_forecast", methods=["GET"])
@_login_required
def api_reports_revenue_forecast():
    """
    Revenue forecast based on open deals and win probability.
    ✅ НОВОЕ: предиктивная аналитика
    """
    user = g.user
    
    # Get open deals grouped by stage
    deals = query_db(
        """
        SELECT stage, COUNT(*) as cnt, SUM(amount) as total
        FROM deals
        WHERE org_id=? AND status='open'
        GROUP BY stage
        """,
        (user["org_id"],)
    )
    
    # Define win probability by stage (can be configured in workflow_stages)
    stage_probability = {
        "new": 0.1,
        "qualify": 0.25,
        "proposal": 0.5,
        "negotiation": 0.75,
        "closed_won": 1.0
    }
    
    forecast_data = []
    total_weighted = 0
    
    for deal_group in deals:
        stage = deal_group["stage"]
        count = deal_group["cnt"]
        total = float(deal_group["total"] or 0)
        
        probability = stage_probability.get(stage, 0.3)  # Default 30%
        weighted_revenue = total * probability
        total_weighted += weighted_revenue
        
        forecast_data.append({
            "stage": stage,
            "deals_count": count,
            "total_value": total,
            "win_probability": probability,
            "weighted_revenue": weighted_revenue
        })
    
    return jsonify(ok=True, forecast=forecast_data, total_forecast=total_weighted)


@app.route("/api/reports/customer_lifetime_value", methods=["GET"])
@_login_required
def api_reports_clv():
    """
    Customer Lifetime Value report.
    Shows total revenue per company.
    """
    user = g.user
    limit = min(int(request.args.get("limit", 50)), 200)
    
    companies = query_db(
        """
        SELECT 
            c.id,
            c.name,
            COALESCE(SUM(d.amount), 0) as total_revenue,
            COUNT(d.id) as deals_count
        FROM companies c
        LEFT JOIN deals d ON c.id = d.company_id AND d.status='won'
        WHERE c.org_id=?
        GROUP BY c.id
        ORDER BY total_revenue DESC
        LIMIT ?
        """,
        (user["org_id"], limit)
    )
    
    return jsonify(ok=True, items=[dict(c) for c in companies])


# ===== BLOCK: EXPORT FUNCTIONALITY =====

@app.route("/api/export/inbox/csv", methods=["GET"])
@_login_required
def export_inbox_csv():
    """
    Export inbox threads to CSV.
    ✅ НОВОЕ: для data export
    """
    user = g.user
    
    # Get threads (apply same filters as inbox list)
    threads = query_db(
        """
        SELECT 
            t.id,
            t.subject,
            t.status,
            t.priority,
            c.name AS channel_name,
            u.username AS assignee,
            t.created_at,
            t.last_message_at
        FROM inbox_threads t
        LEFT JOIN channels c ON t.channel_id = c.id
        LEFT JOIN users u ON t.assignee_id = u.id
        WHERE t.org_id=?
        ORDER BY t.last_message_at DESC
        LIMIT 5000
        """,
        (user["org_id"],)
    )
    
    # Generate CSV
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Header
    writer.writerow(['ID', 'Subject', 'Status', 'Priority', 'Channel', 'Assignee', 'Created', 'Last Message'])
    
    # Data
    for t in threads:
        writer.writerow([
            t["id"],
            t["subject"] or "",
            t["status"] or "",
            t["priority"] or "",
            t["channel_name"] or "",
            t["assignee"] or "",
            t["created_at"] or "",
            t["last_message_at"] or ""
        ])
    
    # Return as file
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=inbox_export.csv"
        }
    )


@app.route("/api/export/tasks/csv", methods=["GET"])
@_login_required
def export_tasks_csv():
    """Export tasks to CSV"""
    user = g.user
    
    tasks = query_db(
        """
        SELECT 
            t.id,
            t.title,
            t.status,
            t.priority,
            u.username AS assignee,
            c.name AS company,
            t.due_at,
            t.created_at,
            t.completed_at
        FROM tasks t
        LEFT JOIN users u ON t.assignee_id = u.id
        LEFT JOIN companies c ON t.company_id = c.id
        WHERE t.org_id=?
        ORDER BY t.created_at DESC
        LIMIT 5000
        """,
        (user["org_id"],)
    )
    
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    
    writer.writerow(['ID', 'Title', 'Status', 'Priority', 'Assignee', 'Company', 'Due Date', 'Created', 'Completed'])
    
    for t in tasks:
        writer.writerow([
            t["id"],
            t["title"] or "",
            t["status"] or "",
            t["priority"] or "",
            t["assignee"] or "",
            t["company"] or "",
            t["due_at"] or "",
            t["created_at"] or "",
            t["completed_at"] or ""
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=tasks_export.csv"}
    )


@app.route("/api/export/deals/excel", methods=["GET"])
@_login_required
def export_deals_excel():
    """
    Export deals to Excel.
    ✅ НОВОЕ: Excel export для deals
    """
    user = g.user
    
    deals = query_db(
        """
        SELECT 
            d.id,
            d.title,
            d.amount,
            d.currency,
            d.status,
            d.stage,
            u.username AS assignee,
            c.name AS company,
            d.created_at
        FROM deals d
        LEFT JOIN users u ON d.assignee_id = u.id
        LEFT JOIN companies c ON d.company_id = c.id
        WHERE d.org_id=?
        ORDER BY d.created_at DESC
        LIMIT 5000
        """,
        (user["org_id"],)
    )
    
    try:
        # Try to use openpyxl if available
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Deals"
        
        # Header
        headers = ['ID', 'Title', 'Amount', 'Currency', 'Status', 'Stage', 'Assignee', 'Company', 'Created']
        ws.append(headers)
        
        # Data
        for d in deals:
            ws.append([
                d["id"],
                d["title"] or "",
                float(d["amount"] or 0),
                d["currency"] or "RUB",
                d["status"] or "",
                d["stage"] or "",
                d["assignee"] or "",
                d["company"] or "",
                d["created_at"] or ""
            ])
        
        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="deals_export.xlsx"
        )
    
    except ImportError:
        # Fallback to CSV if openpyxl not available
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        
        writer.writerow(['ID', 'Title', 'Amount', 'Currency', 'Status', 'Stage', 'Assignee', 'Company', 'Created'])
        
        for d in deals:
            writer.writerow([
                d["id"],
                d["title"] or "",
                d["amount"] or 0,
                d["currency"] or "RUB",
                d["status"] or "",
                d["stage"] or "",
                d["assignee"] or "",
                d["company"] or "",
                d["created_at"] or ""
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=deals_export.csv"}
        )


# ===== BLOCK: LEAD SCORING API =====

@app.route("/api/scoring/rules/list", methods=["GET"])
@_login_required
@_require_role("admin")
def api_scoring_rules_list():
    """
    List lead scoring rules.
    ✅ НОВОЕ: управление правилами scoring
    """
    user = g.user
    
    rules = query_db(
        "SELECT * FROM lead_scoring_rules WHERE org_id=? ORDER BY id",
        (user["org_id"],)
    )
    
    return jsonify(ok=True, items=[dict(r) for r in rules])


@app.route("/api/scoring/rule/create", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_scoring_rule_create():
    """Create lead scoring rule"""
    user = g.user
    data = request.get_json() or {}
    
    name = data.get("name", "").strip()
    entity_type = data.get("entity_type", "deal")
    field = data.get("field", "")
    operator = data.get("operator", "==")
    value = data.get("value", "")
    score_delta = int(data.get("score_delta", 10))
    
    if not name or not field:
        return jsonify(ok=False, error="Name and field required"), 400
    
    rule_id = exec_db(
        """
        INSERT INTO lead_scoring_rules (org_id, name, entity_type, field, operator, value, score_delta, active, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (user["org_id"], name, entity_type, field, operator, value, score_delta, 1, utc_now())
    )
    
    return jsonify(ok=True, id=rule_id)


@app.route("/api/scoring/rule/<int:rule_id>/toggle", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_scoring_rule_toggle(rule_id):
    """Toggle scoring rule active status"""
    user = g.user
    
    rule = query_db("SELECT active FROM lead_scoring_rules WHERE id=? AND org_id=?", (rule_id, user["org_id"]), one=True)
    if not rule:
        return jsonify(ok=False, error="Rule not found"), 404
    
    new_active = 0 if rule["active"] else 1
    exec_db("UPDATE lead_scoring_rules SET active=? WHERE id=?", (new_active, rule_id))
    
    return jsonify(ok=True, active=bool(new_active))


@app.route("/api/scoring/recalculate", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_scoring_recalculate():
    """
    Recalculate scores for all deals/companies.
    Background job - returns immediately.
    """
    user = g.user
    data = request.get_json() or {}
    
    entity_type = data.get("entity_type", "deal")
    
    if entity_type not in ("deal", "company"):
        return jsonify(ok=False, error="Invalid entity type"), 400
    
    # Get all entities
    if entity_type == "deal":
        entities = query_db("SELECT id FROM deals WHERE org_id=?", (user["org_id"],))
    else:
        entities = query_db("SELECT id FROM companies WHERE org_id=?", (user["org_id"],))
    
    # Recalculate in background (for production, use Celery)
    count = 0
    for entity in entities:
        try:
            recalculate_score(entity_type, entity["id"], user["org_id"])
            count += 1
        except Exception as e:
            log("ERROR", f"Score recalculation failed for {entity_type} {entity['id']}: {e}")
    
    return jsonify(ok=True, recalculated=count)


# ===== BLOCK: EMAIL SEQUENCES API =====

@app.route("/api/sequences/list", methods=["GET"])
@_login_required
def api_sequences_list():
    """List email sequences"""
    user = g.user
    
    sequences = query_db(
        "SELECT * FROM email_sequences WHERE org_id=? ORDER BY created_at DESC",
        (user["org_id"],)
    )
    
    return jsonify(ok=True, items=[dict(s) for s in sequences])


@app.route("/api/sequence/create", methods=["POST"])
@_login_required
@_csrf_protect
def api_sequence_create():
    """
    Create email sequence.
    ✅ НОВОЕ: для drip campaigns
    """
    user = g.user
    data = request.get_json() or {}
    
    name = data.get("name", "").strip()
    if not name:
        return jsonify(ok=False, error="Name required"), 400
    
    sequence_id = exec_db(
        "INSERT INTO email_sequences (org_id, name, active, created_at) VALUES (?, ?, ?, ?)",
        (user["org_id"], name, 1, utc_now())
    )
    
    return jsonify(ok=True, id=sequence_id)


@app.route("/api/sequence/<int:seq_id>/steps", methods=["GET"])
@_login_required
def api_sequence_steps_list(seq_id):
    """Get sequence steps"""
    user = g.user
    
    # Verify sequence
    seq = query_db("SELECT id FROM email_sequences WHERE id=? AND org_id=?", (seq_id, user["org_id"]), one=True)
    if not seq:
        return jsonify(ok=False, error="Sequence not found"), 404
    
    steps = query_db(
        "SELECT * FROM sequence_steps WHERE sequence_id=? ORDER BY step_num",
        (seq_id,)
    )
    
    return jsonify(ok=True, items=[dict(s) for s in steps])


@app.route("/api/sequence/<int:seq_id>/step/add", methods=["POST"])
@_login_required
@_csrf_protect
def api_sequence_step_add(seq_id):
    """Add step to sequence"""
    user = g.user
    data = request.get_json() or {}
    
    # Verify sequence
    seq = query_db("SELECT id FROM email_sequences WHERE id=? AND org_id=?", (seq_id, user["org_id"]), one=True)
    if not seq:
        return jsonify(ok=False, error="Sequence not found"), 404
    
    step_num = int(data.get("step_num", 1))
    delay_hours = int(data.get("delay_hours", 24))
    subject = data.get("subject", "")
    body_template = data.get("body_template", "")
    
    step_id = exec_db(
        """
        INSERT INTO sequence_steps (sequence_id, step_num, delay_hours, subject, body_template, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (seq_id, step_num, delay_hours, subject, body_template, utc_now())
    )
    
    return jsonify(ok=True, id=step_id)


@app.route("/api/sequence/enroll", methods=["POST"])
@_login_required
@_csrf_protect
def api_sequence_enroll():
    """Enroll contact/company in sequence"""
    user = g.user
    data = request.get_json() or {}
    
    sequence_id = data.get("sequence_id")
    email = data.get("email", "").strip()
    contact_id = data.get("contact_id")
    company_id = data.get("company_id")
    
    if not sequence_id or not email:
        return jsonify(ok=False, error="Sequence ID and email required"), 400
    
    # Verify sequence
    seq = query_db("SELECT id FROM email_sequences WHERE id=? AND org_id=?", (sequence_id, user["org_id"]), one=True)
    if not seq:
        return jsonify(ok=False, error="Sequence not found"), 404
    
    # Check if already enrolled
    existing = query_db(
        "SELECT id FROM sequence_enrollments WHERE sequence_id=? AND email=? AND status='active'",
        (sequence_id, email),
        one=True
    )
    
    if existing:
        return jsonify(ok=False, error="Already enrolled"), 400
    
    # Enroll
    enrollment_id = exec_db(
        """
        INSERT INTO sequence_enrollments (sequence_id, contact_id, company_id, email, status, enrolled_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (sequence_id, contact_id, company_id, email, "active", utc_now())
    )
    
    return jsonify(ok=True, id=enrollment_id)


@app.route("/api/sequence/enrollments/<int:seq_id>", methods=["GET"])
@_login_required
def api_sequence_enrollments_list(seq_id):
    """Get sequence enrollments"""
    user = g.user
    
    # Verify sequence
    seq = query_db("SELECT id FROM email_sequences WHERE id=? AND org_id=?", (seq_id, user["org_id"]), one=True)
    if not seq:
        return jsonify(ok=False, error="Sequence not found"), 404
    
    enrollments = query_db(
        "SELECT * FROM sequence_enrollments WHERE sequence_id=? ORDER BY enrolled_at DESC LIMIT 500",
        (seq_id,)
    )
    
    return jsonify(ok=True, items=[dict(e) for e in enrollments])


# ===== BLOCK: AI AGENT ACTIONS LOGGING =====

@app.route("/api/agent/action/log", methods=["POST"])
@_login_required
@_csrf_protect
def api_agent_action_log():
    """
    Log AI agent action.
    ✅ НОВОЕ: для audit trail AI-агентов
    
    Body: {
        "agent_name": "CustomerSupportBot",
        "action_type": "task.created",
        "entity_type": "task",
        "entity_id": 123,
        "reasoning": "User requested follow-up",
        "success": true
    }
    """
    user = g.user
    data = request.get_json() or {}
    
    agent_name = data.get("agent_name", "unknown")
    action_type = data.get("action_type", "")
    entity_type = data.get("entity_type")
    entity_id = data.get("entity_id")
    reasoning = data.get("reasoning", "")
    success = data.get("success", True)
    error = data.get("error")
    
    action_id = exec_db(
        """
        INSERT INTO agent_actions (org_id, agent_name, action_type, entity_type, entity_id, reasoning, success, error, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (user["org_id"], agent_name, action_type, entity_type, entity_id, reasoning, int(success), error, utc_now())
    )
    
    return jsonify(ok=True, id=action_id)


@app.route("/api/agent/actions/list", methods=["GET"])
@_login_required
def api_agent_actions_list():
    """Get recent agent actions"""
    user = g.user
    
    limit = min(int(request.args.get("limit", 100)), 500)
    agent_name = request.args.get("agent_name", "")
    
    where_parts = ["org_id = ?"]
    params = [user["org_id"]]
    
    if agent_name:
        where_parts.append("agent_name = ?")
        params.append(agent_name)
    
    where_clause = " AND ".join(where_parts)
    
    actions = query_db(
        f"""
        SELECT * FROM agent_actions
        WHERE {where_clause}
        ORDER BY created_at DESC
        LIMIT ?
        """,
        (*params, limit)
    )
    
    return jsonify(ok=True, items=[dict(a) for a in actions])


# ===== BLOCK: AI FEEDBACK API =====

@app.route("/api/ai/feedback", methods=["POST"])
@_login_required
@_csrf_protect
def api_ai_feedback():
    """
    Submit feedback on AI job output.
    ✅ НОВОЕ: для обучения AI
    
    Body: {
        "ai_job_id": 123,
        "rating": 4,
        "correction": "Better answer would be..."
    }
    """
    user = g.user
    data = request.get_json() or {}
    
    ai_job_id = data.get("ai_job_id")
    rating = data.get("rating")  # 1-5
    correction = data.get("correction", "")
    
    if not ai_job_id:
        return jsonify(ok=False, error="AI job ID required"), 400
    
    # Verify job belongs to org
    job = query_db("SELECT id FROM ai_jobs WHERE id=? AND org_id=?", (ai_job_id, user["org_id"]), one=True)
    if not job:
        return jsonify(ok=False, error="AI job not found"), 404
    
    feedback_id = exec_db(
        "INSERT INTO ai_feedback (ai_job_id, user_id, rating, correction, created_at) VALUES (?, ?, ?, ?, ?)",
        (ai_job_id, user["id"], rating, correction, utc_now())
    )
    
    return jsonify(ok=True, id=feedback_id)


@app.route("/api/ai/feedback/stats", methods=["GET"])
@_login_required
@_require_role("admin")
def api_ai_feedback_stats():
    """Get AI feedback statistics"""
    user = g.user
    
    stats = query_db(
        """
        SELECT 
            AVG(rating) as avg_rating,
            COUNT(*) as total_feedback,
            SUM(CASE WHEN rating >= 4 THEN 1 ELSE 0 END) as positive_count
        FROM ai_feedback f
        JOIN ai_jobs j ON f.ai_job_id = j.id
        WHERE j.org_id = ?
        """,
        (user["org_id"],),
        one=True
    )
    
    return jsonify(
        ok=True,
        avg_rating=float(stats["avg_rating"] or 0),
        total_feedback=stats["total_feedback"],
        positive_count=stats["positive_count"],
        satisfaction_rate=stats["positive_count"] / max(stats["total_feedback"], 1) if stats["total_feedback"] else 0
    )


# ===== BLOCK: EMBEDDINGS API (VECTOR STORAGE) =====

@app.route("/api/embeddings/store", methods=["POST"])
@_login_required
@_csrf_protect
def api_embeddings_store():
    """
    Store embedding vector for entity.
    ✅ НОВОЕ: для semantic search
    
    Body: {
        "entity_type": "task",
        "entity_id": 123,
        "model": "text-embedding-ada-002",
        "vector": [0.123, 0.456, ...]  # Base64-encoded or array
    }
    """
    user = g.user
    data = request.get_json() or {}
    
    entity_type = data.get("entity_type")
    entity_id = data.get("entity_id")
    model = data.get("model", "unknown")
    vector = data.get("vector")
    
    if not entity_type or not entity_id or not vector:
        return jsonify(ok=False, error="Missing required fields"), 400
    
    # Convert vector to bytes (store as BLOB)
    if isinstance(vector, list):
        import struct
        vector_bytes = struct.pack(f"{len(vector)}f", *vector)
    elif isinstance(vector, str):
        # Assume base64
        vector_bytes = base64.b64decode(vector)
    else:
        return jsonify(ok=False, error="Invalid vector format"), 400
    
    # Store (UPSERT)
    existing = query_db(
        "SELECT id FROM embeddings WHERE org_id=? AND entity_type=? AND entity_id=? AND model=?",
        (user["org_id"], entity_type, entity_id, model),
        one=True
    )
    
    if existing:
        exec_db(
            "UPDATE embeddings SET vector=?, created_at=? WHERE id=?",
            (vector_bytes, utc_now(), existing["id"])
        )
        embedding_id = existing["id"]
    else:
        embedding_id = exec_db(
            "INSERT INTO embeddings (org_id, entity_type, entity_id, model, vector, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (user["org_id"], entity_type, entity_id, model, vector_bytes, utc_now())
        )
    
    return jsonify(ok=True, id=embedding_id)


@app.route("/api/embeddings/search", methods=["POST"])
@_login_required
@_csrf_protect
def api_embeddings_search():
    """
    Semantic search using embeddings (cosine similarity).
    Placeholder - requires vector similarity library (faiss, pgvector, etc.)
    """
    user = g.user
    data = request.get_json() or {}
    
    query_vector = data.get("vector")
    entity_type = data.get("entity_type")
    limit = min(int(data.get("limit", 10)), 50)
    
    if not query_vector:
        return jsonify(ok=False, error="Query vector required"), 400
    
    # For now, return placeholder
    # In production, implement cosine similarity calculation
    # using numpy or integrate with pgvector/faiss
    
    return jsonify(
        ok=True,
        message="Semantic search not fully implemented - requires vector similarity calculation",
        results=[]
    )


# ===== BLOCK: MISC UTILITY ENDPOINTS =====

@app.route("/api/utils/phones/normalize", methods=["POST"])
@_login_required
def api_utils_phones_normalize():
    """
    Normalize phone numbers batch.
    ✅ НОВОЕ: utility для bulk phone normalization
    """
    data = request.get_json() or {}
    phones = data.get("phones", [])
    
    if not phones or len(phones) > 100:
        return jsonify(ok=False, error="Provide 1-100 phone numbers"), 400
    
    results = {}
    for phone in phones:
        results[phone] = normalize_phone(phone)
    
    return jsonify(ok=True, results=results)


@app.route("/api/utils/html/sanitize", methods=["POST"])
@_login_required
def api_utils_html_sanitize():
    """Sanitize HTML (utility for preview)"""
    data = request.get_json() or {}
    html = data.get("html", "")
    
    if not html:
        return jsonify(ok=False, error="HTML required"), 400
    
    sanitized = sanitize_html(html)
    
    return jsonify(ok=True, sanitized=sanitized)


@app.route("/api/utils/fts/rebuild", methods=["POST"])
@_login_required
@_csrf_protect
@_require_role("admin")
def api_utils_fts_rebuild():
    """
    Rebuild FTS indexes.
    Admin-only background operation.
    """
    user = g.user
    data = request.get_json() or {}
    
    table_name = data.get("table_name")
    
    if table_name not in ("fts_inbox_messages", "fts_tasks", "fts_chat_messages"):
        return jsonify(ok=False, error="Invalid table name"), 400
    
    try:
        rebuild_fts_table(table_name)
        
        add_audit(user["org_id"], user["id"], "fts.rebuild", details={"table": table_name})
        
        return jsonify(ok=True, message=f"FTS table {table_name} rebuilt")
    except Exception as e:
        log("ERROR", f"FTS rebuild failed: {e}")
        return jsonify(ok=False, error="Rebuild failed"), 500


@app.route("/api/utils/db/stats", methods=["GET"])
@_login_required
@_require_role("admin")
def api_utils_db_stats():
    """Get database statistics"""
    user = g.user
    
    stats = {}
    
    tables = [
        "users", "companies", "contacts", "tasks", "deals",
        "inbox_threads", "inbox_messages", "calls", "chat_messages"
    ]
    
    for table in tables:
        try:
            count = query_db(f"SELECT COUNT(*) as cnt FROM {table} WHERE org_id=?", (user["org_id"],), one=True)
            stats[table] = count["cnt"]
        except Exception:
            stats[table] = 0
    
    return jsonify(ok=True, stats=stats)


# ===== BLOCK: SYSTEM INFO =====

@app.route("/api/system/info", methods=["GET"])
@_login_required
@_require_role("admin")
def api_system_info():
    """Get system information (admin only)"""
    user = g.user
    
    info = {
        "version": SCHEMA_VERSION,
        "debug": DEBUG,
        "env": ENV,
        "storage_backend": STORAGE_BACKEND,
        "ai_provider": AI_PROVIDER,
        "redis_available": REDIS_AVAILABLE,
        "sse_enabled": SSE_ENABLED,
        "rate_limit_enabled": RATE_LIMIT_ENABLED,
    }
    
    return jsonify(ok=True, **info)


# ===== END OF CORE PART 9B/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# CORE PART 10/10 — FINAL: STARTUP, TEMPLATES DECLARATIONS, CLEANUP
# ═════════════════════════════════════════════════════════════════════════════

# ===== BLOCK: REMAINING UTILITY ROUTES =====

@app.route("/api/notifications/unread", methods=["GET"])
@_login_required
def api_notifications_unread():
    """
    Get unread notifications count.
    Placeholder - notifications system can be expanded.
    """
    user = g.user
    
    # Count unread inbox threads assigned to user
    unread_threads = query_db(
        "SELECT COUNT(*) as cnt FROM inbox_threads WHERE org_id=? AND assignee_id=? AND status='open'",
        (user["org_id"], user["id"]),
        one=True
    )["cnt"]
    
    # Count overdue tasks
    overdue_tasks = query_db(
        "SELECT COUNT(*) as cnt FROM tasks WHERE org_id=? AND assignee_id=? AND due_at < datetime('now') AND status NOT IN ('done', 'cancelled')",
        (user["org_id"], user["id"]),
        one=True
    )["cnt"]
    
    return jsonify(
        ok=True,
        unread_threads=unread_threads,
        overdue_tasks=overdue_tasks,
        total=unread_threads + overdue_tasks
    )


@app.route("/api/recent/activity", methods=["GET"])
@_login_required
def api_recent_activity():
    """
    Get recent activity across all entities.
    ✅ НОВОЕ: unified activity feed
    """
    user = g.user
    limit = min(int(request.args.get("limit", 50)), 100)
    
    activities = []
    
    # Recent audit logs
    audits = query_db(
        """
        SELECT 
            'audit' as type,
            action,
            entity_type,
            entity_id,
            created_at,
            u.username
        FROM audit_logs a
        LEFT JOIN users u ON a.user_id = u.id
        WHERE a.org_id=?
        ORDER BY a.created_at DESC
        LIMIT ?
        """,
        (user["org_id"], limit)
    )
    
    activities.extend([dict(a) for a in audits])
    
    # Sort by created_at
    activities.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    
    return jsonify(ok=True, items=activities[:limit])


@app.route("/api/dashboard/widgets", methods=["GET"])
@_login_required
def api_dashboard_widgets():
    """
    Get dashboard widget data.
    ✅ НОВОЕ: для dashboard cards
    """
    user = g.user
    
    # Tasks summary
    tasks_open = query_db(
        "SELECT COUNT(*) as cnt FROM tasks WHERE org_id=? AND status NOT IN ('done', 'cancelled')",
        (user["org_id"],),
        one=True
    )["cnt"]
    
    tasks_my = query_db(
        "SELECT COUNT(*) as cnt FROM tasks WHERE org_id=? AND assignee_id=? AND status NOT IN ('done', 'cancelled')",
        (user["org_id"], user["id"]),
        one=True
    )["cnt"]
    
    # Deals summary
    deals_open = query_db(
        "SELECT COUNT(*) as cnt, SUM(amount) as total FROM deals WHERE org_id=? AND status='open'",
        (user["org_id"],),
        one=True
    )
    
    # Inbox summary
    inbox_open = query_db(
        "SELECT COUNT(*) as cnt FROM inbox_threads WHERE org_id=? AND status='open'",
        (user["org_id"],),
        one=True
    )["cnt"]
    
    return jsonify(
        ok=True,
        widgets={
            "tasks": {
                "open": tasks_open,
                "my": tasks_my
            },
            "deals": {
                "open": deals_open["cnt"] or 0,
                "pipeline_value": float(deals_open["total"] or 0)
            },
            "inbox": {
                "open": inbox_open
            }
        }
    )


# ===== BLOCK: ERROR HANDLERS =====

@app.errorhandler(400)
def handle_400(e):
    """Bad Request handler"""
    if request.path.startswith("/api/"):
        return jsonify(ok=False, error="Bad Request"), 400
    return render_template_string(ERROR_400_TMPL, error=str(e)), 400


@app.errorhandler(403)
def handle_403(e):
    """Forbidden handler"""
    if request.path.startswith("/api/"):
        return jsonify(ok=False, error="Forbidden"), 403
    return render_template_string(ERROR_403_TMPL), 403


@app.errorhandler(404)
def handle_404(e):
    """Not Found handler"""
    if request.path.startswith("/api/"):
        return jsonify(ok=False, error="Not Found"), 404
    return render_template_string(ERROR_404_TMPL), 404


@app.errorhandler(500)
def handle_500(e):
    """Internal Server Error handler"""
    log("ERROR", f"500 error: {e}", exc_info=True)
    
    # Increment error metric
    _increment_metric("errors_total")
    
    if request.path.startswith("/api/"):
        return jsonify(ok=False, error="Internal Server Error"), 500
    return render_template_string(ERROR_500_TMPL), 500


# ===== BLOCK: TEMPLATE VARIABLE DECLARATIONS =====

# Templates are defined in STYLES PART 1-10
# Here we declare them as module-level variables (will be assigned in STYLES parts)

# Authentication templates
LOGIN_TMPL = ""
REGISTER_TMPL = ""
APPROVAL_TMPL = ""

# Layout & common
LAYOUT_TMPL = ""
ERROR_400_TMPL = ""
ERROR_403_TMPL = ""
ERROR_404_TMPL = ""
ERROR_500_TMPL = ""

# Main pages
DASHBOARD_TMPL = ""
PROFILE_TMPL = ""
INBOX_TMPL = ""
THREAD_TMPL = ""
TASKS_TMPL = ""
TASK_VIEW_TMPL = ""
DEALS_TMPL = ""
CLIENTS_TMPL = ""
CLIENT_PAGE_TMPL = ""
CALLS_TMPL = ""
MEETING_TMPL = ""
CHAT_TMPL = ""
SEARCH_TMPL = ""
SETTINGS_TMPL = ""
DOCUMENTS_TMPL = ""
DOCUMENT_VIEW_TMPL = ""
WAREHOUSE_TMPL = ""
IMPORT_TMPL = ""
ANALYTICS_TMPL = ""


# ===== BLOCK: SERVER STARTUP =====

def _start_server():
    """
    Main server startup function.
    ✅ ИСПРАВЛЕНО: graceful initialization sequence
    """
    try:
        # Step 1: Handle CLI migration flag
        run_migrations_cli_if_requested()
    except SystemExit:
        # Migration mode - exit after migrations
        return
    except Exception as e:
        log("ERROR", f"Migration CLI check failed: {e}")
    
    try:
        # Step 2: Ensure database schema
        log("INFO", "Initializing database schema...")
        ensure_schema()
        log("INFO", "Database schema ready")
    except Exception as e:
        log("CRITICAL", f"Database initialization failed: {e}")
        sys.exit(1)
    
    try:
        # Step 3: Seed default data
        log("INFO", "Seeding default data...")
        seed_defaults()
        log("INFO", "Default data seeded")
    except Exception as e:
        log("ERROR", f"Seeding failed: {e}")
    
    try:
        # Step 4: Start background workers
        log("INFO", "Starting background workers...")
        start_workers_once()
        log("INFO", "Background workers started")
    except Exception as e:
        log("ERROR", f"Worker startup failed: {e}")
    
    # Step 5: Log startup info
    log("INFO", "=" * 80)
    log("INFO", f"  🚀 CRM/ERP System v{SCHEMA_VERSION}")
    log("INFO", f"  Environment: {ENV}")
    log("INFO", f"  Debug: {DEBUG}")
    log("INFO", f"  Host: {HOST}:{PORT}")
    log("INFO", f"  Database: {DATABASE_PATH}")
    log("INFO", f"  Storage: {STORAGE_BACKEND}")
    log("INFO", f"  AI Provider: {AI_PROVIDER}")
    log("INFO", f"  Redis: {'✓' if REDIS_AVAILABLE else '✗'}")
    log("INFO", f"  SSE: {'✓' if SSE_ENABLED else '✗'}")
    log("INFO", "=" * 80)
    log("INFO", "🎯 Server starting...")
    
    # Step 6: Start Flask app
    try:
        app.run(
            host=HOST,
            port=PORT,
            debug=DEBUG,
            threaded=True,
            use_reloader=DEBUG  # Auto-reload only in debug mode
        )
    except KeyboardInterrupt:
        log("INFO", "Server stopped by user")
    except Exception as e:
        log("CRITICAL", f"Server startup failed: {e}")
        sys.exit(1)


# ===== BLOCK: CLI UTILITIES =====

def print_help():
    """Print CLI help"""
    help_text = """
╔═══════════════════════════════════════════════════════════════════════════╗
║                    CRM/ERP SYSTEM - CLI HELP                              ║
╚═══════════════════════════════════════════════════════════════════════════╝

Usage: python app.py [OPTIONS]

Options:
  --migrate              Run database migrations and exit
  --help                 Show this help message and exit
  
Environment Variables:
  ENV                    Environment (development|production) [default: development]
  HOST                   Server host [default: 0.0.0.0]
  PORT                   Server port [default: 5000]
  SECRET_KEY             Flask secret key (REQUIRED in production)
  DATABASE_PATH          SQLite database path [default: ./crm.db]
  STORAGE_BACKEND        Storage backend (local|s3) [default: local]
  AI_PROVIDER            AI provider (openai|anthropic) [default: openai]
  AI_API_KEY             AI API key
  REDIS_URL              Redis connection URL
  
Examples:
  # Run migrations
  python app.py --migrate
  
  # Start server (development)
  python app.py
  
  # Start server (production)
  ENV=production SECRET_KEY=xxx python app.py
  
  # Run with custom port
  PORT=8080 python app.py

Default credentials (after first run):
  Username: admin
  Password: admin
  
  ⚠️  CHANGE DEFAULT PASSWORD IN PRODUCTION!

Documentation: https://github.com/yourorg/crm-erp
"""
    print(help_text)


# ===== BLOCK: MAIN ENTRY POINT PLACEHOLDER =====

# NOTE: Actual entry point will be at the END of STYLES PART 10/10
# This ensures all templates are loaded before server starts

# Placeholder for module-level execution check
# The actual if __name__ == "__main__" block will be in STYLES PART 10/10


# ═════════════════════════════════════════════════════════════════════════════
# END OF CORE PARTS 1-10
# 
# Total: ~8500 lines of backend code
# 
# Next: STYLES PARTS 1-10 (templates, frontend JS, CSS)
# ═════════════════════════════════════════════════════════════════════════════
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 1/10 — LAYOUT, CSS, AUTH TEMPLATES, ERROR PAGES
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: LAYOUT (BASE) =====

LAYOUT_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{% block title %}CRM/ERP System{% endblock %}</title>
    <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>📊</text></svg>">
    <style>
        /* ===== CSS DESIGN SYSTEM ===== */
        
        :root {
            /* Colors - Light theme */
            --bg: #ffffff;
            --fg: #1a1a1a;
            --surface: #f5f5f5;
            --panel: #fafafa;
            --border: #e0e0e0;
            --muted: #757575;
            --accent: #2bd66a;
            --warn: #ff9800;
            --err: #f44336;
            --ok: #4caf50;
            
            /* Spacing */
            --spacing-xs: 4px;
            --spacing-sm: 8px;
            --spacing-md: 16px;
            --spacing-lg: 24px;
            --spacing-xl: 32px;
            
            /* Border radius */
            --radius-sm: 6px;
            --radius-md: 10px;
            --radius-lg: 12px;
            
            /* Shadows */
            --shadow-sm: 0 1px 3px rgba(0,0,0,0.05);
            --shadow-md: 0 4px 6px rgba(0,0,0,0.07);
            --shadow-lg: 0 10px 15px rgba(0,0,0,0.1);
            
            /* Typography */
            --font-sans: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            --font-mono: "SF Mono", Monaco, "Cascadia Code", "Roboto Mono", Consolas, monospace;
        }
        
        /* Dark theme */
        [data-theme="dark"] {
            --bg: #1a1a1a;
            --fg: #e0e0e0;
            --surface: #2a2a2a;
            --panel: #242424;
            --border: #404040;
            --muted: #9e9e9e;
            --accent: #2bd66a;
            --warn: #ff9800;
            --err: #f44336;
            --ok: #4caf50;
        }
        
        /* Reset & Base */
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        html {
            font-size: 16px;
        }
        
        body {
            font-family: var(--font-sans);
            background: var(--bg);
            color: var(--fg);
            line-height: 1.6;
            min-height: 100vh;
        }
        
        /* Layout */
        .container {
            max-width: 1400px;
            margin: 0 auto;
            padding: 0 var(--spacing-md);
        }
        
        .split {
            display: grid;
            gap: var(--spacing-md);
            grid-template-columns: 1fr 1fr;
        }
        
        @media (max-width: 768px) {
            .split {
                grid-template-columns: 1fr;
            }
        }
        
        /* Topbar */
        .topbar {
            background: var(--surface);
            border-bottom: 1px solid var(--border);
            padding: var(--spacing-sm) 0;
            position: sticky;
            top: 0;
            z-index: 100;
        }
        
        .topbar-inner {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: var(--spacing-md);
        }
        
        .topbar-brand {
            font-weight: 700;
            font-size: 1.125rem;
            color: var(--fg);
            text-decoration: none;
        }
        
        .topbar-search {
            flex: 1;
            max-width: 400px;
        }
        
        .topbar-actions {
            display: flex;
            align-items: center;
            gap: var(--spacing-sm);
        }
        
        /* Navigation */
        .nav {
            display: flex;
            gap: var(--spacing-xs);
            list-style: none;
            flex-wrap: wrap;
        }
        
        .nav a {
            padding: var(--spacing-sm) var(--spacing-md);
            border-radius: var(--radius-sm);
            color: var(--fg);
            text-decoration: none;
            transition: background 0.2s;
        }
        
        .nav a:hover {
            background: var(--panel);
        }
        
        .nav a.active {
            background: var(--accent);
            color: #000;
        }
        
        /* Main content */
        .main {
            padding: var(--spacing-lg) 0;
            min-height: calc(100vh - 120px);
        }
        
        /* Card */
        .card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius-lg);
            padding: var(--spacing-md);
            box-shadow: var(--shadow-sm);
        }
        
        /* Form elements */
        .input,
        .select,
        textarea.input {
            width: 100%;
            padding: var(--spacing-sm) var(--spacing-md);
            background: var(--bg);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            color: var(--fg);
            font-family: inherit;
            font-size: 0.875rem;
            transition: border-color 0.2s;
        }
        
        .input:focus,
        .select:focus,
        textarea.input:focus {
            outline: none;
            border-color: var(--accent);
        }
        
        .input::placeholder {
            color: var(--muted);
        }
        
        /* Buttons */
        .button {
            display: inline-flex;
            align-items: center;
            gap: var(--spacing-xs);
            padding: var(--spacing-sm) var(--spacing-md);
            background: var(--accent);
            color: #000;
            border: none;
            border-radius: var(--radius-sm);
            font-family: inherit;
            font-size: 0.875rem;
            font-weight: 600;
            cursor: pointer;
            transition: filter 0.2s;
        }
        
        .button:hover {
            filter: brightness(0.95);
        }
        
        .button:disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }
        
        .button.secondary {
            background: var(--surface);
            color: var(--fg);
            border: 1px solid var(--border);
        }
        
        .button.ghost {
            background: transparent;
            color: var(--fg);
            border: 1px solid var(--border);
        }
        
        .button.warn {
            background: var(--warn);
            color: #fff;
        }
        
        .button.small {
            padding: 4px 8px;
            font-size: 0.75rem;
        }
        
        .iconbtn {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 4px 8px;
            background: transparent;
            color: var(--fg);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            font-size: 0.75rem;
            cursor: pointer;
            text-decoration: none;
            transition: background 0.2s;
        }
        
        .iconbtn:hover {
            background: var(--panel);
        }
        
        .iconbtn.phone {
            background: var(--ok);
            color: #fff;
            border: none;
        }
        
        /* Table */
        .table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.875rem;
        }
        
        .table thead {
            background: var(--panel);
        }
        
        .table th,
        .table td {
            padding: var(--spacing-sm);
            text-align: left;
            border-bottom: 1px solid var(--border);
        }
        
        .table th {
            font-weight: 600;
        }
        
        .table tbody tr:hover {
            background: var(--panel);
        }
        
        .table tbody tr:focus {
            outline: 2px solid var(--accent);
            outline-offset: -2px;
        }
        
        /* Badge */
        .badge {
            display: inline-block;
            padding: 2px 6px;
            font-size: 0.75rem;
            border-radius: 4px;
            font-weight: 600;
        }
        
        .badge.ok {
            background: rgba(76, 175, 80, 0.15);
            color: var(--ok);
        }
        
        .badge.warn {
            background: rgba(255, 152, 0, 0.15);
            color: var(--warn);
        }
        
        .badge.err {
            background: rgba(244, 67, 54, 0.15);
            color: var(--err);
        }
        
        /* Modal */
        .modal-backdrop {
            display: none;
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0, 0, 0, 0.5);
            z-index: 1000;
            align-items: center;
            justify-content: center;
            padding: var(--spacing-md);
        }
        
        .modal-backdrop.show {
            display: flex;
        }
        
        .modal {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius-lg);
            padding: var(--spacing-lg);
            max-width: 600px;
            width: 100%;
            max-height: 90vh;
            overflow: auto;
            box-shadow: var(--shadow-lg);
        }
        
        /* Helper classes */
        .help {
            color: var(--muted);
            font-size: 0.875rem;
        }
        
        .row {
            display: flex;
            align-items: center;
            gap: var(--spacing-sm);
        }
        
        .grid-filters {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: var(--spacing-sm);
        }
        
        /* Toast notifications */
        .toast {
            position: fixed;
            bottom: var(--spacing-lg);
            right: var(--spacing-lg);
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius-md);
            padding: var(--spacing-md);
            box-shadow: var(--shadow-lg);
            z-index: 2000;
            animation: slideIn 0.3s ease;
        }
        
        @keyframes slideIn {
            from {
                transform: translateY(100px);
                opacity: 0;
            }
            to {
                transform: translateY(0);
                opacity: 1;
            }
        }
        
        /* Details/summary */
        details summary {
            cursor: pointer;
            user-select: none;
        }
        
        /* Responsive */
        @media (max-width: 768px) {
            .topbar-search {
                display: none;
            }
            
            .nav {
                overflow-x: auto;
                white-space: nowrap;
            }
        }
        
        /* Utility */
        .form-fixed label {
            display: block;
            margin-bottom: var(--spacing-xs);
            font-size: 0.875rem;
            font-weight: 500;
        }
    </style>
</head>
<body data-theme="{{ user.get('theme', 'light') if user else 'light' }}">
    {% if user %}
    <div class="topbar">
        <div class="container topbar-inner">
            <a href="/" class="topbar-brand">📊 CRM/ERP</a>
            
            <div class="topbar-search">
                <form action="/search" method="get" style="margin:0;">
                    <input class="input" name="q" placeholder="Поиск... (Ctrl+K)" autocomplete="off">
                </form>
            </div>
            
            <div class="topbar-actions">
                <span class="help">{{ user.username }}</span>
                <a href="/profile" class="iconbtn">Профиль</a>
                {% if user.role == 'admin' %}
                <a href="/settings" class="iconbtn">Настройки</a>
                {% endif %}
                <form method="post" action="/logout" style="margin:0;">
                    <input type="hidden" name="csrf_token" value="{{ session.get('csrf_token', '') }}">
                    <button class="button ghost small" type="submit">Выход</button>
                </form>
            </div>
        </div>
    </div>
    
    <nav class="container" style="margin-top:8px;">
        <ul class="nav">
            <li><a href="/" {% if request.path == '/' %}class="active"{% endif %}>Главная</a></li>
            <li><a href="/inbox" {% if request.path.startswith('/inbox') or request.path.startswith('/thread') %}class="active"{% endif %}>Inbox</a></li>
            <li><a href="/tasks" {% if request.path.startswith('/task') %}class="active"{% endif %}>Задачи</a></li>
            <li><a href="/deals" {% if request.path.startswith('/deal') %}class="active"{% endif %}>Сделки</a></li>
            <li><a href="/clients" {% if request.path.startswith('/client') %}class="active"{% endif %}>Клиенты</a></li>
            <li><a href="/calls" {% if request.path.startswith('/call') %}class="active"{% endif %}>Звонки</a></li>
            <li><a href="/meetings" {% if request.path.startswith('/meeting') %}class="active"{% endif %}>Встречи</a></li>
            <li><a href="/chat" {% if request.path.startswith('/chat') %}class="active"{% endif %}>Чат</a></li>
            <li><a href="/documents" {% if request.path.startswith('/document') %}class="active"{% endif %}>Документы</a></li>
            <li><a href="/analytics" {% if request.path.startswith('/analytics') %}class="active"{% endif %}>Аналитика</a></li>
        </ul>
    </nav>
    {% endif %}
    
    <main class="main container">
        {% block content %}{% endblock %}
    </main>
    
    <div id="toastContainer"></div>
    
    <script nonce="{{ csp_nonce }}">
        // ===== GLOBAL CONSTANTS =====
        const CSRF = '{{ session.get("csrf_token", "") }}';
        const USER_ID = {{ user.id if user else 'null' }};
        
        // ===== UTILITY FUNCTIONS =====
        
        /**
         * Escape HTML to prevent XSS
         */
        function esc(str) {
            if (str == null) return '';
            const div = document.createElement('div');
            div.textContent = String(str);
            return div.innerHTML;
        }
        
        /**
         * Show toast notification
         */
        function toast(message, duration = 3000) {
            const container = document.getElementById('toastContainer');
            if (!container) return;
            
            const toast = document.createElement('div');
            toast.className = 'toast';
            toast.textContent = message;
            
            container.appendChild(toast);
            
            setTimeout(() => {
                toast.style.animation = 'slideIn 0.3s ease reverse';
                setTimeout(() => toast.remove(), 300);
            }, duration);
        }
        
        /**
         * Normalize phone number (client-side helper)
         */
        function normPhoneRU(phone) {
            if (!phone) return '';
            const digits = phone.replace(/\D+/g, '');
            
            if (digits.length === 11 && digits.startsWith('8')) {
                return '+7' + digits.slice(1);
            }
            if (digits.length === 11 && digits.startsWith('7')) {
                return '+' + digits;
            }
            if (digits.length === 10) {
                return '+7' + digits;
            }
            if (digits.length > 10) {
                return '+' + digits;
            }
            
            return phone;
        }
        
        /**
         * Format date to local string
         */
        function formatDate(isoString) {
            if (!isoString) return '';
            try {
                const date = new Date(isoString);
                return date.toLocaleString('ru-RU');
            } catch (e) {
                return isoString;
            }
        }
        
        /**
         * Confirm before action (for .js-confirm forms)
         */
        document.addEventListener('submit', (e) => {
            const form = e.target;
            if (form && form.classList.contains('js-confirm')) {
                const msg = form.getAttribute('data-confirm') || 'Вы уверены?';
                if (!confirm(msg)) {
                    e.preventDefault();
                }
            }
        });
        
        /**
         * Auto-close modals on backdrop click
         */
        document.addEventListener('click', (e) => {
            if (e.target.classList.contains('modal-backdrop')) {
                e.target.classList.remove('show');
            }
        });
        
        /**
         * Keyboard shortcut: Ctrl+K for search
         */
        document.addEventListener('keydown', (e) => {
            if ((e.ctrlKey || e.metaKey) && e.key === 'k') {
                e.preventDefault();
                const searchInput = document.querySelector('.topbar-search input');
                if (searchInput) searchInput.focus();
            }
        });
        
        /**
         * Auto-focus first input in modals
         */
        const observer = new MutationObserver((mutations) => {
            mutations.forEach((mutation) => {
                mutation.addedNodes.forEach((node) => {
                    if (node.classList && node.classList.contains('modal-backdrop') && node.classList.contains('show')) {
                        const firstInput = node.querySelector('input:not([type=hidden]), textarea, select');
                        if (firstInput) {
                            setTimeout(() => firstInput.focus(), 100);
                        }
                    }
                });
            });
        });
        
        if (document.body) {
            observer.observe(document.body, { childList: true, subtree: true });
        }
    </script>
    
    {% block scripts %}{% endblock %}
</body>
</html>
"""


# ===== TEMPLATE: LOGIN =====

LOGIN_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Вход — CRM/ERP</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .login-box {
            background: #fff;
            border-radius: 12px;
            padding: 40px;
            width: 100%;
            max-width: 400px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
        }
        h1 {
            text-align: center;
            margin-bottom: 30px;
            color: #333;
        }
        .form-group {
            margin-bottom: 20px;
        }
        label {
            display: block;
            margin-bottom: 6px;
            font-size: 14px;
            font-weight: 500;
            color: #555;
        }
        input {
            width: 100%;
            padding: 10px 12px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 14px;
            transition: border-color 0.2s;
        }
        input:focus {
            outline: none;
            border-color: #667eea;
        }
        button {
            width: 100%;
            padding: 12px;
            background: #667eea;
            color: #fff;
            border: none;
            border-radius: 6px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: background 0.2s;
        }
        button:hover {
            background: #5568d3;
        }
        .error {
            background: #fee;
            color: #c33;
            padding: 10px;
            border-radius: 6px;
            margin-bottom: 20px;
            font-size: 14px;
        }
        .help {
            text-align: center;
            margin-top: 20px;
            font-size: 14px;
            color: #888;
        }
        .help a {
            color: #667eea;
            text-decoration: none;
        }
    </style>
</head>
<body>
    <div class="login-box">
        <h1>📊 Вход в CRM</h1>
        
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        
        <form method="post">
            <div class="form-group">
                <label>Логин</label>
                <input type="text" name="username" required autofocus>
            </div>
            
            <div class="form-group">
                <label>Пароль</label>
                <input type="password" name="password" required>
            </div>
            
            {% if show_2fa %}
            <div class="form-group">
                <label>2FA код</label>
                <input type="text" name="totp_token" pattern="[0-9]{6}" placeholder="123456" required>
            </div>
            {% endif %}
            
            <button type="submit">Войти</button>
        </form>
        
        <div class="help">
            Нет аккаунта? <a href="/register">Зарегистрироваться</a>
        </div>
    </div>
</body>
</html>
"""


# ===== TEMPLATE: REGISTER =====

REGISTER_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Регистрация — CRM/ERP</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            padding: 20px;
        }
        .register-box {
            background: #fff;
            border-radius: 12px;
            padding: 40px;
            width: 100%;
            max-width: 450px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
        }
        h1 {
            text-align: center;
            margin-bottom: 30px;
            color: #333;
        }
        .form-group {
            margin-bottom: 20px;
        }
        label {
            display: block;
            margin-bottom: 6px;
            font-size: 14px;
            font-weight: 500;
            color: #555;
        }
        input {
            width: 100%;
            padding: 10px 12px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 14px;
        }
        input:focus {
            outline: none;
            border-color: #667eea;
        }
        button {
            width: 100%;
            padding: 12px;
            background: #667eea;
            color: #fff;
            border: none;
            border-radius: 6px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
        }
        button:hover {
            background: #5568d3;
        }
        .error {
            background: #fee;
            color: #c33;
            padding: 10px;
            border-radius: 6px;
            margin-bottom: 20px;
            font-size: 14px;
        }
        .help {
            text-align: center;
            margin-top: 20px;
            font-size: 14px;
            color: #888;
        }
        .help a {
            color: #667eea;
            text-decoration: none;
        }
    </style>
</head>
<body>
    <div class="register-box">
        <h1>📊 Регистрация</h1>
        
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        
        <form method="post">
            <div class="form-group">
                <label>Логин</label>
                <input type="text" name="username" required autofocus>
            </div>
            
            <div class="form-group">
                <label>Email</label>
                <input type="email" name="email">
            </div>
            
            <div class="form-group">
                <label>Пароль (минимум 12 символов)</label>
                <input type="password" name="password" minlength="12" required>
            </div>
            
            <div class="form-group">
                <label>Slug организации</label>
                <input type="text" name="org_slug" pattern="[a-z0-9_-]+" required placeholder="mycompany">
            </div>
            
            <button type="submit">Зарегистрироваться</button>
        </form>
        
        <div class="help">
            Уже есть аккаунт? <a href="/login">Войти</a>
        </div>
    </div>
</body>
</html>
"""


# ===== TEMPLATE: APPROVAL (PUBLIC) =====

APPROVAL_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Согласование</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: #f5f5f5;
            padding: 20px;
        }
        .approval-box {
            max-width: 600px;
            margin: 40px auto;
            background: #fff;
            border-radius: 12px;
            padding: 30px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }
        h1 {
            margin-bottom: 20px;
            color: #333;
        }
        .entity-info {
            background: #f9f9f9;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
        }
        .entity-info p {
            margin: 8px 0;
        }
        label {
            display: block;
            margin-bottom: 6px;
            font-weight: 500;
        }
        textarea {
            width: 100%;
            padding: 10px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-family: inherit;
            margin-bottom: 20px;
        }
        .button-group {
            display: flex;
            gap: 10px;
        }
        button {
            flex: 1;
            padding: 12px;
            border: none;
            border-radius: 6px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
        }
        .approve {
            background: #4caf50;
            color: #fff;
        }
        .reject {
            background: #f44336;
            color: #fff;
        }
    </style>
</head>
<body>
    <div class="approval-box">
        <h1>Согласование</h1>
        
        <div class="entity-info">
            <p><strong>Тип:</strong> {{ entity_type }}</p>
            <p><strong>ID:</strong> {{ entity.id }}</p>
            {% if entity.title %}
            <p><strong>Название:</strong> {{ entity.title }}</p>
            {% endif %}
            {% if entity.amount %}
            <p><strong>Сумма:</strong> {{ entity.amount }} {{ entity.currency or 'RUB' }}</p>
            {% endif %}
        </div>
        
        <form method="post">
            <label>Комментарий (опционально)</label>
            <textarea name="comment" rows="4" placeholder="Ваш комментарий..."></textarea>
            
            <div class="button-group">
                <button type="submit" name="decision" value="approve" class="approve">✓ Одобрить</button>
                <button type="submit" name="decision" value="reject" class="reject">✗ Отклонить</button>
            </div>
        </form>
    </div>
</body>
</html>
"""


# ===== TEMPLATES: ERROR PAGES =====

ERROR_400_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>400 — Bad Request</title>
    <style>
        body { font-family: -apple-system, sans-serif; text-align: center; padding: 50px; background: #f5f5f5; }
        h1 { font-size: 48px; margin: 0; color: #333; }
        p { color: #666; margin: 20px 0; }
        a { color: #2bd66a; text-decoration: none; font-weight: 600; }
    </style>
</head>
<body>
    <h1>400</h1>
    <p>Некорректный запрос</p>
    <p style="font-size: 14px; color: #999;">{{ error }}</p>
    <a href="/">← На главную</a>
</body>
</html>
"""

ERROR_403_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>403 — Forbidden</title>
    <style>
        body { font-family: -apple-system, sans-serif; text-align: center; padding: 50px; background: #f5f5f5; }
        h1 { font-size: 48px; margin: 0; color: #333; }
        p { color: #666; margin: 20px 0; }
        a { color: #2bd66a; text-decoration: none; font-weight: 600; }
    </style>
</head>
<body>
    <h1>403</h1>
    <p>Доступ запрещен</p>
    <a href="/">← На главную</a>
</body>
</html>
"""

ERROR_404_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>404 — Не найдено</title>
    <style>
        body { font-family: -apple-system, sans-serif; text-align: center; padding: 50px; background: #f5f5f5; }
        h1 { font-size: 48px; margin: 0; color: #333; }
        p { color: #666; margin: 20px 0; }
        a { color: #2bd66a; text-decoration: none; font-weight: 600; }
    </style>
</head>
<body>
    <h1>404</h1>
    <p>Страница не найдена</p>
    <a href="/">← На главную</a>
</body>
</html>
"""

ERROR_500_TMPL = """<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <title>500 — Внутренняя ошибка</title>
    <style>
        body { font-family: -apple-system, sans-serif; text-align: center; padding: 50px; background: #f5f5f5; }
        h1 { font-size: 48px; margin: 0; color: #333; }
        p { color: #666; margin: 20px 0; }
        a { color: #2bd66a; text-decoration: none; font-weight: 600; }
    </style>
</head>
<body>
    <h1>500</h1>
    <p>Внутренняя ошибка сервера</p>
    <p style="font-size: 14px; color: #999;">Попробуйте позже или обратитесь к администратору</p>
    <a href="/">← На главную</a>
</body>
</html>
"""


# ===== END OF STYLES PART 1/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 2/10 — DASHBOARD, PROFILE, INBOX TEMPLATES
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: DASHBOARD =====

DASHBOARD_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="margin-bottom: 20px;">
    <h1 style="margin: 0 0 8px 0;">Добро пожаловать, {{ user.username }}!</h1>
    <p class="help">Обзор активности</p>
</div>

<div class="split" style="margin-bottom: 20px;">
    <div class="card">
        <h3 style="margin: 0 0 12px 0;">📋 Задачи</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 12px;">
            <div>
                <div style="font-size: 32px; font-weight: 700; color: var(--accent);">{{ open_tasks }}</div>
                <div class="help">Открытых</div>
            </div>
            <div>
                <div style="font-size: 32px; font-weight: 700; color: var(--warn);">{{ my_tasks }}</div>
                <div class="help">Моих</div>
            </div>
        </div>
        <a href="/tasks" class="button" style="margin-top: 16px; width: 100%;">Перейти к задачам →</a>
    </div>
    
    <div class="card">
        <h3 style="margin: 0 0 12px 0;">💼 Сделки</h3>
        <div>
            <div style="font-size: 32px; font-weight: 700; color: var(--ok);">{{ open_deals }}</div>
            <div class="help">Активных сделок</div>
        </div>
        <a href="/deals" class="button" style="margin-top: 16px; width: 100%;">Перейти к сделкам →</a>
    </div>
</div>

<div class="card">
    <h3 style="margin: 0 0 12px 0;">Быстрые действия</h3>
    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 8px;">
        <button class="button secondary" onclick="location.href='/tasks'">Создать задачу</button>
        <button class="button secondary" onclick="location.href='/deals'">Создать сделку</button>
        <button class="button secondary" onclick="location.href='/clients'">Добавить клиента</button>
        <button class="button secondary" onclick="location.href='/inbox'">Открыть Inbox</button>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
// Load dashboard widgets via API
async function loadWidgets() {
    try {
        const r = await fetch('/api/dashboard/widgets');
        const j = await r.json();
        if (j.ok) {
            console.log('Dashboard widgets loaded:', j.widgets);
        }
    } catch (e) {
        console.error('Failed to load widgets:', e);
    }
}

loadWidgets();
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: PROFILE =====

PROFILE_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<h2 style="margin: 0 0 16px 0;">Профиль пользователя</h2>

<div class="split">
    <div>
        <div class="card" style="margin-bottom: 16px;">
            <h3 style="margin: 0 0 12px 0;">Основная информация</h3>
            <div style="display: grid; gap: 12px;">
                <div>
                    <label class="help">Логин</label>
                    <div>{{ user.username }}</div>
                </div>
                <div>
                    <label class="help">Роль</label>
                    <div>{{ user.role }}</div>
                </div>
                <div>
                    <label class="help">Email</label>
                    <input class="input" id="email" value="{{ user.email or '' }}" placeholder="example@mail.com">
                </div>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                    <div>
                        <label class="help">Имя</label>
                        <input class="input" id="firstName" value="{{ user.first_name or '' }}">
                    </div>
                    <div>
                        <label class="help">Фамилия</label>
                        <input class="input" id="lastName" value="{{ user.last_name or '' }}">
                    </div>
                </div>
                <div>
                    <label class="help">Телефон</label>
                    <input class="input" id="phone" value="{{ user.phone or '' }}" placeholder="+7 495 ...">
                </div>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                    <div>
                        <label class="help">Часовой пояс</label>
                        <select class="select" id="timezone">
                            {% for tz in timezones %}
                            <option value="{{ tz }}" {% if user.timezone == tz %}selected{% endif %}>{{ tz }}</option>
                            {% endfor %}
                        </select>
                    </div>
                    <div>
                        <label class="help">Язык</label>
                        <select class="select" id="locale">
                            <option value="ru" {% if user.locale == 'ru' %}selected{% endif %}>Русский</option>
                            <option value="en" {% if user.locale == 'en' %}selected{% endif %}>English</option>
                        </select>
                    </div>
                </div>
                <button class="button" id="btnSaveProfile">Сохранить изменения</button>
            </div>
        </div>
        
        <div class="card">
            <h3 style="margin: 0 0 12px 0;">Смена пароля</h3>
            <div style="display: grid; gap: 12px;">
                <input class="input" type="password" id="currentPassword" placeholder="Текущий пароль">
                <input class="input" type="password" id="newPassword" placeholder="Новый пароль (мин. 12 символов)" minlength="12">
                <button class="button secondary" id="btnChangePassword">Сменить пароль</button>
            </div>
        </div>
    </div>
    
    <div>
        <div class="card" style="margin-bottom: 16px;">
            <h3 style="margin: 0 0 12px 0;">Аватар</h3>
            <div style="text-align: center;">
                {% if user.avatar_url %}
                <img src="{{ user.avatar_url }}" alt="Avatar" style="width: 128px; height: 128px; border-radius: 50%; object-fit: cover; margin-bottom: 12px;">
                {% else %}
                <div style="width: 128px; height: 128px; border-radius: 50%; background: var(--surface); display: flex; align-items: center; justify-content: center; margin: 0 auto 12px; font-size: 48px;">
                    {{ user.username[0].upper() }}
                </div>
                {% endif %}
                <input type="file" id="avatarFile" accept="image/*" style="display: none;">
                <button class="button secondary" onclick="document.getElementById('avatarFile').click()">Загрузить фото</button>
            </div>
        </div>
        
        <div class="card">
            <h3 style="margin: 0 0 12px 0;">Двухфакторная аутентификация</h3>
            {% if user.totp_enabled %}
            <div class="badge ok" style="margin-bottom: 12px;">✓ Включена</div>
            <button class="button warn" id="btn2FADisable">Отключить 2FA</button>
            {% else %}
            <div class="badge warn" style="margin-bottom: 12px;">Отключена</div>
            <button class="button" id="btn2FAEnable">Включить 2FA</button>
            {% endif %}
        </div>
        
        <div class="card" style="margin-top: 16px;">
            <h3 style="margin: 0 0 12px 0;">Статистика</h3>
            <div class="help">Создано задач: {{ tasks_created }}</div>
        </div>
    </div>
</div>

<div class="modal-backdrop" id="modal2FA">
    <div class="modal">
        <h3>Настройка 2FA</h3>
        <div id="qrContainer" style="text-align: center; margin: 20px 0;"></div>
        <div class="help" style="margin-bottom: 12px;">Отсканируйте QR-код в приложении (Google Authenticator, Authy)</div>
        <div class="help" style="margin-bottom: 12px;">Секрет: <code id="totpSecret"></code></div>
        <div class="help" style="margin-bottom: 12px; font-weight: bold;">Backup коды (сохраните их!):</div>
        <div id="backupCodes" style="background: var(--panel); padding: 12px; border-radius: 6px; font-family: var(--font-mono); font-size: 14px; margin-bottom: 12px;"></div>
        <button class="button" onclick="document.getElementById('modal2FA').classList.remove('show')">Закрыть</button>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
// Save profile
document.getElementById('btnSaveProfile')?.addEventListener('click', async () => {
    const data = {
        email: document.getElementById('email').value,
        first_name: document.getElementById('firstName').value,
        last_name: document.getElementById('lastName').value,
        phone: document.getElementById('phone').value,
        timezone: document.getElementById('timezone').value,
        locale: document.getElementById('locale').value
    };
    
    try {
        const r = await fetch('/api/profile', {
            method: 'PATCH',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) {
            toast('Профиль сохранён');
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Change password
document.getElementById('btnChangePassword')?.addEventListener('click', async () => {
    const current = document.getElementById('currentPassword').value;
    const newPass = document.getElementById('newPassword').value;
    
    if (!current || !newPass) return alert('Заполните оба поля');
    if (newPass.length < 12) return alert('Пароль должен быть не менее 12 символов');
    
    try {
        const r = await fetch('/api/profile/password', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify({
                current_password: current,
                new_password: newPass
            })
        });
        const j = await r.json();
        if (j.ok) {
            toast('Пароль изменён');
            document.getElementById('currentPassword').value = '';
            document.getElementById('newPassword').value = '';
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Avatar upload
document.getElementById('avatarFile')?.addEventListener('change', async (e) => {
    const file = e.target.files[0];
    if (!file) return;
    
    const formData = new FormData();
    formData.append('file', file);
    
    try {
        const r = await fetch('/api/profile/avatar', {
            method: 'POST',
            headers: {
                'X-CSRFToken': CSRF
            },
            body: formData
        });
        const j = await r.json();
        if (j.ok) {
            toast('Аватар загружен');
            location.reload();
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка загрузки');
    }
});

// Enable 2FA
document.getElementById('btn2FAEnable')?.addEventListener('click', async () => {
    try {
        const r = await fetch('/api/profile/2fa/enable', {
            method: 'POST',
            headers: {
                'X-CSRFToken': CSRF
            }
        });
        const j = await r.json();
        if (j.ok) {
            document.getElementById('totpSecret').textContent = j.secret;
            document.getElementById('backupCodes').textContent = (j.backup_codes || []).join('\\n');
            
            const qrContainer = document.getElementById('qrContainer');
            qrContainer.innerHTML = '<img src="https://api.qrserver.com/v1/create-qr-code/?size=200x200&data=' + encodeURIComponent(j.qr_url) + '" alt="QR Code">';
            
            document.getElementById('modal2FA').classList.add('show');
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Disable 2FA
document.getElementById('btn2FADisable')?.addEventListener('click', async () => {
    const password = prompt('Введите пароль для отключения 2FA:');
    if (!password) return;
    
    try {
        const r = await fetch('/api/profile/2fa/disable', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify({ password })
        });
        const j = await r.json();
        if (j.ok) {
            toast('2FA отключена');
            location.reload();
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: INBOX LIST =====

INBOX_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<h2 style="margin: 0 0 8px 0;">Входящие</h2>

<div class="card">
    <details open>
        <summary class="button ghost">Фильтры</summary>
        <form method="get" class="grid-filters" action="/inbox">
            <label>Статус
                <select class="select" name="status">
                    <option value="">— все —</option>
                    <option value="open" {% if filters.status=='open' %}selected{% endif %}>Open</option>
                    <option value="pending" {% if filters.status=='pending' %}selected{% endif %}>Pending</option>
                    <option value="resolved" {% if filters.status=='resolved' %}selected{% endif %}>Resolved</option>
                    <option value="snoozed" {% if filters.status=='snoozed' %}selected{% endif %}>Snoozed</option>
                </select>
            </label>
            <label>Канал
                <select class="select" name="channel">
                    <option value="">— все —</option>
                    {% for c in channels %}
                    <option value="{{ c.id }}" {% if filters.channel|string==c.id|string %}selected{% endif %}>{{ c.name }} ({{ c.type }})</option>
                    {% endfor %}
                </select>
            </label>
            <label>Назначено
                <select class="select" name="assignee">
                    <option value="">— любой —</option>
                    {% for a in agents %}
                    <option value="{{ a.id }}" {% if filters.assignee|string==a.id|string %}selected{% endif %}>{{ a.username }}</option>
                    {% endfor %}
                </select>
            </label>
            <label>Мои
                <select class="select" name="who">
                    <option value="">—</option>
                    <option value="me" {% if filters.who=='me' %}selected{% endif %}>Да</option>
                </select>
            </label>
            <label>Поиск
                <input class="input" name="q" value="{{ filters.q or '' }}" placeholder="по теме">
            </label>
            <div style="grid-column: 1/-1; display: flex; gap: 8px; justify-content: flex-end; margin-top: 4px;">
                <button class="button" type="submit">Применить</button>
                <a class="button ghost" href="/inbox">Сбросить</a>
            </div>
        </form>
    </details>
</div>

<div class="card" style="margin-top: 10px;">
    <table class="table">
        <thead>
            <tr>
                <th>ID</th>
                <th>Тема</th>
                <th>Канал</th>
                <th>Статус</th>
                <th>Приоритет</th>
                <th>Назначен</th>
                <th>FRT</th>
                <th>Обновлен</th>
                <th></th>
            </tr>
        </thead>
        <tbody id="inboxTBody">
            {% for t in rows %}
            <tr data-id="{{ t.id }}" tabindex="0">
                <td>#{{ t.id }}</td>
                <td>{{ t.subject or '—' }}</td>
                <td>{{ t.channel_name or '—' }}</td>
                <td>
                    <select class="select th-status" data-id="{{ t.id }}">
                        {% set st = t.status or 'open' %}
                        <option value="open" {% if st=='open' %}selected{% endif %}>open</option>
                        <option value="pending" {% if st=='pending' %}selected{% endif %}>pending</option>
                        <option value="resolved" {% if st=='resolved' %}selected{% endif %}>resolved</option>
                        <option value="snoozed" {% if st=='snoozed' %}selected{% endif %}>snoozed</option>
                    </select>
                </td>
                <td>
                    <select class="select th-priority" data-id="{{ t.id }}">
                        {% set pr = t.priority or 'normal' %}
                        <option value="low" {% if pr=='low' %}selected{% endif %}>low</option>
                        <option value="normal" {% if pr=='normal' %}selected{% endif %}>normal</option>
                        <option value="high" {% if pr=='high' %}selected{% endif %}>high</option>
                        <option value="urgent" {% if pr=='urgent' %}selected{% endif %}>urgent</option>
                    </select>
                </td>
                <td>
                    <select class="select th-assignee" data-id="{{ t.id }}">
                        <option value="">—</option>
                        {% for a in agents %}
                        <option value="{{ a.id }}" {% if t.assignee_id==a.id %}selected{% endif %}>{{ a.username }}</option>
                        {% endfor %}
                    </select>
                </td>
                <td>
                    {% if t.first_response_due_at and (not t.first_response_at) %}
                    <span class="badge {% if now.isoformat(' ','seconds') > t.first_response_due_at %}err{% else %}warn{% endif %}">
                        до {{ t.first_response_due_at }}
                    </span>
                    {% elif t.first_response_at %}
                    <span class="badge ok">ответ: {{ t.first_response_at }}</span>
                    {% else %}
                    —
                    {% endif %}
                </td>
                <td>{{ t.last_message_at or t.created_at }}</td>
                <td><a class="iconbtn small" href="/thread/{{ t.id }}">Открыть</a></td>
            </tr>
            {% else %}
            <tr><td colspan="9"><div class="help">Ничего не найдено</div></td></tr>
            {% endfor %}
        </tbody>
    </table>
    <button class="button ghost" id="loadMoreBtn" style="margin-top: 8px;">Загрузить больше</button>
</div>

<script nonce="{{ csp_nonce }}">
const INBOX_AGENTS = {{ agents|tojson }};
let inboxPage = 1;

function badgeFRT(t) {
    try {
        const due = t.first_response_due_at || '';
        const at = t.first_response_at || '';
        if (due && !at) {
            const nowISO = new Date().toISOString().slice(0,19).replace('T',' ');
            const cls = (nowISO > due) ? 'err' : 'warn';
            return `<span class="badge ${cls}">до ${esc(due)}</span>`;
        }
        if (at) {
            return `<span class="badge ok">ответ: ${esc(at)}</span>`;
        }
    } catch (_) {}
    return '—';
}

function assigneeSelectHTML(id, assignee_id) {
    let opts = '<option value="">—</option>';
    for (const a of INBOX_AGENTS) {
        const sel = (String(assignee_id||'')===String(a.id)) ? ' selected' : '';
        opts += `<option value="${a.id}"${sel}>${esc(a.username)}</option>`;
    }
    return `<select class="select th-assignee" data-id="${id}">${opts}</select>`;
}

function statusSelectHTML(id, status) {
    const sts = ['open','pending','resolved','snoozed'];
    return `<select class="select th-status" data-id="${id}">` +
        sts.map(s => `<option value="${s}"${s===(status||'open')?' selected':''}>${s}</option>`).join('') +
        `</select>`;
}

function prioritySelectHTML(id, pr) {
    const prs = ['low','normal','high','urgent'];
    return `<select class="select th-priority" data-id="${id}">` +
        prs.map(s => `<option value="${s}"${s===(pr||'normal')?' selected':''}>${s}</option>`).join('') +
        `</select>`;
}

function rowHTML(t) {
    return `<tr data-id="${t.id}" tabindex="0">
        <td>#${t.id}</td>
        <td>${esc(t.subject||'—')}</td>
        <td>${esc(t.channel_name||'—')}</td>
        <td>${statusSelectHTML(t.id, t.status)}</td>
        <td>${prioritySelectHTML(t.id, t.priority)}</td>
        <td>${assigneeSelectHTML(t.id, t.assignee_id)}</td>
        <td>${badgeFRT(t)}</td>
        <td>${esc(t.last_message_at || t.created_at || '')}</td>
        <td><a class="iconbtn small" href="/thread/${t.id}">Открыть</a></td>
    </tr>`;
}

function buildParams(page) {
    const params = new URLSearchParams(window.location.search);
    if (page) params.set('page', page);
    params.set('per_page', '50');
    return params;
}

async function loadMoreInbox() {
    inboxPage++;
    try {
        const params = buildParams(inboxPage);
        const url = '/api/inbox/list?' + params.toString();
        const r = await fetch(url);
        const j = await r.json();
        if (!j.ok) {
            alert(j.error || 'Ошибка');
            return;
        }
        const tb = document.getElementById('inboxTBody');
        let added = 0;
        for (const t of (j.items || [])) {
            tb.insertAdjacentHTML('beforeend', rowHTML(t));
            added++;
        }
        if (added < (j.per_page || 50)) {
            document.getElementById('loadMoreBtn').style.display = 'none';
        }
        toast('Загружено больше');
    } catch (e) {
        alert('Ошибка загрузки');
    }
}

async function thUpdateGeneric(id, patch) {
    try {
        const r = await fetch('/api/thread/update', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify(Object.assign({id}, patch))
        });
        const j = await r.json();
        if (!j.ok) {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
}

document.addEventListener('change', (e) => {
    const t = e.target;
    if (!t) return;
    const id = t.getAttribute('data-id');
    if (!id) return;
    if (t.classList.contains('th-status')) thUpdateGeneric(parseInt(id,10), {status: t.value});
    else if (t.classList.contains('th-priority')) thUpdateGeneric(parseInt(id,10), {priority: t.value});
    else if (t.classList.contains('th-assignee')) thUpdateGeneric(parseInt(id,10), {assignee_id: t.value||null});
});

document.getElementById('loadMoreBtn')?.addEventListener('click', (e) => {
    e.preventDefault();
    loadMoreInbox();
});

document.addEventListener('keydown', e => {
    const tag = (e.target && e.target.tagName || '').toUpperCase();
    if (tag === 'INPUT' || tag === 'TEXTAREA' || e.ctrlKey || e.metaKey) return;
    const rows = [...document.querySelectorAll('#inboxTBody tr')];
    const active = (document.activeElement && document.activeElement.closest) ? document.activeElement.closest('tr') : null;
    const idx = rows.indexOf(active);
    if (e.key === 'ArrowDown') {
        e.preventDefault();
        (rows[Math.min(rows.length-1, idx+1)] || rows[0]).focus();
    }
    if (e.key === 'ArrowUp') {
        e.preventDefault();
        (rows[Math.max(0, idx-1)] || rows[0]).focus();
    }
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: THREAD VIEW =====

THREAD_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<h2>Диалог #{{ r.id }}</h2>

<div class="split">
    <div>
        <div class="card" style="margin-bottom: 10px;">
            <div style="display: grid; grid-template-columns: 2fr 1fr 1fr; gap: 8px;">
                <div><strong>Тема:</strong> {{ r.subject or '—' }}</div>
                <div><strong>Канал:</strong> {{ r.channel_id or '—' }}</div>
                <div>
                    <strong>Статус:</strong>
                    <select class="select" id="thStatus">
                        {% set st = r.status or 'open' %}
                        <option value="open" {% if st=='open' %}selected{% endif %}>open</option>
                        <option value="pending" {% if st=='pending' %}selected{% endif %}>pending</option>
                        <option value="resolved" {% if st=='resolved' %}selected{% endif %}>resolved</option>
                        <option value="snoozed" {% if st=='snoozed' %}selected{% endif %}>snoozed</option>
                    </select>
                </div>
            </div>
            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 8px; margin-top: 8px;">
                <div>
                    <strong>Приоритет:</strong>
                    <select class="select" id="thPriority">
                        {% set pr = r.priority or 'normal' %}
                        <option value="low" {% if pr=='low' %}selected{% endif %}>low</option>
                        <option value="normal" {% if pr=='normal' %}selected{% endif %}>normal</option>
                        <option value="high" {% if pr=='high' %}selected{% endif %}>high</option>
                        <option value="urgent" {% if pr=='urgent' %}selected{% endif %}>urgent</option>
                    </select>
                </div>
                <div>
                    <strong>Назначено:</strong>
                    <select class="select" id="assigneeSelect">
                        <option value="">—</option>
                        {% for u in query_db('SELECT id,username FROM users WHERE org_id=? AND active=1 ORDER BY username',(user.org_id,)) %}
                        <option value="{{ u.id }}" {% if r.assignee_id==u.id %}selected{% endif %}>{{ u.username }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div>
                    <strong>Теги:</strong>
                    <input class="input" id="tagsInput" placeholder="tag1, tag2" value="{{ r.tags_csv or '' }}">
                    <button class="iconbtn small" id="btnSaveTags">Сохранить</button>
                </div>
            </div>
        </div>
        
        <div class="card">
            <div id="msgList" style="max-height: 60vh; overflow: auto;">
                {% for m in messages %}
                <div class="msg {{ 'system' if m.internal_note else m.sender_type }}" data-mid="{{ m.id }}" style="margin: 8px 0; padding: 8px; border: 1px solid var(--border); border-radius: 8px;">
                    <div class="meta" style="font-size: 12px; color: var(--muted); margin-bottom: 4px;">
                        [{{ 'Внутренняя' if m.internal_note else m.sender_type }}] {{ m.created_at }} • {{ m.username or m.external_user_id or '—' }}
                    </div>
                    <div class="body" style="white-space: pre-wrap;">{{ m.body or '' }}</div>
                    {% if not m.internal_note %}
                    <div style="margin-top: 6px;">
                        <button class="iconbtn small btnMsgToTask" data-mid="{{ m.id }}">В задачу</button>
                    </div>
                    {% endif %}
                </div>
                {% else %}
                <div class="help">Сообщений пока нет</div>
                {% endfor %}
            </div>
            <button class="button ghost" id="loadOlderBtnThread" style="margin: 8px auto; display: block;">Загрузить ещё</button>
        </div>
        
        <div class="card" style="margin-top: 10px;">
            <div id="dropZone" style="border: 2px dashed var(--border); padding: 10px; text-align: center; cursor: pointer;">
                Перетащите файлы или кликните для загрузки
            </div>
            <input type="file" id="fileInput" multiple style="display: none;">
            <div id="attachments" class="help" style="margin: 6px 0; display: none;"></div>
            <textarea id="body" class="input" rows="4" placeholder="Напишите сообщение..."></textarea>
            <div style="display: flex; gap: 8px; margin-top: 8px; flex-wrap: wrap;">
                <label><input type="checkbox" id="internalNote"> внутренняя заметка</label>
                <div style="margin-left: auto; display: flex; gap: 8px;">
                    <button class="button secondary" type="button" id="btnAIDraft">AI черновик</button>
                    <button class="button" type="button" id="btnSendInternal">Отправить (внутр.)</button>
                    <button class="button warn" type="button" id="btnSendExternal">Клиенту</button>
                </div>
            </div>
        </div>
    </div>
    
    <div>
        <div class="card">
            <h3>Сведения</h3>
            <div><strong>ID треда:</strong> {{ r.id }}</div>
            <div style="margin-top: 8px; display: flex; gap: 8px; flex-wrap: wrap;">
                <a class="button ghost" href="/inbox">← к списку</a>
                <button class="button ghost" type="button" id="btnSnooze">Snooze</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
const TID = {{ r.id }};
let msgPage = 1;
let attList = [];

function thPatch(patch) {
    return fetch('/api/thread/update', {
        method: 'POST',
        headers: {
            'Content-Type': 'application/json',
            'X-CSRFToken': CSRF
        },
        body: JSON.stringify(Object.assign({id: TID}, patch))
    });
}

function saveTags() {
    const raw = (document.getElementById('tagsInput').value || '');
    const arr = [...new Set(raw.split(',').map(x => x.trim()).filter(Boolean))];
    thPatch({tags_json: arr}).then(r => r.json()).then(j => {
        if (!j.ok) {
            alert(j.error || 'Ошибка');
        }
    });
}

function updateAttView() {
    const el = document.getElementById('attachments');
    if (!attList.length) {
        el.style.display = 'none';
        el.textContent = '';
        return;
    }
    el.style.display = 'block';
    el.innerHTML = 'Вложения: ' + attList.map(a => `<a href="${a.url}" target="_blank">${esc(a.name)}</a>`).join(', ');
}

async function uploadFiles(files) {
    try {
        const list = files || (document.getElementById('fileInput')?.files || []);
        if (!list || !list.length) return;
        for (let i = 0; i < list.length; i++) {
            const f = list[i];
            const fd = new FormData();
            fd.append('file', f);
            const r = await fetch('/api/message/upload', {
                method: 'POST',
                headers: {
                    'X-CSRFToken': CSRF
                },
                body: fd
            });
            const j = await r.json();
            if (j.ok && j.file) {
                attList.push({file_id: j.file.id, name: j.file.name, url: j.file.url});
            } else {
                alert((j.error || 'Ошибка загрузки') + ': ' + (f && f.name ? f.name : ''));
            }
        }
        updateAttView();
        toast('Файлы загружены');
    } catch (e) {
        alert('Ошибка загрузки файлов');
    }
}

async function sendMsg() {
    const body = (document.getElementById('body').value || '').trim();
    const internal = document.getElementById('internalNote').checked;
    if (!body && !attList.length) {
        return alert('Пустое сообщение');
    }
    try {
        const r = await fetch('/api/message/send', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify({
                thread_id: TID,
                body,
                internal_note: internal,
                attachments: attList
            })
        });
        const j = await r.json();
        if (j.ok) {
            location.reload();
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
}

async function sendExternal() {
    const body = (document.getElementById('body').value || '').trim();
    if (!body) {
        return alert('Пустое сообщение');
    }
    try {
        const r = await fetch('/api/channel/send', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify({
                thread_id: TID,
                text: body
            })
        });
        const j = await r.json();
        if (j.ok) {
            location.reload();
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
}

async function aiDraft() {
    try {
        const r = await fetch('/api/ai/draft_reply', {
            method: 'POST',
            headers: {
                'Content-Type': 'application/json',
                'X-CSRFToken': CSRF
            },
            body: JSON.stringify({
                thread_id: TID,
                tone: 'neutral'
            })
        });
        const j = await r.json();
        if (j.ok && j.variants && j.variants.length) {
            document.getElementById('body').value = j.variants[0];
            toast('Черновик сгенерирован');
        } else {
            alert(j.error || 'Ошибка AI');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
}

document.getElementById('thStatus')?.addEventListener('change', e => {
    e.preventDefault();
    thPatch({status: e.target.value});
});

document.getElementById('thPriority')?.addEventListener('change', e => {
    e.preventDefault();
    thPatch({priority: e.target.value});
});

document.getElementById('assigneeSelect')?.addEventListener('change', e => {
    e.preventDefault();
    thPatch({assignee_id: e.target.value || null});
});

document.getElementById('btnSaveTags')?.addEventListener('click', e => {
    e.preventDefault();
    saveTags();
});

document.getElementById('btnAIDraft')?.addEventListener('click', e => {
    e.preventDefault();
    aiDraft();
});

document.getElementById('btnSendInternal')?.addEventListener('click', e => {
    e.preventDefault();
    sendMsg();
});

document.getElementById('btnSendExternal')?.addEventListener('click', e => {
    e.preventDefault();
    document.getElementById('internalNote').checked = false;
    sendExternal();
});

document.getElementById('btnSnooze')?.addEventListener('click', e => {
    e.preventDefault();
    thPatch({status: 'snoozed'});
});

document.getElementById('fileInput')?.addEventListener('change', e => uploadFiles());

document.getElementById('dropZone')?.addEventListener('click', e => {
    e.preventDefault();
    document.getElementById('fileInput')?.click();
});

document.getElementById('dropZone')?.addEventListener('dragover', e => {
    e.preventDefault();
});

document.getElementById('dropZone')?.addEventListener('drop', e => {
    e.preventDefault();
    const files = e.dataTransfer.files;
    if (files && files.length) uploadFiles(files);
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== END OF STYLES PART 2/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 3/10 — TASKS PAGE & TASK VIEW TEMPLATES
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: TASKS LIST =====

TASKS_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px;">
    <h2 style="margin: 0;">Задачи</h2>
    <button class="button" onclick="document.getElementById('modalTaskCreate').classList.add('show')">+ Создать задачу</button>
</div>

<div class="card" style="margin-bottom: 12px;">
    <div style="display: flex; gap: 8px; flex-wrap: wrap;">
        <a class="button {% if current_filter=='open' %}primary{% else %}ghost{% endif %}" href="/tasks?f=open">Открытые</a>
        <a class="button {% if current_filter=='today' %}primary{% else %}ghost{% endif %}" href="/tasks?f=today">На сегодня</a>
        <a class="button {% if current_filter=='overdue' %}primary{% else %}ghost{% endif %}" href="/tasks?f=overdue">Просроченные</a>
        <a class="button {% if current_filter=='done' %}primary{% else %}ghost{% endif %}" href="/tasks?f=done">Завершённые</a>
        <details style="margin-left: auto;">
            <summary class="button ghost">Фильтры</summary>
            <div class="card" style="position: absolute; right: 0; z-index: 10; margin-top: 4px; min-width: 300px;">
                <form method="get" class="grid-filters">
                    <label>Статус
                        <select class="select" name="status">
                            <option value="">—</option>
                            {% for s in statuses %}
                            <option value="{{ s.name }}">{{ s.name }}</option>
                            {% endfor %}
                        </select>
                    </label>
                    <label>Исполнитель
                        <select class="select" name="assignee_id">
                            <option value="">—</option>
                            {% for a in agents %}
                            <option value="{{ a.id }}">{{ a.username }}</option>
                            {% endfor %}
                        </select>
                    </label>
                    <button class="button" type="submit">Применить</button>
                </form>
            </div>
        </details>
    </div>
</div>

<div class="card">
    <table class="table">
        <thead>
            <tr>
                <th><input type="checkbox" id="selectAll"></th>
                <th>ID</th>
                <th>Название</th>
                <th>Статус</th>
                <th>Приоритет</th>
                <th>Исполнитель</th>
                <th>Срок</th>
                <th>Компания</th>
                <th></th>
            </tr>
        </thead>
        <tbody>
            {% for t in tasks %}
            <tr data-id="{{ t.id }}" tabindex="0">
                <td><input type="checkbox" class="task-check" value="{{ t.id }}"></td>
                <td>#{{ t.id }}</td>
                <td><a href="/task/{{ t.id }}" style="color: var(--fg); text-decoration: none;">{{ t.title|e }}</a></td>
                <td>
                    <select class="select task-status" data-id="{{ t.id }}">
                        {% for s in statuses %}
                        <option value="{{ s.name }}" {% if t.status==s.name %}selected{% endif %}>{{ s.name }}</option>
                        {% endfor %}
                    </select>
                </td>
                <td>{{ t.priority or 'normal' }}</td>
                <td>
                    <select class="select task-assignee" data-id="{{ t.id }}">
                        <option value="">—</option>
                        {% for a in agents %}
                        <option value="{{ a.id }}" {% if t.assignee_id==a.id %}selected{% endif %}>{{ a.username }}</option>
                        {% endfor %}
                    </select>
                </td>
                <td>
                    {% if t.due_at %}
                    <span class="{% if t.due_at < now.isoformat(' ','seconds') and t.status not in ('done','cancelled') %}badge err{% endif %}">
                        {{ t.due_at }}
                    </span>
                    {% else %}—{% endif %}
                </td>
                <td>{{ t.company_name or '—' }}</td>
                <td>
                    <button class="iconbtn small task-toggle" data-id="{{ t.id }}">
                        {% if t.status == 'done' %}↩{% else %}✓{% endif %}
                    </button>
                </td>
            </tr>
            {% else %}
            <tr><td colspan="9" class="help">Задач нет</td></tr>
            {% endfor %}
        </tbody>
    </table>
    
    <div id="bulkActions" style="margin-top: 12px; display: none; gap: 8px;">
        <button class="button secondary" id="btnBulkAssign">Назначить</button>
        <button class="button secondary" id="btnBulkStatus">Сменить статус</button>
        <button class="button warn" id="btnBulkDelete">Удалить</button>
    </div>
</div>

<div class="modal-backdrop" id="modalTaskCreate">
    <div class="modal">
        <h3>Создать задачу</h3>
        <div class="form-fixed">
            <label>Название *
                <input class="input" id="newTaskTitle" required>
            </label>
            <label>Описание
                <textarea class="input" id="newTaskDesc" rows="3"></textarea>
            </label>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                <label>Срок
                    <input class="input" type="datetime-local" id="newTaskDue">
                </label>
                <label>Исполнитель
                    <select class="select" id="newTaskAssignee">
                        <option value="">—</option>
                        {% for a in agents %}
                        <option value="{{ a.id }}">{{ a.username }}</option>
                        {% endfor %}
                    </select>
                </label>
            </div>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnCreateTask">Создать</button>
                <button class="button ghost" onclick="document.getElementById('modalTaskCreate').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
// Task inline updates
document.addEventListener('change', async (e) => {
    const t = e.target;
    const id = t.getAttribute('data-id');
    if (!id) return;
    
    let patch = {};
    if (t.classList.contains('task-status')) patch.status = t.value;
    else if (t.classList.contains('task-assignee')) patch.assignee_id = t.value || null;
    else return;
    
    try {
        const r = await fetch('/api/task/update', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({id: parseInt(id), ...patch})
        });
        const j = await r.json();
        if (!j.ok) alert(j.error || 'Ошибка');
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Toggle task
document.addEventListener('click', async (e) => {
    if (!e.target.classList.contains('task-toggle')) return;
    const id = e.target.getAttribute('data-id');
    
    try {
        const r = await fetch('/api/task/toggle', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({id: parseInt(id)})
        });
        const j = await r.json();
        if (j.ok) location.reload();
        else alert(j.error || 'Ошибка');
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Select all
document.getElementById('selectAll')?.addEventListener('change', (e) => {
    document.querySelectorAll('.task-check').forEach(cb => cb.checked = e.target.checked);
    updateBulkActions();
});

document.querySelectorAll('.task-check').forEach(cb => {
    cb.addEventListener('change', updateBulkActions);
});

function updateBulkActions() {
    const checked = document.querySelectorAll('.task-check:checked');
    const bulk = document.getElementById('bulkActions');
    if (bulk) bulk.style.display = checked.length > 0 ? 'flex' : 'none';
}

// Create task
document.getElementById('btnCreateTask')?.addEventListener('click', async () => {
    const title = document.getElementById('newTaskTitle').value.trim();
    if (!title) return alert('Название обязательно');
    
    const data = {
        title,
        description: document.getElementById('newTaskDesc').value,
        due_at: document.getElementById('newTaskDue').value || null,
        assignee_id: document.getElementById('newTaskAssignee').value || null
    };
    
    try {
        const r = await fetch('/api/task/create', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) location.reload();
        else alert(j.error || 'Ошибка');
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Keyboard nav
document.addEventListener('keydown', e => {
    const tag = (e.target && e.target.tagName || '').toUpperCase();
    if (tag === 'INPUT' || tag === 'TEXTAREA' || tag === 'SELECT' || e.ctrlKey || e.metaKey) return;
    
    const rows = [...document.querySelectorAll('tbody tr[data-id]')];
    const active = document.activeElement?.closest('tr');
    const idx = rows.indexOf(active);
    
    if (e.key === 'ArrowDown') {
        e.preventDefault();
        (rows[Math.min(rows.length-1, idx+1)] || rows[0])?.focus();
    } else if (e.key === 'ArrowUp') {
        e.preventDefault();
        (rows[Math.max(0, idx-1)] || rows[0])?.focus();
    } else if (e.key === 'Enter' && active) {
        const id = active.getAttribute('data-id');
        if (id) location.href = '/task/' + id;
    }
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: TASK VIEW =====

TASK_VIEW_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; gap: 12px; margin-bottom: 16px;">
    <a href="/tasks" class="button ghost">← Назад</a>
    <h2 style="margin: 0;">Задача #{{ t.id }}</h2>
    <button class="button task-toggle-detail" data-id="{{ t.id }}">
        {% if t.status == 'done' %}↩ Открыть снова{% else %}✓ Завершить{% endif %}
    </button>
</div>

<div class="split">
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <h3 style="margin: 0 0 12px 0;" contenteditable="true" id="taskTitle">{{ t.title|e }}</h3>
            <div style="display: grid; gap: 12px;">
                <label class="help">Описание
                    <textarea class="input" id="taskDesc" rows="4">{{ t.description or '' }}</textarea>
                </label>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                    <label class="help">Статус
                        <select class="select" id="taskStatus">
                            {% for s in statuses %}
                            <option value="{{ s.name }}" {% if t.status==s.name %}selected{% endif %}>{{ s.name }}</option>
                            {% endfor %}
                        </select>
                    </label>
                    <label class="help">Приоритет
                        <select class="select" id="taskPriority">
                            <option value="low" {% if t.priority=='low' %}selected{% endif %}>Low</option>
                            <option value="normal" {% if t.priority=='normal' %}selected{% endif %}>Normal</option>
                            <option value="high" {% if t.priority=='high' %}selected{% endif %}>High</option>
                            <option value="urgent" {% if t.priority=='urgent' %}selected{% endif %}>Urgent</option>
                        </select>
                    </label>
                </div>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                    <label class="help">Срок
                        <input class="input" type="datetime-local" id="taskDue" value="{{ t.due_at[:16] if t.due_at else '' }}">
                    </label>
                    <label class="help">Исполнитель
                        <select class="select" id="taskAssignee">
                            <option value="">—</option>
                            {% for u in query_db('SELECT id,username FROM users WHERE org_id=? AND active=1',(user.org_id,)) %}
                            <option value="{{ u.id }}" {% if t.assignee_id==u.id %}selected{% endif %}>{{ u.username }}</option>
                            {% endfor %}
                        </select>
                    </label>
                </div>
                <button class="button" id="btnSaveTask">Сохранить изменения</button>
            </div>
        </div>
        
        <div class="card" style="margin-bottom: 12px;">
            <details>
                <summary style="font-weight: 600; margin-bottom: 8px;">Чек-лист</summary>
                <div id="checklistContainer">
                    {% for item in query_db('SELECT * FROM task_checklists WHERE task_id=? ORDER BY sort_order',(t.id,)) %}
                    <div class="checklist-item" style="display: flex; gap: 8px; margin: 4px 0;">
                        <input type="checkbox" {% if item.checked %}checked{% endif %}>
                        <input class="input" value="{{ item.item|e }}" style="flex: 1;">
                        <button class="iconbtn small">✕</button>
                    </div>
                    {% endfor %}
                </div>
                <button class="button ghost small" id="btnAddChecklistItem">+ Добавить пункт</button>
            </details>
        </div>
        
        <div class="card">
            <h4 style="margin: 0 0 12px 0;">Комментарии ({{ comments|length }})</h4>
            <div id="commentsContainer">
                {% for c in comments %}
                <div class="comment" style="border-bottom: 1px solid var(--border); padding: 12px 0;">
                    <div class="help">{{ c.username or 'Система' }} • {{ c.created_at }}</div>
                    <div style="margin-top: 4px; white-space: pre-wrap;">{{ c.body|e }}</div>
                    {% if c.attachments %}
                    <div class="help" style="margin-top: 4px;">
                        Вложения: {% for a in c.attachments %}<a href="{{ a.url }}" target="_blank">{{ a.name }}</a>{% if not loop.last %}, {% endif %}{% endfor %}
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
            </div>
            <div style="margin-top: 12px;">
                <textarea class="input" id="newComment" rows="3" placeholder="Добавить комментарий..."></textarea>
                <button class="button" id="btnAddComment" style="margin-top: 8px;">Отправить</button>
            </div>
        </div>
    </div>
    
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">Информация</h4>
            <div class="help">Создана: {{ t.created_at }}</div>
            <div class="help">Обновлена: {{ t.updated_at or '—' }}</div>
            {% if t.completed_at %}
            <div class="help">Завершена: {{ t.completed_at }}</div>
            {% endif %}
        </div>
        
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">Участники</h4>
            {% for p in participants %}
            <div class="help">{{ p.username }} ({{ p.role }})</div>
            {% else %}
            <div class="help">Нет участников</div>
            {% endfor %}
        </div>
        
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">История</h4>
            {% for a in activity[:5] %}
            <div class="help">{{ a.kind }} • {{ a.created_at }}</div>
            {% else %}
            <div class="help">Нет активности</div>
            {% endfor %}
        </div>
        
        <div class="card">
            <h4 style="margin: 0 0 8px 0;">Быстрые действия</h4>
            <div style="display: flex; flex-direction: column; gap: 4px;">
                <button class="button secondary small" id="btnDuplicate">Дублировать</button>
                <button class="button secondary small" id="btnDelegate">Делегировать</button>
                <button class="button warn small" id="btnDelete">Удалить</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
const TASK_ID = {{ t.id }};

// Save task
document.getElementById('btnSaveTask')?.addEventListener('click', async () => {
    const data = {
        id: TASK_ID,
        title: document.getElementById('taskTitle').textContent.trim(),
        description: document.getElementById('taskDesc').value,
        status: document.getElementById('taskStatus').value,
        priority: document.getElementById('taskPriority').value,
        due_at: document.getElementById('taskDue').value || null,
        assignee_id: document.getElementById('taskAssignee').value || null
    };
    
    try {
        const r = await fetch('/api/task/update', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) toast('Сохранено');
        else alert(j.error || 'Ошибка');
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Toggle task
document.querySelector('.task-toggle-detail')?.addEventListener('click', async () => {
    try {
        const r = await fetch('/api/task/toggle', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({id: TASK_ID})
        });
        const j = await r.json();
        if (j.ok) location.reload();
        else alert(j.error || 'Ошибка');
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Add comment
document.getElementById('btnAddComment')?.addEventListener('click', async () => {
    const body = document.getElementById('newComment').value.trim();
    if (!body) return;
    
    try {
        const r = await fetch('/api/task/comment', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({task_id: TASK_ID, body})
        });
        const j = await r.json();
        if (j.ok) location.reload();
        else alert(j.error || 'Ошибка');
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Delete task
document.getElementById('btnDelete')?.addEventListener('click', async () => {
    if (!confirm('Удалить задачу?')) return;
    // API endpoint not implemented yet
    alert('Функция в разработке');
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== END OF STYLES PART 3/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 4/10 — DEALS PAGE (KANBAN VIEW) TEMPLATE
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: DEALS (KANBAN) =====

DEALS_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px;">
    <h2 style="margin: 0;">Сделки</h2>
    <div style="display: flex; gap: 8px;">
        <button class="button ghost" id="btnToggleView">📋 Список</button>
        <button class="button" onclick="document.getElementById('modalDealCreate').classList.add('show')">+ Создать сделку</button>
    </div>
</div>

<div class="card" style="margin-bottom: 12px;">
    <div style="display: flex; gap: 8px; align-items: center;">
        <label class="help">Воронка:
            <select class="select" id="pipelineSelect" style="display: inline-block; width: auto;">
                <option value="default">Основная</option>
            </select>
        </label>
        <div class="help" style="margin-left: auto;">
            Всего в воронке: <strong id="totalDealsCount">{{ deals|length }}</strong> • 
            Сумма: <strong id="totalPipelineValue">0</strong> ₽
        </div>
    </div>
</div>

<div id="kanbanBoard" style="display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 12px; margin-bottom: 20px;">
    <!-- Kanban columns will be rendered by JS -->
</div>

<div id="listView" style="display: none;">
    <div class="card">
        <table class="table">
            <thead>
                <tr>
                    <th>ID</th>
                    <th>Название</th>
                    <th>Сумма</th>
                    <th>Стадия</th>
                    <th>Статус</th>
                    <th>Исполнитель</th>
                    <th>Компания</th>
                    <th>Создана</th>
                </tr>
            </thead>
            <tbody id="dealsTableBody">
                {% for d in deals %}
                <tr>
                    <td>#{{ d.id }}</td>
                    <td>{{ d.title|e }}</td>
                    <td>{{ d.amount|round(2) }} {{ d.currency }}</td>
                    <td>{{ d.stage }}</td>
                    <td><span class="badge {% if d.status=='won' %}ok{% elif d.status=='lost' %}err{% else %}warn{% endif %}">{{ d.status }}</span></td>
                    <td>{{ users_map.get(d.assignee_id|string, '—') }}</td>
                    <td>{{ d.company_name or '—' }}</td>
                    <td>{{ d.created_at }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</div>

<div class="modal-backdrop" id="modalDealCreate">
    <div class="modal">
        <h3>Создать сделку</h3>
        <div class="form-fixed">
            <label>Название *
                <input class="input" id="newDealTitle" required>
            </label>
            <div style="display: grid; grid-template-columns: 2fr 1fr; gap: 8px;">
                <label>Сумма
                    <input class="input" type="number" id="newDealAmount" value="0" min="0" step="0.01">
                </label>
                <label>Валюта
                    <select class="select" id="newDealCurrency">
                        <option value="RUB">RUB</option>
                        <option value="USD">USD</option>
                        <option value="EUR">EUR</option>
                    </select>
                </label>
            </div>
            <label>Стадия
                <select class="select" id="newDealStage">
                    <option value="new">Новая</option>
                    <option value="qualify">Квалификация</option>
                    <option value="proposal">Предложение</option>
                    <option value="negotiation">Переговоры</option>
                </select>
            </label>
            <label>Исполнитель
                <select class="select" id="newDealAssignee">
                    <option value="">—</option>
                    {% for u in users %}
                    <option value="{{ u.id }}">{{ u.username }}</option>
                    {% endfor %}
                </select>
            </label>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnCreateDeal">Создать</button>
                <button class="button ghost" onclick="document.getElementById('modalDealCreate').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<div class="modal-backdrop" id="modalDealEdit">
    <div class="modal">
        <h3>Редактировать сделку #<span id="editDealId"></span></h3>
        <div class="form-fixed">
            <label>Название
                <input class="input" id="editDealTitle">
            </label>
            <div style="display: grid; grid-template-columns: 2fr 1fr; gap: 8px;">
                <label>Сумма
                    <input class="input" type="number" id="editDealAmount">
                </label>
                <label>Валюта
                    <select class="select" id="editDealCurrency">
                        <option value="RUB">RUB</option>
                        <option value="USD">USD</option>
                        <option value="EUR">EUR</option>
                    </select>
                </label>
            </div>
            <label>Статус
                <select class="select" id="editDealStatus">
                    <option value="open">Open</option>
                    <option value="won">Won</option>
                    <option value="lost">Lost</option>
                </select>
            </label>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnSaveDeal">Сохранить</button>
                <button class="button ghost" onclick="document.getElementById('modalDealEdit').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<style>
.kanban-column {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 12px;
    min-height: 400px;
}

.kanban-column-header {
    font-weight: 600;
    margin-bottom: 12px;
    padding-bottom: 8px;
    border-bottom: 2px solid var(--border);
    display: flex;
    justify-content: space-between;
    align-items: center;
}

.kanban-card {
    background: var(--bg);
    border: 1px solid var(--border);
    border-radius: var(--radius-md);
    padding: 12px;
    margin-bottom: 8px;
    cursor: grab;
    transition: box-shadow 0.2s, transform 0.2s;
}

.kanban-card:hover {
    box-shadow: var(--shadow-md);
    transform: translateY(-2px);
}

.kanban-card.dragging {
    opacity: 0.5;
    cursor: grabbing;
}

.kanban-column.drag-over {
    background: var(--panel);
    border-color: var(--accent);
}
</style>

<script nonce="{{ csp_nonce }}">
const USERS_MAP = {{ users_map|tojson }};
let kanbanData = { columns: [], items: {} };
let currentDraggedDeal = null;

// Toggle view
document.getElementById('btnToggleView')?.addEventListener('click', () => {
    const kanban = document.getElementById('kanbanBoard');
    const list = document.getElementById('listView');
    if (kanban.style.display === 'none') {
        kanban.style.display = 'grid';
        list.style.display = 'none';
    } else {
        kanban.style.display = 'none';
        list.style.display = 'block';
    }
});

// Load kanban data
async function loadKanban() {
    try {
        const pipeline = document.getElementById('pipelineSelect').value;
        const r = await fetch(`/api/deals/kanban?pipeline=${pipeline}`);
        const j = await r.json();
        
        if (!j.ok) {
            alert(j.error || 'Ошибка загрузки');
            return;
        }
        
        kanbanData = j;
        renderKanban();
        updateStats();
    } catch (e) {
        console.error('Failed to load kanban:', e);
        alert('Ошибка загрузки канбана');
    }
}

function renderKanban() {
    const board = document.getElementById('kanbanBoard');
    if (!board) return;
    
    board.innerHTML = '';
    
    for (const column of (kanbanData.columns || [])) {
        const deals = kanbanData.items[column] || [];
        const columnEl = document.createElement('div');
        columnEl.className = 'kanban-column';
        columnEl.setAttribute('data-stage', column);
        
        columnEl.innerHTML = `
            <div class="kanban-column-header">
                <span>${esc(column)}</span>
                <span class="badge">${deals.length}</span>
            </div>
            <div class="kanban-cards" data-stage="${esc(column)}">
                ${deals.map(d => renderDealCard(d)).join('')}
            </div>
        `;
        
        board.appendChild(columnEl);
        
        // Enable drop
        const cardsContainer = columnEl.querySelector('.kanban-cards');
        cardsContainer.addEventListener('dragover', handleDragOver);
        cardsContainer.addEventListener('drop', handleDrop);
        cardsContainer.addEventListener('dragleave', handleDragLeave);
    }
    
    // Enable drag on cards
    document.querySelectorAll('.kanban-card').forEach(card => {
        card.addEventListener('dragstart', handleDragStart);
        card.addEventListener('dragend', handleDragEnd);
    });
}

function renderDealCard(deal) {
    const assignee = USERS_MAP[String(deal.assignee_id)] || '—';
    return `
        <div class="kanban-card" draggable="true" data-id="${deal.id}" ondblclick="openDealEdit(${deal.id})">
            <div style="font-weight: 600; margin-bottom: 4px;">${esc(deal.title || 'Без названия')}</div>
            <div class="help" style="margin-bottom: 4px;">${parseFloat(deal.amount || 0).toFixed(0)} ${esc(deal.currency || 'RUB')}</div>
            <div class="help" style="font-size: 12px;">👤 ${esc(assignee)}</div>
        </div>
    `;
}

function handleDragStart(e) {
    currentDraggedDeal = parseInt(e.target.getAttribute('data-id'));
    e.target.classList.add('dragging');
    e.dataTransfer.effectAllowed = 'move';
}

function handleDragEnd(e) {
    e.target.classList.remove('dragging');
    document.querySelectorAll('.kanban-column').forEach(col => {
        col.classList.remove('drag-over');
    });
}

function handleDragOver(e) {
    e.preventDefault();
    e.dataTransfer.dropEffect = 'move';
    e.currentTarget.closest('.kanban-column').classList.add('drag-over');
}

function handleDragLeave(e) {
    e.currentTarget.closest('.kanban-column').classList.remove('drag-over');
}

async function handleDrop(e) {
    e.preventDefault();
    const targetStage = e.currentTarget.getAttribute('data-stage');
    
    if (!currentDraggedDeal || !targetStage) return;
    
    try {
        const r = await fetch('/api/deals/kanban/update', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({
                id: currentDraggedDeal,
                stage: targetStage
            })
        });
        
        const j = await r.json();
        if (j.ok) {
            toast('Сделка перемещена');
            loadKanban(); // Reload
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
    
    currentDraggedDeal = null;
}

function updateStats() {
    let total = 0;
    let count = 0;
    
    for (const column in kanbanData.items) {
        const deals = kanbanData.items[column] || [];
        count += deals.length;
        deals.forEach(d => {
            total += parseFloat(d.amount || 0);
        });
    }
    
    document.getElementById('totalDealsCount').textContent = count;
    document.getElementById('totalPipelineValue').textContent = total.toFixed(0);
}

// Create deal
document.getElementById('btnCreateDeal')?.addEventListener('click', async () => {
    const title = document.getElementById('newDealTitle').value.trim();
    if (!title) return alert('Название обязательно');
    
    const data = {
        title,
        amount: parseFloat(document.getElementById('newDealAmount').value) || 0,
        currency: document.getElementById('newDealCurrency').value,
        stage: document.getElementById('newDealStage').value,
        assignee_id: document.getElementById('newDealAssignee').value || null
    };
    
    try {
        const r = await fetch('/api/deal/create', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) {
            document.getElementById('modalDealCreate').classList.remove('show');
            loadKanban();
            toast('Сделка создана');
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Open deal edit (double-click)
function openDealEdit(dealId) {
    // Find deal in kanbanData
    let deal = null;
    for (const column in kanbanData.items) {
        const found = kanbanData.items[column].find(d => d.id === dealId);
        if (found) {
            deal = found;
            break;
        }
    }
    
    if (!deal) return;
    
    document.getElementById('editDealId').textContent = deal.id;
    document.getElementById('editDealTitle').value = deal.title || '';
    document.getElementById('editDealAmount').value = deal.amount || 0;
    document.getElementById('editDealCurrency').value = deal.currency || 'RUB';
    document.getElementById('editDealStatus').value = deal.status || 'open';
    
    // Store current deal ID for save
    document.getElementById('btnSaveDeal').setAttribute('data-id', dealId);
    
    document.getElementById('modalDealEdit').classList.add('show');
}

// Save deal edit
document.getElementById('btnSaveDeal')?.addEventListener('click', async () => {
    const id = document.getElementById('btnSaveDeal').getAttribute('data-id');
    if (!id) return;
    
    const data = {
        id: parseInt(id),
        title: document.getElementById('editDealTitle').value,
        amount: parseFloat(document.getElementById('editDealAmount').value),
        currency: document.getElementById('editDealCurrency').value,
        status: document.getElementById('editDealStatus').value
    };
    
    try {
        const r = await fetch('/api/deal/update', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) {
            document.getElementById('modalDealEdit').classList.remove('show');
            loadKanban();
            toast('Сохранено');
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Initial load
loadKanban();
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== END OF STYLES PART 4/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 5/10 — CLIENTS LIST & CLIENT DETAIL PAGE TEMPLATES
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: CLIENTS LIST =====

CLIENTS_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px;">
    <h2 style="margin: 0;">Клиенты</h2>
    <button class="button" onclick="document.getElementById('modalClientCreate').classList.add('show')">+ Добавить клиента</button>
</div>

<div class="card" style="margin-bottom: 12px;">
    <form method="get" style="display: flex; gap: 8px;">
        <input class="input" name="q" placeholder="Поиск по названию, ИНН, телефону, email..." style="flex: 1;" value="{{ request.args.get('q', '') }}">
        <button class="button" type="submit">Найти</button>
        <a class="button ghost" href="/clients">Сбросить</a>
    </form>
</div>

<div class="card">
    <table class="table">
        <thead>
            <tr>
                <th>ID</th>
                <th>Название</th>
                <th>ИНН</th>
                <th>Телефон</th>
                <th>Email</th>
                <th>Сделки</th>
                <th>Адрес</th>
                <th></th>
            </tr>
        </thead>
        <tbody id="clientsTBody">
            {% for c in clients %}
            <tr data-id="{{ c.id }}" tabindex="0">
                <td>#{{ c.id }}</td>
                <td><a href="/client/{{ c.id }}" style="color: var(--fg); text-decoration: none; font-weight: 500;">{{ c.name|e }}</a></td>
                <td>{{ c.inn or '—' }}</td>
                <td>
                    {% if c.phone %}
                    <a href="tel:{{ c.phone }}" class="iconbtn phone small">📞 {{ c.phone }}</a>
                    {% else %}—{% endif %}
                </td>
                <td>
                    {% if c.email %}
                    <a href="mailto:{{ c.email }}" style="color: var(--fg);">{{ c.email }}</a>
                    {% else %}—{% endif %}
                </td>
                <td>
                    {% if c.deals > 0 %}
                    <span class="badge ok">{{ c.deals }}</span>
                    {% else %}
                    <span class="help">0</span>
                    {% endif %}
                </td>
                <td class="help">{{ (c.address or '')[:30] }}{% if c.address and c.address|length > 30 %}...{% endif %}</td>
                <td>
                    <button class="iconbtn small" onclick="editClient({{ c.id }}, '{{ c.name|e }}', '{{ c.inn or '' }}', '{{ c.phone or '' }}', '{{ c.email or '' }}', '{{ c.address|e or '' }}')">✏️</button>
                </td>
            </tr>
            {% else %}
            <tr><td colspan="8" class="help">Клиентов нет</td></tr>
            {% endfor %}
        </tbody>
    </table>
    
    <div style="margin-top: 12px; display: flex; justify-content: space-between; align-items: center;">
        <div class="help">Всего: {{ clients|length }}</div>
        <button class="button ghost" id="loadMoreClients">Загрузить больше</button>
    </div>
</div>

<div class="modal-backdrop" id="modalClientCreate">
    <div class="modal">
        <h3>Добавить клиента</h3>
        <div class="form-fixed">
            <label>Название компании *
                <input class="input" id="newClientName" required>
            </label>
            <label>ИНН
                <input class="input" id="newClientINN" pattern="[0-9]{10}|[0-9]{12}" placeholder="10 или 12 цифр">
            </label>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                <label>Телефон
                    <input class="input" id="newClientPhone" type="tel" placeholder="+7...">
                </label>
                <label>Email
                    <input class="input" id="newClientEmail" type="email">
                </label>
            </div>
            <label>Адрес
                <textarea class="input" id="newClientAddress" rows="2"></textarea>
            </label>
            <label>Заметки
                <textarea class="input" id="newClientNotes" rows="3"></textarea>
            </label>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnCreateClient">Создать</button>
                <button class="button ghost" onclick="document.getElementById('modalClientCreate').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<div class="modal-backdrop" id="modalClientEdit">
    <div class="modal">
        <h3>Редактировать клиента #<span id="editClientId"></span></h3>
        <div class="form-fixed">
            <label>Название компании
                <input class="input" id="editClientName">
            </label>
            <label>ИНН
                <input class="input" id="editClientINN" pattern="[0-9]{10}|[0-9]{12}">
            </label>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                <label>Телефон
                    <input class="input" id="editClientPhone">
                </label>
                <label>Email
                    <input class="input" id="editClientEmail">
                </label>
            </div>
            <label>Адрес
                <textarea class="input" id="editClientAddress" rows="2"></textarea>
            </label>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnSaveClient">Сохранить</button>
                <button class="button ghost" onclick="document.getElementById('modalClientEdit').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
let clientsPage = 1;

// Create client via form
document.getElementById('btnCreateClient')?.addEventListener('click', async () => {
    const name = document.getElementById('newClientName').value.trim();
    if (!name) return alert('Название обязательно');
    
    const data = {
        name,
        inn: document.getElementById('newClientINN').value.trim(),
        phone: document.getElementById('newClientPhone').value.trim(),
        email: document.getElementById('newClientEmail').value.trim(),
        address: document.getElementById('newClientAddress').value.trim(),
        notes: document.getElementById('newClientNotes').value.trim()
    };
    
    try {
        // Using form POST to /clients (from CORE routes)
        const form = document.createElement('form');
        form.method = 'POST';
        form.action = '/clients';
        
        const csrfInput = document.createElement('input');
        csrfInput.type = 'hidden';
        csrfInput.name = 'csrf_token';
        csrfInput.value = CSRF;
        form.appendChild(csrfInput);
        
        for (const [key, value] of Object.entries(data)) {
            const input = document.createElement('input');
            input.type = 'hidden';
            input.name = key;
            input.value = value;
            form.appendChild(input);
        }
        
        document.body.appendChild(form);
        form.submit();
    } catch (e) {
        alert('Ошибка создания');
    }
});

// Edit client (open modal)
function editClient(id, name, inn, phone, email, address) {
    document.getElementById('editClientId').textContent = id;
    document.getElementById('editClientName').value = name;
    document.getElementById('editClientINN').value = inn;
    document.getElementById('editClientPhone').value = phone;
    document.getElementById('editClientEmail').value = email;
    document.getElementById('editClientAddress').value = address;
    
    document.getElementById('btnSaveClient').setAttribute('data-id', id);
    document.getElementById('modalClientEdit').classList.add('show');
}

// Save client edit
document.getElementById('btnSaveClient')?.addEventListener('click', async () => {
    const id = document.getElementById('btnSaveClient').getAttribute('data-id');
    if (!id) return;
    
    const data = {
        name: document.getElementById('editClientName').value,
        inn: document.getElementById('editClientINN').value,
        phone: document.getElementById('editClientPhone').value,
        email: document.getElementById('editClientEmail').value,
        address: document.getElementById('editClientAddress').value
    };
    
    try {
        const r = await fetch(`/api/clients/${id}`, {
            method: 'PATCH',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) {
            document.getElementById('modalClientEdit').classList.remove('show');
            location.reload();
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Load more
document.getElementById('loadMoreClients')?.addEventListener('click', async () => {
    clientsPage++;
    const q = new URLSearchParams(window.location.search).get('q') || '';
    
    try {
        const r = await fetch(`/api/clients/list?page=${clientsPage}&q=${encodeURIComponent(q)}`);
        const j = await r.json();
        if (!j.ok) return alert(j.error || 'Ошибка');
        
        const tbody = document.getElementById('clientsTBody');
        for (const c of (j.items || [])) {
            const tr = document.createElement('tr');
            tr.setAttribute('data-id', c.id);
            tr.innerHTML = `
                <td>#${c.id}</td>
                <td><a href="/client/${c.id}" style="color: var(--fg); text-decoration: none; font-weight: 500;">${esc(c.name)}</a></td>
                <td>${esc(c.inn || '—')}</td>
                <td>${c.phone ? '<a href="tel:' + esc(c.phone) + '" class="iconbtn phone small">📞 ' + esc(c.phone) + '</a>' : '—'}</td>
                <td>${c.email ? '<a href="mailto:' + esc(c.email) + '">' + esc(c.email) + '</a>' : '—'}</td>
                <td>${c.deals > 0 ? '<span class="badge ok">' + c.deals + '</span>' : '<span class="help">0</span>'}</td>
                <td class="help">${esc((c.address || '').slice(0, 30))}</td>
                <td><button class="iconbtn small" onclick="editClient(${c.id}, '${esc(c.name)}', '${esc(c.inn||'')}', '${esc(c.phone||'')}', '${esc(c.email||'')}', '${esc(c.address||'')}')">✏️</button></td>
            `;
            tbody.appendChild(tr);
        }
        
        toast('Загружено больше');
    } catch (e) {
        alert('Ошибка загрузки');
    }
});

// Keyboard navigation
document.addEventListener('keydown', e => {
    const tag = (e.target && e.target.tagName || '').toUpperCase();
    if (tag === 'INPUT' || tag === 'TEXTAREA' || tag === 'SELECT' || e.ctrlKey || e.metaKey) return;
    
    const rows = [...document.querySelectorAll('#clientsTBody tr[data-id]')];
    const active = document.activeElement?.closest('tr');
    const idx = rows.indexOf(active);
    
    if (e.key === 'ArrowDown') {
        e.preventDefault();
        (rows[Math.min(rows.length-1, idx+1)] || rows[0])?.focus();
    } else if (e.key === 'ArrowUp') {
        e.preventDefault();
        (rows[Math.max(0, idx-1)] || rows[0])?.focus();
    } else if (e.key === 'Enter' && active) {
        const id = active.getAttribute('data-id');
        if (id) location.href = '/client/' + id;
    }
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: CLIENT PAGE (DETAIL) =====

CLIENT_PAGE_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; gap: 12px; margin-bottom: 16px;">
    <a href="/clients" class="button ghost">← Назад</a>
    <h2 style="margin: 0;">{{ c.name|e }}</h2>
</div>

<div class="split">
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <h3 style="margin: 0 0 12px 0;">Информация</h3>
            <div style="display: grid; gap: 12px;">
                <div>
                    <label class="help">Название компании</label>
                    <input class="input" id="companyName" value="{{ c.name|e }}">
                </div>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                    <div>
                        <label class="help">ИНН</label>
                        <input class="input" id="companyINN" value="{{ c.inn or '' }}">
                    </div>
                    <div>
                        <label class="help">Телефон</label>
                        <input class="input" id="companyPhone" value="{{ c.phone or '' }}">
                    </div>
                </div>
                <div>
                    <label class="help">Email</label>
                    <input class="input" type="email" id="companyEmail" value="{{ c.email or '' }}">
                </div>
                <div>
                    <label class="help">Адрес</label>
                    <textarea class="input" id="companyAddress" rows="2">{{ c.address or '' }}</textarea>
                </div>
                <div>
                    <label class="help">Заметки</label>
                    <textarea class="input" id="companyNotes" rows="4">{{ c.notes or '' }}</textarea>
                </div>
                <button class="button" id="btnSaveCompany">Сохранить изменения</button>
            </div>
        </div>
        
        <div class="card">
            <h3 style="margin: 0 0 12px 0;">История звонков ({{ calls|length }})</h3>
            <div style="max-height: 400px; overflow: auto;">
                {% for call in calls %}
                <div style="border-bottom: 1px solid var(--border); padding: 8px 0;">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div>
                            <span class="badge {% if call.direction=='inbound' %}ok{% else %}warn{% endif %}">
                                {{ '📞 Входящий' if call.direction=='inbound' else '📱 Исходящий' }}
                            </span>
                            <span class="help">{{ call.from_e164 }} → {{ call.to_e164 }}</span>
                        </div>
                        <div class="help">{{ call.started_at }}</div>
                    </div>
                    <div class="help" style="margin-top: 4px;">
                        Длительность: {{ call.duration_sec }}с • 
                        Статус: {{ call.status or '—' }}
                        {% if call.recording_url %}
                        • <a href="/cti/recording/{{ call.id }}" target="_blank">Запись</a>
                        {% endif %}
                    </div>
                </div>
                {% else %}
                <div class="help">Звонков пока нет</div>
                {% endfor %}
            </div>
        </div>
    </div>
    
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <h3 style="margin: 0 0 8px 0;">Сведения</h3>
            <div class="help">ID: {{ c.id }}</div>
            <div class="help">Создана: {{ c.created_at }}</div>
            <div class="help">Обновлена: {{ c.updated_at or '—' }}</div>
            {% if c.score %}
            <div style="margin-top: 8px;">
                <span class="help">Lead Score:</span>
                <span class="badge ok">{{ c.score }}</span>
            </div>
            {% endif %}
        </div>
        
        <div class="card" style="margin-bottom: 12px;">
            <h3 style="margin: 0 0 8px 0;">Быстрые действия</h3>
            <div style="display: flex; flex-direction: column; gap: 6px;">
                {% if c.phone %}
                <button class="button" onclick="clickToCall('{{ c.phone }}')">📞 Позвонить</button>
                {% endif %}
                {% if c.email %}
                <a class="button secondary" href="mailto:{{ c.email }}">✉️ Написать email</a>
                {% endif %}
                <button class="button secondary" onclick="location.href='/tasks?company_id={{ c.id }}'">📋 Задачи</button>
                <button class="button secondary" onclick="location.href='/deals?company_id={{ c.id }}'">💼 Сделки</button>
                <button class="button secondary" onclick="createDealForClient({{ c.id }}, '{{ c.name|e }}')">+ Создать сделку</button>
            </div>
        </div>
        
        <div class="card">
            <h3 style="margin: 0 0 8px 0;">Контакты ({{ query_db('SELECT COUNT(*) as cnt FROM contacts WHERE company_id=?', (c.id,), one=True).cnt }})</h3>
            <div id="contactsList" style="max-height: 300px; overflow: auto;">
                {% for contact in query_db('SELECT * FROM contacts WHERE company_id=? LIMIT 10', (c.id,)) %}
                <div style="padding: 6px 0; border-bottom: 1px solid var(--border);">
                    <div style="font-weight: 500;">{{ contact.name|e }}</div>
                    <div class="help">{{ contact.position or '—' }}</div>
                    {% if contact.phone %}
                    <div class="help">📞 {{ contact.phone }}</div>
                    {% endif %}
                </div>
                {% else %}
                <div class="help">Контактов нет</div>
                {% endfor %}
            </div>
            <button class="button ghost small" style="margin-top: 8px; width: 100%;" onclick="alert('Функция в разработке')">+ Добавить контакт</button>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
const CLIENT_ID = {{ c.id }};

// Save company info
document.getElementById('btnSaveCompany')?.addEventListener('click', async () => {
    const data = {
        name: document.getElementById('companyName').value,
        inn: document.getElementById('companyINN').value,
        phone: document.getElementById('companyPhone').value,
        email: document.getElementById('companyEmail').value,
        address: document.getElementById('companyAddress').value,
        notes: document.getElementById('companyNotes').value
    };
    
    try {
        const r = await fetch(`/api/clients/${CLIENT_ID}`, {
            method: 'PATCH',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify(data)
        });
        const j = await r.json();
        if (j.ok) {
            toast('Сохранено');
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
});

// Click to call
async function clickToCall(phone) {
    if (!confirm(`Позвонить на ${phone}?`)) return;
    
    try {
        const r = await fetch('/api/cti/click_to_call', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({to: phone})
        });
        const j = await r.json();
        if (j.ok) {
            toast(j.message || 'Звонок инициирован');
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка звонка');
    }
}

// Create deal for client
function createDealForClient(clientId, clientName) {
    // Redirect to deals page with prefilled company
    location.href = `/deals?new=1&company_id=${clientId}&company_name=${encodeURIComponent(clientName)}`;
}
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== END OF STYLES PART 5/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 6/10 — CALLS PAGE TEMPLATE
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: CALLS =====

CALLS_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px;">
    <h2 style="margin: 0;">Звонки</h2>
    <div style="display: flex; gap: 8px;">
        <button class="button ghost" onclick="document.getElementById('modalCallStats').classList.add('show')">📊 Статистика</button>
        <button class="button" onclick="promptClickToCall()">📞 Позвонить</button>
    </div>
</div>

<div class="card" style="margin-bottom: 12px;">
    <details>
        <summary class="button ghost">Фильтры</summary>
        <form method="get" class="grid-filters" style="margin-top: 12px;">
            <label>Период от
                <input class="input" type="date" name="date_from" value="{{ request.args.get('date_from', '') }}">
            </label>
            <label>Период до
                <input class="input" type="date" name="date_to" value="{{ request.args.get('date_to', '') }}">
            </label>
            <label>Направление
                <select class="select" name="direction">
                    <option value="">— все —</option>
                    <option value="inbound" {% if request.args.get('direction')=='inbound' %}selected{% endif %}>Входящие</option>
                    <option value="outbound" {% if request.args.get('direction')=='outbound' %}selected{% endif %}>Исходящие</option>
                </select>
            </label>
            <label>Агент
                <select class="select" name="agent_id">
                    <option value="">— все —</option>
                    {% for a in agents_rows %}
                    <option value="{{ a.id }}" {% if request.args.get('agent_id')|string==a.id|string %}selected{% endif %}>{{ a.username }}</option>
                    {% endfor %}
                </select>
            </label>
            <label>Мои звонки
                <input type="checkbox" name="mine" value="1" {% if request.args.get('mine')=='1' %}checked{% endif %}>
            </label>
            <div style="grid-column: 1/-1; display: flex; gap: 8px; justify-content: flex-end; margin-top: 4px;">
                <button class="button" type="submit">Применить</button>
                <a class="button ghost" href="/calls">Сбросить</a>
            </div>
        </form>
    </details>
</div>

<div class="split" style="margin-bottom: 12px;">
    <div class="card">
        <h4 style="margin: 0 0 8px 0;">Сегодня</h4>
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 12px;">
            <div>
                <div style="font-size: 24px; font-weight: 700; color: var(--ok);" id="todayInbound">0</div>
                <div class="help">Входящих</div>
            </div>
            <div>
                <div style="font-size: 24px; font-weight: 700; color: var(--warn);" id="todayOutbound">0</div>
                <div class="help">Исходящих</div>
            </div>
        </div>
    </div>
    
    <div class="card">
        <h4 style="margin: 0 0 8px 0;">Среднее время</h4>
        <div style="font-size: 24px; font-weight: 700; color: var(--accent);" id="avgDuration">0</div>
        <div class="help">секунд на звонок</div>
    </div>
</div>

<div class="card">
    <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px;">
        <h3 style="margin: 0;">История звонков</h3>
        <button class="button ghost small" id="btnExportCalls">Экспорт CSV</button>
    </div>
    
    <table class="table">
        <thead>
            <tr>
                <th>ID</th>
                <th>Направление</th>
                <th>От кого</th>
                <th>Кому</th>
                <th>Агент</th>
                <th>Статус</th>
                <th>Длительность</th>
                <th>Начало</th>
                <th>Компания</th>
                <th></th>
            </tr>
        </thead>
        <tbody id="callsTBody">
            <!-- Calls will be loaded via JS -->
        </tbody>
    </table>
    
    <button class="button ghost" id="loadMoreCalls" style="margin-top: 12px; width: 100%;">Загрузить больше</button>
</div>

<div class="modal-backdrop" id="modalCallStats">
    <div class="modal">
        <h3>Статистика звонков</h3>
        <div id="statsContainer">
            <canvas id="callsChart" style="max-height: 300px;"></canvas>
            <div style="margin-top: 16px; display: grid; gap: 8px;">
                <div class="help">Общая статистика за период загружается...</div>
            </div>
        </div>
        <button class="button" onclick="document.getElementById('modalCallStats').classList.remove('show')" style="margin-top: 16px;">Закрыть</button>
    </div>
</div>

<div class="modal-backdrop" id="modalCallDetail">
    <div class="modal">
        <h3>Детали звонка #<span id="callDetailId"></span></h3>
        <div id="callDetailContent" style="display: grid; gap: 8px;">
            <!-- Populated by JS -->
        </div>
        <div id="callRecordingPlayer" style="margin-top: 12px; display: none;">
            <h4>Запись звонка</h4>
            <audio controls style="width: 100%;"></audio>
        </div>
        <div style="margin-top: 16px; display: flex; gap: 8px;">
            <button class="button secondary" id="btnCallToTask">Создать задачу</button>
            <button class="button ghost" onclick="document.getElementById('modalCallDetail').classList.remove('show')">Закрыть</button>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
let callsPage = 1;
let currentCallId = null;

// Load calls
async function loadCalls(page = 1) {
    try {
        const params = new URLSearchParams(window.location.search);
        params.set('page', page);
        params.set('per_page', 50);
        
        const r = await fetch('/api/calls/list?' + params.toString());
        const j = await r.json();
        
        if (!j.ok) {
            alert(j.error || 'Ошибка загрузки');
            return;
        }
        
        const tbody = document.getElementById('callsTBody');
        
        if (page === 1) {
            tbody.innerHTML = '';
        }
        
        for (const call of (j.items || [])) {
            tbody.insertAdjacentHTML('beforeend', renderCallRow(call));
        }
        
        if ((j.items || []).length < (j.per_page || 50)) {
            document.getElementById('loadMoreCalls').style.display = 'none';
        }
        
        updateStats(j.items || []);
    } catch (e) {
        console.error('Failed to load calls:', e);
        alert('Ошибка загрузки звонков');
    }
}

function renderCallRow(call) {
    const directionBadge = call.direction === 'inbound' 
        ? '<span class="badge ok">📞 Входящий</span>' 
        : '<span class="badge warn">📱 Исходящий</span>';
    
    const statusColor = call.status === 'completed' ? 'ok' : (call.status === 'missed' ? 'err' : 'warn');
    
    return `
        <tr data-id="${call.id}" onclick="showCallDetail(${call.id})" style="cursor: pointer;">
            <td>#${call.id}</td>
            <td>${directionBadge}</td>
            <td>${esc(call.from_e164 || '—')}</td>
            <td>${esc(call.to_e164 || '—')}</td>
            <td>${esc(call.agent_name || '—')}</td>
            <td><span class="badge ${statusColor}">${esc(call.status || 'unknown')}</span></td>
            <td>${call.duration_sec || 0}с</td>
            <td>${esc((call.started_at || '').slice(0, 16))}</td>
            <td>${esc(call.company_name || '—')}</td>
            <td>
                ${call.recording_url ? '<a href="/cti/recording/' + call.id + '" target="_blank" onclick="event.stopPropagation()" class="iconbtn small">🎧</a>' : ''}
            </td>
        </tr>
    `;
}

function updateStats(calls) {
    const today = new Date().toISOString().slice(0, 10);
    let inbound = 0, outbound = 0, totalDuration = 0, count = 0;
    
    for (const call of calls) {
        const callDate = (call.started_at || '').slice(0, 10);
        
        if (callDate === today) {
            if (call.direction === 'inbound') inbound++;
            else if (call.direction === 'outbound') outbound++;
        }
        
        totalDuration += (call.duration_sec || 0);
        count++;
    }
    
    document.getElementById('todayInbound').textContent = inbound;
    document.getElementById('todayOutbound').textContent = outbound;
    document.getElementById('avgDuration').textContent = count > 0 ? Math.round(totalDuration / count) : 0;
}

async function showCallDetail(callId) {
    currentCallId = callId;
    
    // Find call in current data (or fetch)
    const row = document.querySelector(`tr[data-id="${callId}"]`);
    if (!row) return;
    
    // Extract data from row (simplified - in production, fetch from API)
    const cells = row.querySelectorAll('td');
    
    const content = `
        <div><strong>ID:</strong> ${callId}</div>
        <div><strong>Направление:</strong> ${cells[1].textContent}</div>
        <div><strong>От кого:</strong> ${cells[2].textContent}</div>
        <div><strong>Кому:</strong> ${cells[3].textContent}</div>
        <div><strong>Агент:</strong> ${cells[4].textContent}</div>
        <div><strong>Статус:</strong> ${cells[5].textContent}</div>
        <div><strong>Длительность:</strong> ${cells[6].textContent}</div>
        <div><strong>Начало:</strong> ${cells[7].textContent}</div>
        <div><strong>Компания:</strong> ${cells[8].textContent}</div>
    `;
    
    document.getElementById('callDetailId').textContent = callId;
    document.getElementById('callDetailContent').innerHTML = content;
    
    // Check for recording
    const recordingLink = row.querySelector('a[href*="/cti/recording/"]');
    const recordingPlayer = document.getElementById('callRecordingPlayer');
    
    if (recordingLink) {
        const audio = recordingPlayer.querySelector('audio');
        audio.src = recordingLink.href;
        recordingPlayer.style.display = 'block';
    } else {
        recordingPlayer.style.display = 'none';
    }
    
    document.getElementById('modalCallDetail').classList.add('show');
}

// Create task from call
document.getElementById('btnCallToTask')?.addEventListener('click', async () => {
    if (!currentCallId) return;
    
    const title = prompt('Название задачи:', 'Обработать звонок #' + currentCallId);
    if (!title) return;
    
    try {
        const r = await fetch('/api/call/to_task', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({call_id: currentCallId, title})
        });
        const j = await r.json();
        
        if (j.ok) {
            toast('Задача создана');
            document.getElementById('modalCallDetail').classList.remove('show');
            if (confirm('Перейти к задаче?')) {
                location.href = '/task/' + j.task_id;
            }
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка создания задачи');
    }
});

// Load more
document.getElementById('loadMoreCalls')?.addEventListener('click', () => {
    callsPage++;
    loadCalls(callsPage);
});

// Export CSV
document.getElementById('btnExportCalls')?.addEventListener('click', () => {
    const params = new URLSearchParams(window.location.search);
    window.open('/api/export/calls/csv?' + params.toString(), '_blank');
});

// Click to call prompt
function promptClickToCall() {
    const phone = prompt('Введите номер телефона:');
    if (!phone) return;
    
    clickToCall(phone);
}

async function clickToCall(phone) {
    try {
        const r = await fetch('/api/cti/click_to_call', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({to: phone})
        });
        const j = await r.json();
        
        if (j.ok) {
            toast(j.message || 'Звонок инициирован');
            
            // Reload after 2 seconds to show new call
            setTimeout(() => {
                loadCalls(1);
            }, 2000);
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка звонка');
    }
}

// Initial load
loadCalls(1);

// Auto-refresh every 30 seconds
setInterval(() => {
    if (document.visibilityState === 'visible') {
        loadCalls(1);
    }
}, 30000);
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== END OF STYLES PART 6/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 7/10 — MEETINGS & CHAT TEMPLATES
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: MEETING (LIST & DETAIL) =====

MEETING_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
{% if meeting %}
{# Meeting detail view #}
<div style="display: flex; align-items: center; gap: 12px; margin-bottom: 16px;">
    <a href="/meetings" class="button ghost">← Назад</a>
    <h2 style="margin: 0;">{{ meeting.title or 'Встреча' }}</h2>
</div>

<div class="split">
    <div>
        <div class="card" style="margin-bottom: 12px; text-align: center; background: linear-gradient(135deg, var(--accent) 0%, var(--ok) 100%); color: #000;">
            <h3 style="margin: 0 0 12px 0;">🎥 Комната: {{ meeting.room }}</h3>
            <div style="margin-bottom: 16px;">
                {% if meeting.start_at %}
                <div class="help" style="color: rgba(0,0,0,0.7);">Начало: {{ meeting.start_at }}</div>
                {% endif %}
                {% if meeting.end_at %}
                <div class="help" style="color: rgba(0,0,0,0.7);">Окончание: {{ meeting.end_at }}</div>
                {% endif %}
            </div>
            <button class="button" onclick="joinMeeting('{{ meeting.room }}')" style="background: #fff; color: #000; font-size: 16px; padding: 12px 24px;">
                🚀 Присоединиться к встрече
            </button>
        </div>
        
        <div class="card">
            <h4 style="margin: 0 0 12px 0;">Ссылка для приглашения</h4>
            <div style="display: flex; gap: 8px;">
                <input class="input" id="meetingLink" value="{{ request.url_root }}meeting/{{ meeting.id }}" readonly>
                <button class="button secondary" onclick="copyMeetingLink()">📋 Копировать</button>
            </div>
        </div>
        
        <div id="jitsiContainer" style="margin-top: 12px; height: 600px; border-radius: var(--radius-lg); overflow: hidden; display: none;">
            <!-- Jitsi iframe will be injected here -->
        </div>
    </div>
    
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">Информация</h4>
            <div class="help">ID: {{ meeting.id }}</div>
            <div class="help">Создал: {{ query_db('SELECT username FROM users WHERE id=?', (meeting.created_by,), one=True).username if meeting.created_by else '—' }}</div>
            <div class="help">Создана: {{ meeting.created_at }}</div>
            {% if meeting.recording_started_at %}
            <div class="help">Запись начата: {{ meeting.recording_started_at }}</div>
            {% endif %}
        </div>
        
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">Участники</h4>
            {% set participants = meeting.participants_json|fromjson if meeting.participants_json else [] %}
            {% if participants %}
                {% for p in participants %}
                <div class="help">• {{ p }}</div>
                {% endfor %}
            {% else %}
                <div class="help">Участники не указаны</div>
            {% endif %}
        </div>
        
        <div class="card">
            <h4 style="margin: 0 0 8px 0;">Управление</h4>
            <div style="display: flex; flex-direction: column; gap: 6px;">
                <button class="button secondary small" id="btnStartRecording">🔴 Начать запись</button>
                <button class="button secondary small" id="btnStopRecording" style="display: none;">⏹ Остановить запись</button>
                <button class="button warn small" onclick="deleteMeeting({{ meeting.id }})">🗑 Удалить встречу</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
const MEETING_ID = {{ meeting.id }};
const JITSI_ROOM = '{{ meeting.room }}';

function joinMeeting(room) {
    const container = document.getElementById('jitsiContainer');
    container.style.display = 'block';
    container.scrollIntoView({behavior: 'smooth'});
    
    const domain = '{{ jitsi_base }}'.replace('https://', '').replace('http://', '');
    const options = {
        roomName: room,
        width: '100%',
        height: '100%',
        parentNode: container,
        userInfo: {
            displayName: '{{ user.username }}'
        }
    };
    
    // Load Jitsi API if not loaded
    if (!window.JitsiMeetExternalAPI) {
        const script = document.createElement('script');
        script.src = '{{ jitsi_base }}/external_api.js';
        script.onload = () => {
            new JitsiMeetExternalAPI(domain, options);
        };
        document.head.appendChild(script);
    } else {
        new JitsiMeetExternalAPI(domain, options);
    }
}

function copyMeetingLink() {
    const input = document.getElementById('meetingLink');
    input.select();
    document.execCommand('copy');
    toast('Ссылка скопирована');
}

async function deleteMeeting(id) {
    if (!confirm('Удалить встречу?')) return;
    
    try {
        const r = await fetch(`/api/meetings/${id}`, {
            method: 'DELETE',
            headers: {'X-CSRFToken': CSRF}
        });
        const j = await r.json();
        
        if (j.ok) {
            location.href = '/meetings';
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка удаления');
    }
}

document.getElementById('btnStartRecording')?.addEventListener('click', async () => {
    try {
        const r = await fetch(`/api/meeting/${MEETING_ID}/start_recording`, {
            method: 'POST',
            headers: {'X-CSRFToken': CSRF}
        });
        const j = await r.json();
        
        if (j.ok) {
            toast('Запись начата');
            document.getElementById('btnStartRecording').style.display = 'none';
            document.getElementById('btnStopRecording').style.display = 'block';
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка');
    }
});

document.getElementById('btnStopRecording')?.addEventListener('click', async () => {
    try {
        const r = await fetch(`/api/meeting/${MEETING_ID}/stop_recording`, {
            method: 'POST',
            headers: {'X-CSRFToken': CSRF}
        });
        const j = await r.json();
        
        if (j.ok) {
            toast('Запись остановлена: ' + j.summary);
            document.getElementById('btnStartRecording').style.display = 'block';
            document.getElementById('btnStopRecording').style.display = 'none';
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка');
    }
});
</script>

{% else %}
{# Meetings list view #}
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px;">
    <h2 style="margin: 0;">Встречи</h2>
    <button class="button" onclick="document.getElementById('modalMeetingCreate').classList.add('show')">+ Запланировать встречу</button>
</div>

<div class="card">
    <div id="meetingsContainer">
        <div id="meetingsList"></div>
    </div>
</div>

<div class="modal-backdrop" id="modalMeetingCreate">
    <div class="modal">
        <h3>Запланировать встречу</h3>
        <div class="form-fixed">
            <label>Название
                <input class="input" id="newMeetingTitle" placeholder="Планерка, Демо для клиента...">
            </label>
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 8px;">
                <label>Начало
                    <input class="input" type="datetime-local" id="newMeetingStart">
                </label>
                <label>Окончание
                    <input class="input" type="datetime-local" id="newMeetingEnd">
                </label>
            </div>
            <label>Участники (по одному на строку)
                <textarea class="input" id="newMeetingParticipants" rows="3" placeholder="user1&#10;user2&#10;client@example.com"></textarea>
            </label>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnCreateMeeting">Создать</button>
                <button class="button ghost" onclick="document.getElementById('modalMeetingCreate').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
async function loadMeetings() {
    try {
        const r = await fetch('/api/meetings');
        const j = await r.json();
        
        if (!j.ok) {
            alert(j.error || 'Ошибка загрузки');
            return;
        }
        
        const container = document.getElementById('meetingsList');
        container.innerHTML = '';
        
        const meetings = j.items || [];
        
        if (meetings.length === 0) {
            container.innerHTML = '<div class="help" style="padding: 40px; text-align: center;">Встреч пока нет</div>';
            return;
        }
        
        for (const m of meetings) {
            const card = document.createElement('div');
            card.className = 'card';
            card.style.marginBottom = '12px';
            card.style.cursor = 'pointer';
            card.onclick = () => location.href = '/meeting/' + m.id;
            
            card.innerHTML = `
                <div style="display: flex; justify-content: space-between; align-items: start;">
                    <div>
                        <h4 style="margin: 0 0 4px 0;">🎥 ${esc(m.title || 'Встреча #' + m.id)}</h4>
                        <div class="help">Комната: ${esc(m.room)}</div>
                        ${m.start_at ? '<div class="help">Начало: ' + esc(m.start_at) + '</div>' : ''}
                    </div>
                    <button class="button small" onclick="event.stopPropagation(); location.href='/meeting/${m.id}'">Открыть</button>
                </div>
            `;
            
            container.appendChild(card);
        }
    } catch (e) {
        alert('Ошибка загрузки встреч');
    }
}

document.getElementById('btnCreateMeeting')?.addEventListener('click', async () => {
    const title = document.getElementById('newMeetingTitle').value.trim();
    const start = document.getElementById('newMeetingStart').value;
    const end = document.getElementById('newMeetingEnd').value;
    const participantsText = document.getElementById('newMeetingParticipants').value;
    
    const participants = participantsText.split('\\n').map(p => p.trim()).filter(Boolean);
    
    try {
        const r = await fetch('/api/meetings/schedule', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({
                title,
                start_at: start || null,
                end_at: end || null,
                participants
            })
        });
        const j = await r.json();
        
        if (j.ok) {
            document.getElementById('modalMeetingCreate').classList.remove('show');
            location.href = '/meeting/' + j.id;
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка создания');
    }
});

loadMeetings();
</script>
{% endif %}
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: CHAT =====

CHAT_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<h2 style="margin: 0 0 16px 0;">Чат</h2>

<div style="display: grid; grid-template-columns: 280px 1fr; gap: 12px; height: calc(100vh - 200px);">
    <div class="card" style="overflow: auto;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px;">
            <h4 style="margin: 0;">Каналы</h4>
            <button class="iconbtn small" onclick="document.getElementById('modalChannelCreate').classList.add('show')">+</button>
        </div>
        
        <div id="channelsList">
            {% for ch in channels %}
            <a href="/chat/{{ ch.id }}" class="channel-item {% if current and current.id == ch.id %}active{% endif %}" style="display: block; padding: 8px; border-radius: 6px; text-decoration: none; color: var(--fg); margin-bottom: 4px;">
                <div style="font-weight: 500;"># {{ ch.title or 'Канал ' ~ ch.id }}</div>
                <div class="help" style="font-size: 12px;">{{ ch.type }}</div>
            </a>
            {% else %}
            <div class="help">Каналов нет</div>
            {% endfor %}
        </div>
    </div>
    
    <div class="card" style="display: flex; flex-direction: column;">
        {% if current %}
        <div style="border-bottom: 1px solid var(--border); padding-bottom: 12px; margin-bottom: 12px;">
            <h3 style="margin: 0;"># {{ current.title or 'Канал ' ~ current.id }}</h3>
            <div class="help">{{ current.type }} • ID: {{ current.id }}</div>
        </div>
        
        <div id="messagesContainer" style="flex: 1; overflow: auto; margin-bottom: 12px; display: flex; flex-direction: column-reverse;">
            {% for m in messages|reverse %}
            <div class="chat-message" style="padding: 8px; margin-bottom: 4px; {% if m.user_id == user.id %}background: var(--panel);{% endif %} border-radius: 8px;">
                <div style="display: flex; gap: 8px;">
                    <div style="font-weight: 600; color: var(--accent);">{{ m.username or 'Система' }}</div>
                    <div class="help" style="font-size: 12px;">{{ m.created_at }}</div>
                </div>
                <div style="margin-top: 4px; white-space: pre-wrap;">{{ m.body|e }}</div>
            </div>
            {% else %}
            <div class="help" style="text-align: center; padding: 40px;">Сообщений пока нет</div>
            {% endfor %}
        </div>
        
        <div>
            <div id="dropZoneChat" style="border: 2px dashed var(--border); padding: 8px; text-align: center; cursor: pointer; border-radius: 6px; margin-bottom: 8px; display: none;">
                📎 Перетащите файлы
            </div>
            <div style="display: flex; gap: 8px;">
                <textarea class="input" id="chatMessageInput" rows="2" placeholder="Написать сообщение... (Ctrl+Enter для отправки)" style="flex: 1;"></textarea>
                <div style="display: flex; flex-direction: column; gap: 4px;">
                    <button class="iconbtn" onclick="document.getElementById('chatFileInput').click()">📎</button>
                    <button class="button" id="btnSendChatMessage">Отправить</button>
                </div>
            </div>
            <input type="file" id="chatFileInput" multiple style="display: none;">
        </div>
        {% else %}
        <div class="help" style="text-align: center; padding: 40px;">
            Выберите канал или создайте новый
        </div>
        {% endif %}
    </div>
</div>

<div class="modal-backdrop" id="modalChannelCreate">
    <div class="modal">
        <h3>Создать канал</h3>
        <div class="form-fixed">
            <label>Название
                <input class="input" id="newChannelTitle" placeholder="Общий чат, Продажи...">
            </label>
            <label>Тип
                <select class="select" id="newChannelType">
                    <option value="public">Публичный (все в организации)</option>
                    <option value="private">Приватный (только участники)</option>
                </select>
            </label>
            <label>Участники (выберите пользователей)
                <select class="select" id="newChannelMembers" multiple size="5">
                    {% for u in query_db('SELECT id, username FROM users WHERE org_id=? AND active=1', (user.org_id,)) %}
                    <option value="{{ u.id }}">{{ u.username }}</option>
                    {% endfor %}
                </select>
            </label>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnCreateChannel">Создать</button>
                <button class="button ghost" onclick="document.getElementById('modalChannelCreate').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<style>
.channel-item:hover {
    background: var(--panel);
}

.channel-item.active {
    background: var(--accent);
    color: #000;
}

.chat-message:hover {
    background: var(--surface);
}
</style>

<script nonce="{{ csp_nonce }}">
{% if current %}
const CHANNEL_ID = {{ current.id }};

// Send message
async function sendMessage() {
    const input = document.getElementById('chatMessageInput');
    const body = input.value.trim();
    
    if (!body) return;
    
    try {
        const r = await fetch('/api/chat/send', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({
                channel_id: CHANNEL_ID,
                body
            })
        });
        const j = await r.json();
        
        if (j.ok) {
            input.value = '';
            location.reload(); // В production использовать WebSocket
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка отправки');
    }
}

document.getElementById('btnSendChatMessage')?.addEventListener('click', sendMessage);

document.getElementById('chatMessageInput')?.addEventListener('keydown', (e) => {
    if (e.ctrlKey && e.key === 'Enter') {
        e.preventDefault();
        sendMessage();
    }
});

// File upload
document.getElementById('chatFileInput')?.addEventListener('change', async (e) => {
    const files = e.target.files;
    if (!files || !files.length) return;
    
    for (let i = 0; i < files.length; i++) {
        const file = files[i];
        const fd = new FormData();
        fd.append('file', file);
        fd.append('channel_id', CHANNEL_ID);
        
        try {
            const r = await fetch('/api/chat/upload', {
                method: 'POST',
                headers: {'X-CSRFToken': CSRF},
                body: fd
            });
            const j = await r.json();
            
            if (j.ok) {
                toast('Файл отправлен: ' + file.name);
            } else {
                alert('Ошибка загрузки: ' + (j.error || file.name));
            }
        } catch (e) {
            alert('Ошибка загрузки файла');
        }
    }
    
    setTimeout(() => location.reload(), 1000);
});

// Auto-scroll to bottom
const container = document.getElementById('messagesContainer');
if (container) {
    container.scrollTop = container.scrollHeight;
}
{% endif %}

// Create channel
document.getElementById('btnCreateChannel')?.addEventListener('click', async () => {
    const title = document.getElementById('newChannelTitle').value.trim();
    const type = document.getElementById('newChannelType').value;
    const membersSelect = document.getElementById('newChannelMembers');
    const members = Array.from(membersSelect.selectedOptions).map(o => parseInt(o.value));
    
    try {
        const r = await fetch('/api/chat/create', {
            method: 'POST',
            headers: {'Content-Type': 'application/json', 'X-CSRFToken': CSRF},
            body: JSON.stringify({
                title,
                type,
                members
            })
        });
        const j = await r.json();
        
        if (j.ok) {
            location.href = '/chat/' + j.id;
        } else {
            alert(j.error || 'Ошибка');
        }
    } catch (e) {
        alert('Ошибка создания канала');
    }
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== END OF STYLES PART 7/10 =====
# ═════════════════════════════════════════════════════════════════════════════
# STYLES PART 8/10 — DOCUMENTS, WAREHOUSE, IMPORT TEMPLATES
# ═════════════════════════════════════════════════════════════════════════════

# ===== TEMPLATE: DOCUMENTS =====

DOCUMENTS_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<h2 style="margin: 0 0 16px 0;">Документы</h2>

<div class="split" style="margin-bottom: 12px;">
    <div class="card">
        <h4 style="margin: 0 0 12px 0;">Шаблоны документов</h4>
        <div id="templatesList">
            {% for t in templates %}
            <div style="padding: 8px; border-bottom: 1px solid var(--border);">
                <div style="font-weight: 500;">{{ t.name|e }}</div>
                <div class="help">Тип: {{ t.type }} • Создан: {{ t.created_at[:10] }}</div>
            </div>
            {% else %}
            <div class="help">Шаблонов нет</div>
            {% endfor %}
        </div>
        <button class="button ghost small" onclick="document.getElementById('modalTemplateCreate').classList.add('show')" style="margin-top: 8px; width: 100%;">+ Добавить шаблон</button>
    </div>
    
    <div class="card">
        <h4 style="margin: 0 0 12px 0;">Создать документ</h4>
        <form method="post" action="/documents">
            <input type="hidden" name="csrf_token" value="{{ session.get('csrf_token', '') }}">
            <div class="form-fixed">
                <label>Шаблон
                    <select class="select" name="template_id" required>
                        <option value="">— выберите —</option>
                        {% for t in templates %}
                        <option value="{{ t.id }}">{{ t.name }}</option>
                        {% endfor %}
                    </select>
                </label>
                <label>Компания
                    <select class="select" name="company_id" required>
                        <option value="">— выберите —</option>
                        {% for c in companies[:100] %}
                        <option value="{{ c.id }}">{{ c.name }}</option>
                        {% endfor %}
                    </select>
                </label>
                <button class="button" type="submit">Создать документ</button>
            </div>
        </form>
    </div>
</div>

<div class="card">
    <h3 style="margin: 0 0 12px 0;">Недавние документы</h3>
    <table class="table">
        <thead>
            <tr>
                <th>ID</th>
                <th>Название</th>
                <th>Тип</th>
                <th>Компания</th>
                <th>Создал</th>
                <th>Дата</th>
                <th></th>
            </tr>
        </thead>
        <tbody>
            {% for d in docs %}
            <tr>
                <td>#{{ d.id }}</td>
                <td><a href="/document/{{ d.id }}" style="color: var(--fg); text-decoration: none;">{{ d.title|e }}</a></td>
                <td><span class="badge">{{ d.doc_type or '—' }}</span></td>
                <td>{{ d.company_name or '—' }}</td>
                <td class="help">{{ query_db('SELECT username FROM users WHERE id=?', (d.user_id,), one=True).username if d.user_id else '—' }}</td>
                <td class="help">{{ d.created_at[:10] }}</td>
                <td>
                    <a href="/document/{{ d.id }}" class="iconbtn small">👁️</a>
                    <button class="iconbtn small" onclick="downloadDocument({{ d.id }})">⬇️</button>
                </td>
            </tr>
            {% else %}
            <tr><td colspan="7" class="help">Документов нет</td></tr>
            {% endfor %}
        </tbody>
    </table>
</div>

<div class="modal-backdrop" id="modalTemplateCreate">
    <div class="modal">
        <h3>Добавить шаблон документа</h3>
        <form method="post" action="/documents/template">
            <input type="hidden" name="csrf_token" value="{{ session.get('csrf_token', '') }}">
            <div class="form-fixed">
                <label>Название шаблона
                    <input class="input" name="name" required placeholder="Договор поставки">
                </label>
                <label>Тип
                    <select class="select" name="type">
                        <option value="contract">Договор</option>
                        <option value="invoice">Счет</option>
                        <option value="offer">Коммерческое предложение</option>
                        <option value="act">Акт</option>
                        <option value="other">Другое</option>
                    </select>
                </label>
                <label>Ключ (уникальный идентификатор)
                    <input class="input" name="tkey" placeholder="contract_supply">
                </label>
                <label>Шаблон (HTML, переменные: {{ '{{' }} org.name {{ '}}' }}, {{ '{{' }} company.name {{ '}}' }}, {{ '{{' }} user.username {{ '}}' }})
                    <textarea class="input" name="body_template" rows="12" required placeholder="<h1>Договор поставки</h1>&#10;<p>Поставщик: {{ '{{' }} org.name {{ '}}' }}</p>&#10;<p>Покупатель: {{ '{{' }} company.name {{ '}}' }}</p>"></textarea>
                </label>
                <div style="display: flex; gap: 8px; margin-top: 12px;">
                    <button class="button" type="submit">Сохранить</button>
                    <button class="button ghost" type="button" onclick="document.getElementById('modalTemplateCreate').classList.remove('show')">Отмена</button>
                </div>
            </div>
        </form>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
function downloadDocument(docId) {
    // В production: генерировать PDF через API
    alert('Функция экспорта в PDF в разработке. ID документа: ' + docId);
}
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: DOCUMENT VIEW =====

DOCUMENT_VIEW_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; gap: 12px; margin-bottom: 16px;">
    <a href="/documents" class="button ghost">← Назад</a>
    <h2 style="margin: 0;">{{ d.title|e }}</h2>
</div>

<div class="split">
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <div style="display: flex; justify-content: space-between; margin-bottom: 12px;">
                <h4 style="margin: 0;">Предпросмотр</h4>
                <div style="display: flex; gap: 4px;">
                    <button class="iconbtn small" onclick="editDocument()">✏️ Редактировать</button>
                    <button class="iconbtn small" onclick="printDocument()">🖨️ Печать</button>
                    <button class="iconbtn small" onclick="downloadPDF()">📄 PDF</button>
                </div>
            </div>
            <div id="documentPreview" style="background: #fff; color: #000; padding: 40px; border: 1px solid var(--border); border-radius: 8px; min-height: 600px; font-family: 'Times New Roman', serif;">
                {{ d.content_html|safe }}
            </div>
        </div>
    </div>
    
    <div>
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">Информация</h4>
            <div class="help">ID: {{ d.id }}</div>
            <div class="help">Тип: {{ d.doc_type or '—' }}</div>
            <div class="help">Создан: {{ d.created_at }}</div>
            {% if d.company_id %}
            <div style="margin-top: 8px;">
                <a href="/client/{{ d.company_id }}" class="button ghost small" style="width: 100%;">Компания: {{ query_db('SELECT name FROM companies WHERE id=?', (d.company_id,), one=True).name }}</a>
            </div>
            {% endif %}
        </div>
        
        <div class="card" style="margin-bottom: 12px;">
            <h4 style="margin: 0 0 8px 0;">Действия</h4>
            <div style="display: flex; flex-direction: column; gap: 6px;">
                <button class="button secondary small" onclick="sendDocumentEmail()">✉️ Отправить по email</button>
                <button class="button secondary small" onclick="duplicateDocument()">📋 Дублировать</button>
                <button class="button secondary small" onclick="createApprovalLink()">🔗 Создать ссылку согласования</button>
                <button class="button warn small" onclick="deleteDocument()">🗑 Удалить</button>
            </div>
        </div>
        
        <div class="card">
            <h4 style="margin: 0 0 8px 0;">История</h4>
            <div class="help">Версия: 1.0</div>
            <div class="help">Последнее изменение: {{ d.created_at[:16] }}</div>
        </div>
    </div>
</div>

<div class="modal-backdrop" id="modalDocumentEdit">
    <div class="modal" style="max-width: 800px;">
        <h3>Редактировать документ</h3>
        <div class="form-fixed">
            <label>HTML содержимое
                <textarea class="input" id="docContentEdit" rows="20" style="font-family: var(--font-mono); font-size: 13px;">{{ d.content_html|e }}</textarea>
            </label>
            <div class="help">⚠️ Будьте осторожны с HTML-кодом. Недопустимый код может нарушить отображение.</div>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" onclick="saveDocumentEdit()">Сохранить</button>
                <button class="button ghost" onclick="document.getElementById('modalDocumentEdit').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
const DOC_ID = {{ d.id }};

function editDocument() {
    document.getElementById('modalDocumentEdit').classList.add('show');
}

function saveDocumentEdit() {
    const content = document.getElementById('docContentEdit').value;
    alert('Функция сохранения в разработке. Используйте API для обновления документа.');
    // TODO: Implement /api/document/<id>/update
}

function printDocument() {
    const preview = document.getElementById('documentPreview').innerHTML;
    const win = window.open('', '_blank');
    win.document.write('<html><head><title>{{ d.title|e }}</title></head><body>' + preview + '</body></html>');
    win.document.close();
    win.print();
}

function downloadPDF() {
    alert('Экспорт в PDF в разработке. Используйте печать → Сохранить как PDF.');
}

function sendDocumentEmail() {
    const email = prompt('Email получателя:');
    if (!email) return;
    alert('Отправка email в разработке. Адрес: ' + email);
}

function duplicateDocument() {
    if (!confirm('Создать копию документа?')) return;
    alert('Дублирование в разработке.');
}

function createApprovalLink() {
    // Generate approval token
    const token = btoa('deal:' + DOC_ID + ':secret'); // Simplified
    const link = location.origin + '/approve/' + token;
    
    prompt('Ссылка для согласования (скопируйте):', link);
}

function deleteDocument() {
    if (!confirm('Удалить документ?')) return;
    alert('Удаление в разработке.');
}
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: WAREHOUSE =====

WAREHOUSE_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 16px;">
    <h2 style="margin: 0;">Склад</h2>
    <button class="button" onclick="document.getElementById('modalProductCreate').classList.add('show')">+ Добавить товар</button>
</div>

<div class="card" style="margin-bottom: 12px;">
    <div style="display: flex; gap: 8px;">
        <input class="input" id="searchProduct" placeholder="Поиск по SKU, названию..." style="flex: 1;">
        <button class="button ghost" onclick="filterLowStock()">⚠️ Низкие остатки</button>
        <button class="button ghost" onclick="exportInventory()">📥 Экспорт</button>
    </div>
</div>

<div class="card">
    <table class="table">
        <thead>
            <tr>
                <th>SKU</th>
                <th>Название</th>
                <th>Описание</th>
                <th>Цена</th>
                <th>Валюта</th>
                <th>Остаток</th>
                <th>Статус</th>
                <th></th>
            </tr>
        </thead>
        <tbody id="productsTBody">
            {% for p in products %}
            <tr data-id="{{ p.id }}">
                <td><code>{{ p.sku or '—' }}</code></td>
                <td style="font-weight: 500;">{{ p.name|e }}</td>
                <td class="help">{{ (p.description or '')[:50] }}{% if p.description and p.description|length > 50 %}...{% endif %}</td>
                <td>{{ p.price|round(2) }}</td>
                <td>{{ p.currency }}</td>
                <td>
                    <input class="input" type="number" value="{{ p.qty }}" data-id="{{ p.id }}" style="width: 80px;" onchange="updateStock({{ p.id }}, this.value)">
                </td>
                <td>
                    {% if p.qty <= 0 %}
                    <span class="badge err">Нет в наличии</span>
                    {% elif p.qty < 10 %}
                    <span class="badge warn">Мало</span>
                    {% else %}
                    <span class="badge ok">В наличии</span>
                    {% endif %}
                </td>
                <td>
                    <button class="iconbtn small" onclick="editProduct({{ p.id }})">✏️</button>
                    <button class="iconbtn small" onclick="deleteProduct({{ p.id }})">🗑</button>
                </td>
            </tr>
            {% else %}
            <tr><td colspan="8" class="help">Товаров нет</td></tr>
            {% endfor %}
        </tbody>
    </table>
</div>

<div class="modal-backdrop" id="modalProductCreate">
    <div class="modal">
        <h3>Добавить товар</h3>
        <div class="form-fixed">
            <label>SKU (артикул)
                <input class="input" id="newProductSKU" placeholder="PROD-001">
            </label>
            <label>Название *
                <input class="input" id="newProductName" required>
            </label>
            <label>Описание
                <textarea class="input" id="newProductDesc" rows="3"></textarea>
            </label>
            <div style="display: grid; grid-template-columns: 2fr 1fr 1fr; gap: 8px;">
                <label>Цена
                    <input class="input" type="number" id="newProductPrice" value="0" min="0" step="0.01">
                </label>
                <label>Валюта
                    <select class="select" id="newProductCurrency">
                        <option value="RUB">RUB</option>
                        <option value="USD">USD</option>
                        <option value="EUR">EUR</option>
                    </select>
                </label>
                <label>Остаток
                    <input class="input" type="number" id="newProductQty" value="0" min="0" step="0.01">
                </label>
            </div>
            <div style="display: flex; gap: 8px; margin-top: 12px;">
                <button class="button" id="btnCreateProduct">Создать</button>
                <button class="button ghost" onclick="document.getElementById('modalProductCreate').classList.remove('show')">Отмена</button>
            </div>
        </div>
    </div>
</div>

<script nonce="{{ csp_nonce }}">
// Search products (client-side simple filter)
document.getElementById('searchProduct')?.addEventListener('input', (e) => {
    const q = e.target.value.toLowerCase();
    document.querySelectorAll('#productsTBody tr').forEach(row => {
        const text = row.textContent.toLowerCase();
        row.style.display = text.includes(q) ? '' : 'none';
    });
});

function filterLowStock() {
    document.querySelectorAll('#productsTBody tr').forEach(row => {
        const qtyInput = row.querySelector('input[type=number]');
        if (!qtyInput) return;
        const qty = parseFloat(qtyInput.value) || 0;
        row.style.display = (qty < 10) ? '' : 'none';
    });
}

function exportInventory() {
    alert('Экспорт в CSV в разработке.');
}

async function updateStock(productId, newQty) {
    const form = new FormData();
    form.append('csrf_token', CSRF);
    form.append('product_id', productId);
    form.append('qty', newQty);
    
    try {
        const r = await fetch('/warehouse/stock/set', {
            method: 'POST',
            body: form
        });
        
        if (r.ok) {
            toast('Остаток обновлён');
        } else {
            alert('Ошибка обновления');
        }
    } catch (e) {
        alert('Ошибка сети');
    }
}

function editProduct(id) {
    alert('Редактирование товара #' + id + ' в разработке.');
}

function deleteProduct(id) {
    if (!confirm('Удалить товар?')) return;
    alert('Удаление в разработке.');
}

// Create product (simplified - use form submit in production)
document.getElementById('btnCreateProduct')?.addEventListener('click', () => {
    const name = document.getElementById('newProductName').value.trim();
    if (!name) return alert('Название обязательно');
    
    alert('Создание товара через API в разработке. Используйте backend form.');
});
</script>
''').replace('{% endblock %}', '{% endblock %}')


# ===== TEMPLATE: IMPORT =====

IMPORT_TMPL = LAYOUT_TMPL.replace('{% block content %}', '''{% block content %}
<h2 style="margin: 0 0 16px 0;">Импорт данных</h2>

<div class="split">
    <div class="card">
        <h3 style="margin: 0 0 12px 0;">Загрузить CSV файл</h3>
        <form method="post" action="/import" enctype="multipart/form-data">
            <input type="hidden" name="csrf_token" value="{{ session.get('csrf_token', '') }}">
            <div class="form-fixed">
                <label>Тип данных
                    <select class="select" name="mode" required>
                        <option value="">— выберите —</option>
                        <option value="companies">Компании</option>
                        <option value="contacts">Контакты</option>
                        <option value="tasks">Задачи</option>
                        <option value="deals">Сделки</option>
                        <option value="products">Товары</option>
                    </select>
                </label>
                
                <label>CSV файл (разделитель: точка с запятой)
                    <input type="file" name="csvfile" accept=".csv,.txt" required class="input">
                </label>
                
                <div class="help" style="margin-bottom: 8px;">
                    ℹ️ Формат файла:<br>
                    <strong>Компании:</strong> name;inn;phone;email;address;notes<br>
                    <strong>Контакты:</strong> name;company_id;position;phone;email;notes<br>
                    <strong>Задачи:</strong> title;description;assignee_id;due_at;company_id<br>
                    <strong>Сделки:</strong> title;amount;currency;stage;assignee_id;company_id<br>
                    <strong>Товары:</strong> sku;name;description;price;currency;qty
                </div>
                
                <details style="margin-bottom: 12px;">
                    <summary class="help" style="cursor: pointer;">Дополнительные опции</summary>
                    <div style="margin-top: 8px;">
                        <label><input type="checkbox" name="skip_duplicates" value="1"> Пропускать дубликаты (по имени/SKU)</label><br>
                        <label><input type="checkbox" name="update_existing" value="1"> Обновлять существующие записи</label><br>
                        <label><input type="checkbox" name="validate_strict" value="1" checked> Строгая валидация (остановка при ошибке)</label>
                    </div>
                </details>
                
                <button class="button" type="submit">📥 Импортировать</button>
            </div>
        </form>
    </div>
    
    <div class="card">
        <h3 style="margin: 0 0 12px 0;">Инструкции</h3>
        
        <details open>
            <summary class="help" style="cursor: pointer; font-weight: 600;">1. Подготовьте файл</summary>
            <div style="margin-top: 8px;" class="help">
                • Используйте Excel/LibreOffice: "Сохранить как → CSV (разделитель точка с запятой)"<br>
                • Кодировка: UTF-8<br>
                • Первая строка — заголовки (headers)<br>
                • Пустые поля оставляйте пустыми (не пишите "—" или "null")
            </div>
        </details>
        
        <details style="margin-top: 12px;">
            <summary class="help" style="cursor: pointer; font-weight: 600;">2. Примеры CSV</summary>
            <div style="margin-top: 8px;">
                <h5>Компании (companies):</h5>
                <pre style="background: var(--panel); padding: 8px; border-radius: 4px; overflow: auto; font-size: 12px;">name;inn;phone;email;address;notes
ООО Ромашка;1234567890;+74951234567;info@romashka.ru;Москва, ул. Ленина 1;Крупный клиент
ИП Иванов;;+79161234567;ivanov@mail.ru;;
</pre>
                
                <h5 style="margin-top: 12px;">Задачи (tasks):</h5>
                <pre style="background: var(--panel); padding: 8px; border-radius: 4px; overflow: auto; font-size: 12px;">title;description;assignee_id;due_at;company_id
Позвонить клиенту;Уточнить детали заказа;2;2024-12-31 15:00;5
